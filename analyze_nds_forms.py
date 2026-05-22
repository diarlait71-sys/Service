"""
Анализ НДС-деклараций (Форма 300) за 2025-2026
Вытягивает ключевые показатели и строит сводку по всем кварталам
"""

import openpyxl
import os
import pandas as pd
from collections import defaultdict
from datetime import datetime

def extract_company_name(ws):
    """Пытается найти название ТОО в заголовке декларации"""
    for r in range(1, min(20, ws.max_row+1)):
        for c in range(1, min(10, ws.max_column+1)):
            cell = ws.cell(r, c)
            if cell.value and isinstance(cell.value, str):
                val = str(cell.value).lower()
                if 'тоо' in val or 'досcar' in val:
                    return str(cell.value).strip()
    return "Unknown"

def find_key_indicators(ws):
    """
    Поиск ключевых показателей в форме 300:
    - Строка 059: Оборот реализации (без НДС)
    - Строка 082: Начисленный НДС
    - Строка 076: Входной НДС к зачету
    - Строка 085: Сумма НДС к возмещению
    """
    indicators = {}
    
    for r in range(1, ws.max_row + 1):
        # ищем по коду строки в первой колонке
        cell_code = ws.cell(r, 1)
        if not cell_code.value:
            continue
        
        code = str(cell_code.value).strip()
        
        # ключевые коды из формы 300
        key_rows = {
            '059': 'turnover_no_vat',
            '082': 'vat_charged',
            '076': 'vat_input_deductible',
            '085': 'vat_to_refund',
            '300.00.003': 'vat_total',
            '300.00.005': 'vat_calculated',
        }
        
        if code in key_rows:
            # значение обычно в колонке 16 (P) или 27, 30
            values = []
            for c in [16, 27, 30]:
                val = ws.cell(r, c)
                if val.value and isinstance(val.value, (int, float)):
                    values.append(val.value)
            if values:
                indicators[key_rows[code]] = max(values)  # берем максимум если несколько
    
    return indicators

def process_nds_files():
    """Обрабатывает все файлы НДС и собирает данные"""
    
    nds_path = r'c:\Users\D.Muldabaev\Desktop\Приложения для сервиса\Ндс'
    files = sorted([f for f in os.listdir(nds_path) if f.endswith('.xlsx')])
    
    results = []
    
    for filename in files:
        print(f"Обработка: {filename}")
        
        # парсим имя файла: Форма 300 1 кв 2025 1 лист.xlsx
        parts = filename.replace('Форма 300 ', '').replace(' лист.xlsx', '').split()
        if len(parts) < 3:
            continue
        
        kv_month = parts[0] + ' ' + parts[1]  # "1 кв" или "2 кв"
        year = int(parts[2])
        sheet_num = parts[3] if len(parts) > 3 else '1'
        
        filepath = os.path.join(nds_path, filename)
        
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            # собираем данные
            company = extract_company_name(ws)
            indicators = find_key_indicators(ws)
            
            results.append({
                'Файл': filename,
                'Период': f"{year}-{kv_month}",
                'Год': year,
                'Квартал': kv_month,
                'Лист': sheet_num,
                'ТОО': company,
                'Оборот_без_НДС': indicators.get('turnover_no_vat', 0),
                'НДС_начисленный': indicators.get('vat_charged', 0),
                'НДС_входной': indicators.get('vat_input_deductible', 0),
                'НДС_к_возмещению': indicators.get('vat_to_refund', 0),
            })
            
        except Exception as e:
            print(f"  Ошибка: {e}")
            results.append({
                'Файл': filename,
                'Период': f"{year}-{kv_month}",
                'ТОО': 'ERROR',
                'Ошибка': str(e)
            })
    
    return results

def main():
    print("=" * 80)
    print("АНАЛИЗ НДС-ДЕКЛАРАЦИЙ (ФОРМА 300)")
    print("=" * 80)
    
    results = process_nds_files()
    
    # создаем DataFrame
    df = pd.DataFrame(results)
    
    print(f"\nОбработано файлов: {len(df)}")
    print(f"Периодов: {df['Период'].nunique()}")
    print(f"Листов на период: {df.groupby('Период').size().describe().T}\n")
    
    # сводка по периодам
    print("\n=== СВОДКА ПО ПЕРИОДАМ ===")
    summary = df.groupby('Период')[['Оборот_без_НДС', 'НДС_начисленный', 'НДС_входной', 'НДС_к_возмещению']].sum()
    print(summary.to_string())
    
    # сохраняем в Excel
    output_path = r'c:\Users\D.Muldabaev\Desktop\Приложения для сервиса\Ндс_сводка_анализ.xlsx'
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Детали', index=False)
        summary.to_excel(writer, sheet_name='Сводка по периодам')
        
        # статистика
        stats = {
            'Показатель': ['Всего файлов', 'Периодов', 'Листов на период', 'Сумма оборотов'],
            'Значение': [len(df), df['Период'].nunique(), df.groupby('Период').size().mean(), 
                        df['Оборот_без_НДС'].sum()]
        }
        pd.DataFrame(stats).to_excel(writer, sheet_name='Статистика', index=False)
    
    print(f"\n✓ Сохранено: {output_path}")
    
    # рекомендации
    print("\n" + "=" * 80)
    print("ЧТО НУЖНО ЕЩЕ ДОБАВИТЬ ДЛЯ ПОЛНОГО АНАЛИЗА:")
    print("=" * 80)
    print("""
1. ВХОДЯЩИЕ ДАННЫЕ (по каждому ДЦ за 2025, минимум Q1-Q4):
   ✓ Декларации по НДС (Форма 300) - ЕСТЬ для Q1 2025 и Q1 2026
   □ Реестры ЭСФ (счета-фактуры) - НУЖНЫ для разбора по видам операций
   □ ОСВ (оборотно-сальдовые ведомости) - НУЖНЫ для выручки, расходов, АВ
   □ Расшифровка выручки:
     - Авто (комиссия)
     - Сервис
     - Запчасти
   □ Расшифровка входного НДС:
     - по сервису
     - по запчастям
     - смешанные расходы (аренда, админ, маркетинг)

2. ДЛЯ МОДЕЛИ НАЛОГООБЛОЖЕНИЯ:
   □ Договоры с дистрибьютором (описание комиссионной базы)
   □ Договоры с ГК и между ДЦ (если есть взаимодействие)
   □ Учетная политика ДЦ по НДС и методика распределения

3. ДЛЯ CASH FLOW АНАЛИЗА (по 1 ДЦ):
   □ Движение денежных средств по месяцам (авто, сервис, запчасти)
   □ Средний срок хранения авто на складе
   □ Отсрочка платежа поставщику (дни)
   □ Отсрочка платежа клиента (факт по дебиторке)
   □ Стоимость финансирования (ставка по кредитам/займам)
   □ Взаимозачеты с дистрибьютором (ретро, бонусы, корректировки)

4. ДЛЯ РИСК-АНАЛИЗА:
   □ История претензий и штрафов КГД
   □ Позиции налоговых консультантов по ст. 394 и ст. 256 НК РК
   □ Прецеденты переквалификации "агента" в "торговца" в вашей отрасли
    """)

if __name__ == '__main__':
    main()
