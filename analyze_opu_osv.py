"""
Анализ ОСВ и ОПУ для НДС-модели
Парсит структуру выручки, расходов и входного НДС
"""

import openpyxl
import os
from pathlib import Path

def analyze_osv_structure(filepath):
    """Анализ структуры ОСВ"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    print(f"\n>>> Файл: {Path(filepath).name}")
    print(f"    Листы: {wb.sheetnames}")
    print(f"    Размер: {ws.max_row} строк × {ws.max_column} колонок")
    
    # ищем ключевые счета: 6010 (комиссия), 7010 (сервис), 7110 (запчасти), 3110 (входной НДС)
    key_accounts = {
        '6010': 'Комиссионный доход (авто)',
        '7010': 'Доход от сервиса',
        '7110': 'Доход от запчастей',
        '3110': 'Входной НДС',
        '3120': 'Входной НДС (деtalь)',
        '1100': 'Кассовые счета',
        '1200': 'Расчетные счета',
        '1300': 'Запасы',
    }
    
    found_accounts = {}
    total_sum = 0
    
    for r in range(1, min(ws.max_row+1, 200)):
        cell_val = ws.cell(r, 1).value
        if not cell_val:
            continue
        
        # проверяем, есть ли код счета
        code = str(cell_val).strip()
        
        for key_code, desc in key_accounts.items():
            if key_code in code:
                # ищем сумму (обычно в колонке 5-10)
                for c in range(2, min(15, ws.max_column+1)):
                    val = ws.cell(r, c).value
                    if isinstance(val, (int, float)) and abs(val) > 100:
                        found_accounts[desc] = found_accounts.get(desc, 0) + val
                        total_sum += abs(val)
                        break
    
    if found_accounts:
        print("    Найденные счета:")
        for desc, amt in found_accounts.items():
            print(f"      {desc}: {amt:,.0f} тг")
    
    # смотрим топ строк с числами
    print("    Структура (первые 20 строк с числами):")
    lines_with_nums = []
    for r in range(1, min(100, ws.max_row+1)):
        row_data = []
        has_num = False
        for c in range(1, min(8, ws.max_column+1)):
            val = ws.cell(r, c).value
            if isinstance(val, (int, float)):
                row_data.append(f"{val:,.0f}")
                has_num = True
            elif isinstance(val, str):
                row_data.append(val[:40])
            else:
                row_data.append("")
        
        if has_num and len(row_data) > 1:
            lines_with_nums.append((r, row_data))
    
    for r, data in lines_with_nums[:10]:
        print(f"      R{r}: {' | '.join(data[:4])}")

def analyze_opu_structure(filepath):
    """Анализ структуры ОПУ"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    print(f"\n>>> Файл: {Path(filepath).name}")
    print(f"    Листы: {wb.sheetnames}")
    print(f"    Размер: {ws.max_row} строк × {ws.max_column} колонок")
    
    # ищем разделы доходов и расходов
    sections = {
        'Доходы': [],
        'Расходы': [],
        'НДС': [],
    }
    
    current_section = None
    
    for r in range(1, min(ws.max_row+1, 300)):
        row_text = str(ws.cell(r, 1).value or "").lower()
        
        if 'доход' in row_text:
            current_section = 'Доходы'
        elif 'расход' in row_text or 'затрат' in row_text:
            current_section = 'Расходы'
        elif 'ндс' in row_text or 'налог' in row_text:
            current_section = 'НДС'
        
        # собираем строки с числами
        for c in range(1, min(10, ws.max_column+1)):
            val = ws.cell(r, c).value
            if isinstance(val, (int, float)) and abs(val) > 1000:
                if current_section:
                    line_desc = str(ws.cell(r, 1).value or "")[:50]
                    sections[current_section].append({
                        'desc': line_desc,
                        'value': val,
                        'col': c
                    })
    
    print("    Разделы в ОПУ:")
    for section, items in sections.items():
        if items:
            print(f"      {section}:")
            # топ-5 по величине
            sorted_items = sorted(items, key=lambda x: abs(x['value']), reverse=True)[:5]
            for item in sorted_items:
                print(f"        {item['desc'][:40]}: {item['value']:,.0f} тг")

# основной анализ
nds_path = r'c:\Users\D.Muldabaev\Desktop\Приложения для сервиса\Ндс'

files_to_check = [
    ('ОСВ 2025.xlsx', 'osv'),
    ('ОСВ 2026 1 квартал.xlsx', 'osv'),
    ('Опиу с разбивкой весь период.xlsx', 'opu'),
]

print("\n" + "=" * 100)
print("АНАЛИЗ НОВЫХ ФАЙЛОВ ДЛЯ НДС-МОДЕЛИ")
print("=" * 100)

for fname, ftype in files_to_check:
    fpath = os.path.join(nds_path, fname)
    if os.path.exists(fpath):
        try:
            if ftype == 'osv':
                analyze_osv_structure(fpath)
            else:
                analyze_opu_structure(fpath)
        except Exception as e:
            print(f"\n✗ Ошибка при обработке {fname}: {e}")
    else:
        print(f"\n✗ Файл не найден: {fname}")

print("\n" + "=" * 100)
print("РЕКОМЕНДАЦИИ ПО СЛЕДУЮЩИМ ШАГАМ")
print("=" * 100)
print("""
1. ПРОВЕРИТЬ КОНСИСТЕНТНОСТЬ:
   - Сумма доходов в ОПУ должна совпадать с оборотом в НДС-декларации (Форма 300)
   - Входной НДС в ОСВ должен совпадать с Формой 300
   - Выручка по направлениям (авто, сервис, запчасти) должна быть разбита

2. ЕСЛИ ДАННЫЕ ВЕРНЫ:
   - Экспортировать ОПУ по месяцам (12 месяцев отдельно)
   - Экспортировать входной НДС по типам расходов (3 категории)
   - Готов считать модель AS IS vs TO BE

3. ЕСЛИ ДАННЫЕ НЕПОЛНЫЕ:
   - Нужны дополнительные отчеты из 1С:
     ☐ Регистр "Входящий НДС" с разбивкой по статьям затрат
     ☐ Расшифровка выручки по комиссиям и услугам
     ☐ Уточнить по какому ДЦ данные (указано ли в файлах)
""")
