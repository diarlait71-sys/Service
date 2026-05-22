"""
Streamlit приложение для расчета бонусов отдела сервиса
Версия 3.0 - С поддержкой реальной структуры данных из PowerBI/1C
"""

import datetime
import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import os
import tempfile
import re
from difflib import SequenceMatcher
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.express as px

from config_loader import load_app_config
from mechanics_calculator import MechanicsCalculator, DEFAULT_CATEGORY_RATES, DEFAULT_WORK_TYPE_CATEGORY_KEYWORDS


CONFIG = load_app_config()


def _compute_weight_amount(w, reper: float, fixed_salary: float, bonus_pool: float):
    """Вычисляет сумму в тенге по весу/формуле показателя."""
    import re
    if w.weight != 0:
        return reper * w.weight
    raw = (w.raw_weight or "").strip()
    if not raw:
        return None
    m = re.search(r'(\d+(?:[.,]\d+)?)\s*%', raw)
    if m:
        pct = float(m.group(1).replace(',', '.')) / 100
        raw_lower = raw.lower().replace(' ', '')
        if 'репер' in raw_lower and 'оклад' in raw_lower:
            base = max(reper - fixed_salary, 0.0)
        elif bonus_pool > 0:
            base = bonus_pool
        else:
            base = reper
        return base * pct
    return None


@st.cache_data(ttl=CONFIG["cache"]["facts_ttl_seconds"])
def calculate_bonus_cached(data_folder: str, dealer_center: str, position: str, month: str, consultant_fio: str = "", staff_count_override: int = 0):
    """Кэшированный расчет бонуса для выбранной комбинации ДЦ/должность/месяц/ФИО."""
    from real_bonus_calculator import BonusCalculator, BonusDataLoader

    local_loader = BonusDataLoader(data_folder)
    calculator = BonusCalculator(
        local_loader,
        dealer_center=dealer_center,
        position=position,
        consultant_fio=consultant_fio,
        staff_count_override=staff_count_override,
    )
    result = calculator.calculate_for_month(month)
    indicator_df = calculator.get_indicator_df(month)

    return {
        "result": result,
        "indicator_df": indicator_df,
        "weights_file": calculator.weights_file_path.name,
        "weights_sheet": calculator.weights_sheet_name,
        "logic_file": calculator.logic_file_path.name if calculator.logic_file_path.exists() else "не найден",
        "has_position_rules": calculator.has_position_rules,
        "has_explicit_weights": calculator.has_explicit_weights,
        "has_explicit_logic": calculator.has_explicit_logic,
    }


def show_file_read_diagnostics(err: Exception, data_folder: str = "", required_files: list | None = None):
    """Показывает понятную диагностику: какой файл/колонки не найдены."""
    message = str(err)
    base_dir = Path(data_folder) if data_folder else Path.cwd()

    file_name = ""
    expected_cols = None
    found_cols = None

    if hasattr(err, "file_name"):
        file_name = getattr(err, "file_name", "") or ""
    if hasattr(err, "expected_columns"):
        expected_cols = getattr(err, "expected_columns", None)
    if hasattr(err, "found_columns"):
        found_cols = getattr(err, "found_columns", None)

    if not file_name:
        m_missing_file = re.search(r"Файл не найден: '([^']+)'", message)
        if m_missing_file:
            file_name = m_missing_file.group(1)

    if not file_name:
        m_missing_cols = re.search(r"Файл '([^']+)': не найдены колонки", message)
        if m_missing_cols:
            file_name = m_missing_cols.group(1)

    with st.expander("🔎 Диагностика чтения файлов", expanded=True):
        st.write(f"Папка данных: {base_dir}")
        if file_name:
            st.write(f"Проблемный файл: {file_name}")

        if expected_cols:
            st.write(f"Ожидались колонки: {expected_cols}")
        if found_cols:
            st.write(f"Найдены колонки: {found_cols}")

        st.code(message)

        if required_files:
            status_rows = []
            for fname in required_files:
                fpath = base_dir / fname
                status_rows.append(
                    {
                        "Файл": fname,
                        "Найден": "✅" if fpath.exists() else "❌",
                    }
                )
            st.dataframe(pd.DataFrame(status_rows), use_container_width=True, hide_index=True)

# ============================================================================
# КОНФИГУРАЦИЯ STREAMLIT
# ============================================================================

st.set_page_config(
    page_title="Расчет бонусов - Сервис",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.title("💰 Расчет бонусов отдела сервиса")
st.markdown("---")


def build_recommendations(result) -> list:
    """Формирует краткие рекомендации по KPI."""
    recommendations = []
    if result.services_execution < 0.95:
        recommendations.append("Услуги ниже 95%: проверьте загрузку ремзоны и конверсию в заказ-наряд.")
    elif result.services_execution > 1.1:
        recommendations.append("Услуги >110%: закрепите практики и масштабируйте на другие смены.")

    if result.spare_parts_execution < 0.95:
        recommendations.append("Запчасти ниже 95%: усилить планирование стока и допродажи в заказ-наряде.")
    elif result.spare_parts_execution > 1.1:
        recommendations.append("Запчасти >110%: зафиксируйте ассортимент-лидеры и приоритет пополнения.")

    marginality_min = CONFIG["thresholds"]["marginality_min"]
    negliquidity_max = CONFIG["thresholds"]["negliquidity_max"]

    if result.marginality_fact < marginality_min:
        recommendations.append(
            f"Маржинальность ниже {marginality_min * 100:.0f}%: проверить скидки и долю низкомаржинальных работ."
        )
    if result.negliquidity_fact > negliquidity_max:
        recommendations.append(
            f"Неликвид выше {negliquidity_max * 100:.0f}%: запустить распродажу остатков и ограничить медленный закуп."
        )

    return recommendations


def style_report_sheet(ws, title: str, integer_columns: list = None):
    """Минимально оформляет лист отчета для презентабельного вида."""
    ws.insert_rows(1)
    ws["A1"] = title
    max_col = ws.max_column
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    header_row = 2
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    thin = Side(border_style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="1F4E78")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right")

    # Формат чисел без десятичных, с разделителями групп разрядов.
    if integer_columns:
        header_map = {}
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(row=header_row, column=col_idx).value
            if header_val is not None:
                header_map[str(header_val).strip()] = col_idx

        for col_name in integer_columns:
            if col_name not in header_map:
                continue
            col_idx = header_map[col_name]
            for row_idx in range(header_row + 1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        # Стартуем со 2-й строки, чтобы не попадать на merged title cell в A1.
        max_len = 0
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            value_len = len(str(value)) if value is not None else 0
            if value_len > max_len:
                max_len = value_len
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = ws.dimensions


def _mechanic_card_kpi(description: str) -> str:
    """Укрупняет строку выработки в управленческий KPI-блок карточки механика."""
    txt = str(description or "").lower().strip()
    if "гаран" in txt:
        return "Отдел по гарантии"
    if "сервис" in txt:
        return "Отдел Сервиса"
    if "ппп" in txt or "продаж" in txt or "доп" in txt:
        return "Отдел продаж"
    return "Прочее"


def build_mechanic_card_df(detail_slice: pd.DataFrame) -> pd.DataFrame:
    """Готовит компактную карточку выработки механика по одному ДЦ."""
    if detail_slice is None or detail_slice.empty:
        return pd.DataFrame(columns=["KPI", "Описание", "Тариф", "Выработка н/ч", "Сумма"])

    category_fallback = {
        "warranty": "Гарантия",
        "internal": "Вн сервис",
        "commercial": "Сервис, Доп сервис",
        "exclude": "Исключено",
    }

    card_df = detail_slice.copy()
    card_df["Описание карточки"] = card_df["Группа маппинга"].astype(str).str.strip()
    empty_desc_mask = card_df["Описание карточки"].isin(["", "nan", "None"])
    card_df.loc[empty_desc_mask, "Описание карточки"] = card_df.loc[empty_desc_mask, "Категория"].map(category_fallback).fillna("Прочее")
    card_df["KPI"] = card_df["Описание карточки"].apply(_mechanic_card_kpi)

    grouped = (
        card_df.groupby(["KPI", "Описание карточки", "Ставка"], as_index=False)
        .agg({"Н/Ч": "sum", "Начисление": "sum"})
        .rename(columns={
            "Описание карточки": "Описание",
            "Ставка": "Тариф",
            "Н/Ч": "Выработка н/ч",
            "Начисление": "Сумма",
        })
    )

    kpi_order = {
        "Отдел продаж": 1,
        "Отдел Сервиса": 2,
        "Отдел по гарантии": 3,
        "Прочее": 4,
    }
    grouped["__kpi_order"] = grouped["KPI"].map(kpi_order).fillna(99)
    grouped = grouped.sort_values(["__kpi_order", "KPI", "Описание"]).drop(columns=["__kpi_order"])
    grouped["Тариф"] = grouped["Тариф"].round(0)
    grouped["Выработка н/ч"] = grouped["Выработка н/ч"].round(2)
    grouped["Сумма"] = grouped["Сумма"].round(0)
    return grouped


def write_mechanic_cards_sheet(ws, detail_df: pd.DataFrame):
    """Формирует лист с карточками механиков в разрезе ДЦ."""
    title_fill = PatternFill("solid", fgColor="F3D37A")
    header_fill = PatternFill("solid", fgColor="8E8E8E")
    footer_fill = PatternFill("solid", fgColor="8E8E8E")
    thin = Side(border_style="thin", color="2B2B2B")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths = {"A": 6, "B": 26, "C": 42, "D": 14, "E": 16, "F": 16}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    current_row = 1
    card_number = 1
    for (fio, dc), slice_df in detail_df.groupby(["ФИО", "ДЦ"], sort=True):
        card_df = build_mechanic_card_df(slice_df)
        if card_df.empty:
            continue

        position = ""
        if "Должность (ставки)" in slice_df.columns:
            position_candidates = (
                slice_df["Должность (ставки)"].dropna().astype(str).str.strip().replace("", np.nan).dropna().unique().tolist()
            )
            if position_candidates:
                position = position_candidates[0]

        title_text = str(fio).strip() if not dc else f"{fio} ({dc})"
        if position:
            title_text = f"{title_text} | {position}"
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        title_cell = ws.cell(row=current_row, column=1, value=title_text)
        title_cell.fill = title_fill
        title_cell.font = Font(name="Calibri", size=16, bold=True, color="000000")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        title_cell.border = border
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        headers = ["№", "KPI", "Описание", "Тариф", "Выработка н/ч", "Сумма"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        body_start_row = current_row + 1
        for _, row in card_df.iterrows():
            ws.cell(row=current_row + 1, column=2, value=row["KPI"])
            ws.cell(row=current_row + 1, column=3, value=row["Описание"])
            ws.cell(row=current_row + 1, column=4, value=float(row["Тариф"]))
            ws.cell(row=current_row + 1, column=5, value=float(row["Выработка н/ч"]))
            ws.cell(row=current_row + 1, column=6, value=float(row["Сумма"]))

            for col_idx in range(2, 7):
                body_cell = ws.cell(row=current_row + 1, column=col_idx)
                body_cell.border = border
                body_cell.alignment = Alignment(horizontal="left" if col_idx in (2, 3) else "center", vertical="center")
                if col_idx in (4, 6):
                    body_cell.number_format = '#,##0'
                if col_idx == 5:
                    body_cell.number_format = '#,##0.00'
            current_row += 1

        body_end_row = current_row
        ws.merge_cells(start_row=body_start_row, start_column=1, end_row=body_end_row, end_column=1)
        number_cell = ws.cell(row=body_start_row, column=1, value=card_number)
        number_cell.border = border
        number_cell.alignment = Alignment(horizontal="center", vertical="center")
        number_cell.font = Font(name="Calibri", size=12, bold=True)
        for row_idx in range(body_start_row, body_end_row + 1):
            ws.cell(row=row_idx, column=1).border = border

        current_row += 1
        ws.cell(row=current_row, column=5, value="ИТОГО")
        ws.cell(row=current_row, column=6, value=float(card_df["Сумма"].sum()))
        for col_idx in range(1, 7):
            total_cell = ws.cell(row=current_row, column=col_idx)
            total_cell.border = border
            total_cell.font = Font(bold=True)
            if col_idx == 6:
                total_cell.number_format = '#,##0'
                total_cell.alignment = Alignment(horizontal="center", vertical="center")

        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        footer_cell = ws.cell(row=current_row, column=1, value="ШТРАФ")
        footer_cell.fill = footer_fill
        footer_cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        footer_cell.alignment = Alignment(horizontal="left", vertical="center")
        footer_cell.border = border
        current_row += 2
        card_number += 1

    ws.freeze_panes = "A2"


def make_safe_sheet_name(base_name: str, used_names: set[str] | None = None, max_len: int = 31) -> str:
    """Возвращает безопасное имя листа Excel с учетом ограничений и уникальности."""
    used_names = used_names or set()
    cleaned = str(base_name or "Лист").strip()
    cleaned = cleaned.replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "")
    cleaned = cleaned.replace("[", "").replace("]", "").replace(":", "-")
    cleaned = cleaned[:max_len] if cleaned else "Лист"

    if cleaned not in used_names:
        return cleaned

    i = 2
    while True:
        suffix = f"_{i}"
        candidate = (cleaned[: max_len - len(suffix)] + suffix).strip()
        if candidate not in used_names:
            return candidate
        i += 1


def build_mechanics_cards_export(detail_df: pd.DataFrame, summary_df: pd.DataFrame) -> bytes:
    """Формирует компактный Excel с итогом/детализацией/карточками механиков."""
    export_buf = BytesIO()
    with pd.ExcelWriter(export_buf, engine="openpyxl") as wr:
        summary_df.to_excel(wr, sheet_name="Итог", index=False)
        detail_df.to_excel(wr, sheet_name="Детализация", index=False)
        style_report_sheet(wr.book["Итог"], "Итог по механикам", integer_columns=["Итого начисление"])
        style_report_sheet(wr.book["Детализация"], "Детализация по механикам", integer_columns=["Ставка", "Начисление"])

        cards_sheet = wr.book.create_sheet("Карточки механиков")
        write_mechanic_cards_sheet(cards_sheet, detail_df)

        used_names = set(wr.book.sheetnames)
        for dc_name, dc_slice in detail_df.groupby("ДЦ", sort=True):
            if dc_slice.empty:
                continue
            dc_title = str(dc_name).strip() if str(dc_name).strip() else "Без ДЦ"
            safe_dc_sheet = make_safe_sheet_name(f"Карточки {dc_title}", used_names)
            used_names.add(safe_dc_sheet)
            dc_sheet = wr.book.create_sheet(safe_dc_sheet)
            write_mechanic_cards_sheet(dc_sheet, dc_slice)

    export_buf.seek(0)
    return export_buf.getvalue()

# Выбор режима
mode = st.radio(
    "🔧 Выберите режим расчета:",
    ["🆕 Бонусы отдела сервиса", "👨‍🔧 Механики", "📊 Остальные сотрудники"],
    horizontal=True
)


# ============================================================================
# РЕЖИМ 1: БОНУСЫ ОТДЕЛА СЕРВИСА
# ============================================================================

if mode == "🆕 Бонусы отдела сервиса":
    st.header("🆕 Бонусы отдела сервиса")
    
    from real_bonus_calculator import (
        BonusCalculator,
        BonusDataLoader,
        DataLoadingError,
        MissingColumnError,
        load_cached_settings,
    )

    # Статические файлы (настройки/планы — меняются редко)
    STATIC_FILES = CONFIG["files"]["static"]
    # Ежемесячные файлы из 1С (загружаются каждый месяц)
    MONTHLY_FILES = CONFIG["files"]["monthly"]
    ALL_FILES = STATIC_FILES + MONTHLY_FILES

    import shutil

    # Папка для хранения данных по месяцам (рядом с приложением)
    APP_DIR = Path(__file__).parent
    ARCHIVE_DIR = APP_DIR / "месяцы"
    ARCHIVE_DIR.mkdir(exist_ok=True)

    MONTH_NAMES = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                   "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
    YEARS = ["2025", "2026", "2027"]
    default_year = CONFIG["app"].get("default_year", "2026")
    current_year_index = YEARS.index(default_year) if default_year in YEARS else 0

    def _saved_months() -> list:
        """Возвращает отсортированный список сохранённых месяцев (имена папок)."""
        if not ARCHIVE_DIR.exists():
            return []
        return sorted([d.name for d in ARCHIVE_DIR.iterdir() if d.is_dir()])

    def _save_month_data(folder: str, year: str, month: str) -> str:
        """Копирует все файлы из folder в ARCHIVE_DIR/{year}-{month}."""
        dest = ARCHIVE_DIR / f"{year}-{month}"
        dest.mkdir(exist_ok=True)
        for f in ALL_FILES:
            src = os.path.join(folder, f)
            if os.path.exists(src):
                shutil.copy2(src, dest / f)
        return str(dest)

    # Инициализация по умолчанию
    found_files = []
    missing_files = []
    current_folder = str(Path.cwd())

    with st.sidebar:
        st.header("📁 Источник данных")

        if st.button("♻️ Очистить кэш", key="clear_cache"):
            st.cache_data.clear()
            st.success("Кэш очищен")

        data_source = st.radio(
            "Выберите режим:",
            ["📂 Все файлы в папке", "📤 Загрузить факты из 1С (4 файла)", "📤 Загрузить все файлы (7)", "📅 Открыть сохранённый месяц"]
        )

        if data_source == "📂 Все файлы в папке":
            current_folder = str(Path.cwd())
            st.info(f"📂 Папка: {current_folder}")

            found_files = []
            missing_files = []
            for file in ALL_FILES:
                if os.path.exists(os.path.join(current_folder, file)):
                    found_files.append(file)
                else:
                    missing_files.append(file)

            st.success(f"✅ Найдено: {len(found_files)}/{len(ALL_FILES)}")
            if missing_files:
                st.warning(f"⚠️ Отсутствует: {len(missing_files)}")
                with st.expander("Показать"):
                    for f in missing_files:
                        st.write(f"• {f}")

            # Сохранение в архив
            st.markdown("---")
            st.markdown("**💾 Сохранить данные в архив:**")
            save_year = st.selectbox("Год:", YEARS, index=current_year_index, key="save_year_folder")
            save_month = st.selectbox("Месяц:", MONTH_NAMES, key="save_month_folder")
            if st.button("💾 Сохранить", key="btn_save_folder"):
                if len(found_files) == len(ALL_FILES):
                    _save_month_data(current_folder, save_year, save_month)
                    st.success(f"✅ Сохранено: {save_year}-{save_month}")
                else:
                    st.error("❌ Не все файлы найдены — сохранение невозможно")

        elif data_source == "📤 Загрузить факты из 1С (4 файла)":
            st.markdown("### 📤 Ежемесячные данные из 1С")
            st.info("💡 Статические файлы (планы, настройки) берутся из папки автоматически")

            # Проверяем что статические файлы есть в папке
            static_folder = str(Path.cwd())
            static_ok = all(os.path.exists(os.path.join(static_folder, f)) for f in STATIC_FILES)

            if not static_ok:
                missing_static = [f for f in STATIC_FILES if not os.path.exists(os.path.join(static_folder, f))]
                st.error("❌ Статические файлы не найдены в папке:")
                for f in missing_static:
                    st.write(f"• {f}")
            else:
                st.success("✅ Статические файлы найдены")

                uploaded_facts = {}
                fact_descriptions = {
                    "Факт Услуги.xlsx": "📦 Факт по услугам",
                    "Факт Запасные части.xlsx": "🔧 Факт по запчастям",
                    "Факт маржанальность.xlsx": "💵 Факт маржинальности",
                    "Факт по остальным показателям.xlsx": "📊 Факт остальных KPI",
                }

                for file_name, label in fact_descriptions.items():
                    uploaded_facts[file_name] = st.file_uploader(
                        label, type=['xlsx'], key=f"fact_{file_name}"
                    )

                all_uploaded = all(uploaded_facts.values())
                uploaded_count = sum(1 for f in uploaded_facts.values() if f is not None)
                st.info(f"📤 Загружено: {uploaded_count}/4")

                if all_uploaded:
                    st.success("✅ Все факты загружены!")
                    temp_dir = tempfile.mkdtemp()
                    # Копируем статические файлы
                    for file_name in STATIC_FILES:
                        src = os.path.join(static_folder, file_name)
                        dst = os.path.join(temp_dir, file_name)
                        shutil.copy2(src, dst)
                    # Сохраняем загруженные факты
                    for file_name, uploaded_file in uploaded_facts.items():
                        with open(os.path.join(temp_dir, file_name), 'wb') as f:
                            f.write(uploaded_file.getvalue())
                    current_folder = temp_dir
                    found_files = ALL_FILES
                    missing_files = []

                    # Сохранение в архив
                    st.markdown("---")
                    st.markdown("**💾 Сохранить данные в архив:**")
                    save_year = st.selectbox("Год:", YEARS, index=current_year_index, key="save_year_facts")
                    save_month = st.selectbox("Месяц:", MONTH_NAMES, key="save_month_facts")
                    if st.button("💾 Сохранить в архив", key="btn_save_facts"):
                        _save_month_data(current_folder, save_year, save_month)
                        st.success(f"✅ Сохранено: {save_year}-{save_month}")

        elif data_source == "📤 Загрузить все файлы (7)":
            st.markdown("### 📤 Загрузите все файлы")

            uploaded_files = {}
            all_descriptions = {
                "Дц,должность,репер.xlsx": "⚙️ Настройки (ДЦ, должность, репер)",
                "Доля показателей для Руководителя отдела сервиса.xlsx": "⚖️ Веса показателей",
                "план по остальным показателям.xlsx": "📋 Планы остальных KPI",
                "Факт Услуги.xlsx": "📦 Факт по услугам",
                "Факт Запасные части.xlsx": "🔧 Факт по запчастям",
                "Факт маржанальность.xlsx": "💵 Факт маржинальности",
                "Факт по остальным показателям.xlsx": "📊 Факт остальных KPI",
            }

            for file_name, label in all_descriptions.items():
                uploaded_files[file_name] = st.file_uploader(
                    label, type=['xlsx'], key=f"all_{file_name}"
                )

            all_uploaded = all(uploaded_files.values())
            uploaded_count = sum(1 for f in uploaded_files.values() if f is not None)
            st.info(f"📤 Загружено: {uploaded_count}/{len(ALL_FILES)}")

            if all_uploaded:
                st.success("✅ Все файлы загружены!")
                temp_dir = tempfile.mkdtemp()
                for file_name, uploaded_file in uploaded_files.items():
                    with open(os.path.join(temp_dir, file_name), 'wb') as f:
                        f.write(uploaded_file.getvalue())
                current_folder = temp_dir
                found_files = ALL_FILES
                missing_files = []

                # Сохранение в архив
                st.markdown("---")
                st.markdown("**💾 Сохранить данные в архив:**")
                save_year = st.selectbox("Год:", YEARS, index=current_year_index, key="save_year_all")
                save_month = st.selectbox("Месяц:", MONTH_NAMES, key="save_month_all")
                if st.button("💾 Сохранить в архив", key="btn_save_all"):
                    _save_month_data(current_folder, save_year, save_month)
                    st.success(f"✅ Сохранено: {save_year}-{save_month}")

        elif data_source == "📅 Открыть сохранённый месяц":
            st.markdown("### 📅 Архив месяцев")
            saved = _saved_months()
            if not saved:
                st.warning("⚠️ Нет сохранённых месяцев.\nЗагрузите данные и нажмите **Сохранить**.")
            else:
                selected_archive = st.selectbox("Выберите месяц:", saved, key="archive_select")
                archive_path = ARCHIVE_DIR / selected_archive

                # Проверяем наличие файлов в архиве
                arch_found = [f for f in ALL_FILES if (archive_path / f).exists()]
                arch_missing = [f for f in ALL_FILES if not (archive_path / f).exists()]

                if arch_missing:
                    st.warning(f"⚠️ Неполный архив: найдено {len(arch_found)}/{len(ALL_FILES)}")
                    with st.expander("Отсутствующие файлы"):
                        for f in arch_missing:
                            st.write(f"• {f}")
                else:
                    st.success(f"✅ Архив полный: {len(arch_found)}/{len(ALL_FILES)} файлов")

                current_folder = str(archive_path)
                found_files = arch_found
                missing_files = arch_missing

                # Удаление архива
                st.markdown("---")
                if st.button("🗑️ Удалить этот архив", key="btn_del_archive"):
                    shutil.rmtree(archive_path, ignore_errors=True)
                    st.warning(f"Архив {selected_archive} удалён. Обновите страницу.")

    # Основной расчет
    if len(found_files) == len(ALL_FILES):
        try:
            # Инициализируем загрузчик данных
            loader = BonusDataLoader(current_folder)

            with st.spinner("Загрузка данных..."):
                all_settings = load_cached_settings(current_folder)
                st.success(f"✅ Данные загружены. Найдено должностей: {len(all_settings)}")

            # Выбор ДЦ и должности
            st.markdown("---")
            st.subheader("🏙️ Выбор дилерского центра и должности")

            col_dc, col_pos = st.columns(2)

            with col_dc:
                dealer_centers = sorted(set(s.dealer_center for s in all_settings))
                selected_dc = st.selectbox("🏢 Дилерский центр:", dealer_centers)

            with col_pos:
                positions_for_dc = sorted(set(
                    s.position for s in all_settings if s.dealer_center == selected_dc
                ))
                selected_position = st.selectbox("👤 Должность:", positions_for_dc)

            # Находим репер для выбранной комбинации
            matched = [s for s in all_settings
                       if s.dealer_center == selected_dc and s.position == selected_position]
            selected_reper = matched[0].reper if matched else 0
            selected_fixed_salary = matched[0].fixed_salary if matched else 0.0
            selected_bonus_pool = matched[0].bonus_pool if matched else 0.0
            selected_manager = matched[0].manager_name if matched else ""
            selected_weights_profile = matched[0].weights_profile if matched else ""
            selected_logic_profile = matched[0].logic_profile if matched else ""
            info_line = f"💰 Репер: **{selected_reper:,.0f} тг**"
            if selected_manager:
                info_line += f"  |  👤 Руководитель: **{selected_manager}**"
            st.info(info_line)

            profile_note = "⚙️ Профиль весов: **по должности (авто)**"
            if selected_weights_profile:
                profile_note = f"⚙️ Профиль весов: **{selected_weights_profile}**"
            profile_note += "  |  📐 Профиль логики: **по умолчанию**"
            if selected_logic_profile:
                profile_note = profile_note.replace("по умолчанию", selected_logic_profile)
            st.caption(profile_note)

            position_low = selected_position.lower()
            is_service_consultant = ("сервис" in position_low and "консульт" in position_low)
            consultant_mode = "Среднее по всем"
            selected_consultant_fio = ""
            consultant_fio_options = []
            staff_count_from_settings = 1
            staff_count_override = 0

            if is_service_consultant:
                try:
                    fio_facts = loader.load_service_consultant_fio_facts(selected_dc)
                    consultant_fio_options = sorted(fio_facts.keys())
                except Exception:
                    consultant_fio_options = []

                try:
                    staff_count_from_settings = loader.count_position_staff(selected_dc, selected_position)
                except Exception:
                    staff_count_from_settings = 1

                fio_count = len(consultant_fio_options)

                # --- Блок: количество человек для деления плана ---
                with st.expander("📊 Деление плана на количество консультантов", expanded=True):
                    col_cnt1, col_cnt2, col_cnt3 = st.columns([1, 1, 2])
                    with col_cnt1:
                        st.metric("По настройкам (файл ДЦ)", staff_count_from_settings, help="Количество сотрудников по файлу Дц,должность,репер.xlsx")
                    with col_cnt2:
                        st.metric("По файлу факта (ФИО)", fio_count, help="Количество уникальных ФИО в файле Сервис консультант факт.xlsx")
                    with col_cnt3:
                        staff_count_override = st.number_input(
                            "Делить план на (ручная корректировка)",
                            min_value=1,
                            max_value=50,
                            value=staff_count_from_settings,
                            step=1,
                            key="staff_count_override",
                            help="Оставьте равным настройкам или измените вручную. Это число используется для деления общего плана ДЦ.",
                        )
                    if fio_count > 0 and fio_count != staff_count_from_settings:
                        if fio_count > staff_count_from_settings:
                            st.warning(
                                f"⚠️ В файле факта найдено **{fio_count}** консультантов, "
                                f"но по настройкам план делится на **{staff_count_from_settings}**. "
                                f"Фактически людей **больше** — план на каждого завышен. "
                                f"Рекомендуется установить деление на **{fio_count}**."
                            )
                        else:
                            st.warning(
                                f"⚠️ В файле факта найдено **{fio_count}** консультантов, "
                                f"но по настройкам план делится на **{staff_count_from_settings}**. "
                                f"Фактически людей **меньше** — план на каждого занижен. "
                                f"Рекомендуется установить деление на **{fio_count}**."
                            )

                if consultant_fio_options:
                    col_mode, col_fio = st.columns([1, 2])
                    with col_mode:
                        consultant_mode = st.radio(
                            "Режим расчета",
                            ["По ФИО", "По всем отдельно", "Среднее по всем"],
                            horizontal=False,
                            key="consultant_mode",
                        )
                    with col_fio:
                        if consultant_mode == "По ФИО":
                            selected_consultant_fio = st.selectbox(
                                "ФИО сервис-консультанта",
                                consultant_fio_options,
                                key="consultant_fio",
                            )
                        elif consultant_mode == "По всем отдельно":
                            st.info(f"Будет рассчитан список по всем консультантам ({len(consultant_fio_options)} чел.)")
                        else:
                            st.info("Будет рассчитано среднее значение по всем консультантам ДЦ")
                else:
                    st.warning("Не найдены ФИО сервис-консультантов в файле факта — используется среднее/общий режим.")

            with st.expander("🧩 Как настраивать веса и логику по ДЦ/должностям"):
                st.markdown(
                    """
                    Добавьте в файл **Дц,должность,репер.xlsx** (лист Лист1) две колонки:

                    - **Профиль весов**
                    - **Профиль логики**

                    Как это работает:
                    - Если в нескольких ДЦ укажете одинаковый профиль, будет применяться один и тот же файл (групповая настройка).
                    - Если для одного ДЦ нужен отдельный расчет, задайте уникальный профиль только для этой строки.
                    - Если колонка пустая, остается авто-режим по должности/по умолчанию.

                    Имена файлов:
                    - Веса: **Доля показателей для {Профиль весов}.xlsx**
                    - Логика: **Формула расчета - {Профиль логики}.xlsx**
                    """
                )

            # Для должности без явных правил не показываем fallback логики/весов от других позиций.
            try:
                has_explicit_weights = loader.has_explicit_weights_for_position(
                    position=selected_position,
                    weights_profile=selected_weights_profile,
                )
            except Exception:
                has_explicit_weights = False

            try:
                has_explicit_logic = loader.has_explicit_logic_for_position(
                    position=selected_position,
                    logic_profile=selected_logic_profile,
                )
            except Exception:
                has_explicit_logic = False

            try:
                current_weights = loader.load_metric_weights(
                    position=selected_position,
                    weights_profile=selected_weights_profile
                )
            except Exception:
                current_weights = {}

            # Важно: правила считаем доступными, если смогли загрузить веса
            # (включая fallback на общий файл), а не только при "явном" профиле.
            position_rules_ready = bool(current_weights)

            # Вкладки
            tab1, tab2, tab3, tab4, tab5 = st.tabs(
                ["📊 Справочные данные", "📋 Формулы расчёта", "🎯 Веса показателей", "💰 Расчет бонуса", "📈 Аналитика"]
            )
            
            # TAB 1: Справочные данные
            with tab1:
                st.subheader("📊 Справочные данные")
                
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Дилерский центр", selected_dc)
                with col2:
                    st.metric("Должность", selected_position)
                with col3:
                    st.metric("Реперный показатель", f"{selected_reper:,.0f} тг" if matched else "-")
                with col4:
                    st.metric("ФИО сотрудника", selected_manager if selected_manager else "-")

                col5, col6, col7 = st.columns(3)
                with col5:
                    st.metric(
                        "Оклад (фиксированная часть)",
                        f"{selected_fixed_salary:,.0f} тг" if selected_fixed_salary else "—",
                        help="Колонка «Фикс» из файла Дц,должность,репер.xlsx",
                    )
                with col6:
                    st.metric(
                        "Бонусный пул",
                        f"{selected_bonus_pool:,.0f} тг" if selected_bonus_pool else "—",
                        help="Колонка «Бонус» из файла Дц,должность,репер.xlsx",
                    )
                with col7:
                    if selected_fixed_salary and selected_bonus_pool:
                        st.metric(
                            "Итого (Фикс + Бонус)",
                            f"{selected_fixed_salary + selected_bonus_pool:,.0f} тг",
                        )
                    elif selected_reper:
                        st.metric("Репер (для расчёта)", f"{selected_reper:,.0f} тг")
                
                st.markdown("---")

                if not position_rules_ready:
                    st.info("Для этой должности не удалось загрузить веса KPI. Доступен расчет только по Репер / Фикс / Бонус.")
                elif not (has_explicit_weights and has_explicit_logic):
                    st.warning("Для этой должности применен fallback-файл правил (без отдельного профиля по строке настроек).")
                
                # Источники данных
                with st.expander("📁 Источники данных для расчёта"):
                    st.markdown(
                        f"""
| Данные | Файл |
|---|---|
| Оклад, бонус, репер | `Дц,должность,репер.xlsx` |
| Веса показателей | `Доля показателей для ... .xlsx` |
| Формулы расчёта | `Формула расчета.xlsx` |
| Факт услуги (план/факт) | `Факт Услуги.xlsx` |
| Факт запчастей (план/факт) | `Факт Запасные части.xlsx` |
| Маржинальность | `Факт маржанальность.xlsx` |
| **Доп. KPI (допы) — план** | `план по остальным показателям.xlsx` |
| **Доп. KPI (допы) — факт** | `Факт по остальным показателям.xlsx` |
| Факт сервис-консультанта | `Сервис конультант факт.xlsx` |
                        """
                    )
                
                # Показываем веса для выбранной должности
                st.subheader("Веса показателей")
                if current_weights:
                    weights_data = []
                    for name, w in current_weights.items():
                        # Продажа доп. оборудования — источник данных уточняется, временно скрыт
                        name_low = name.lower()
                        if "доп" in name_low and ("оборудован" in name_low or "допоборуд" in name_low):
                            continue
                        weight_logic = w.raw_weight if w.raw_weight else f"{w.weight * 100:.1f}%"
                        amt = _compute_weight_amount(w, selected_reper, selected_fixed_salary, selected_bonus_pool)
                        weights_data.append({
                            "Показатель": name,
                            "Логика веса": weight_logic,
                            "Сумма ₸": f"{amt:,.0f}" if amt is not None else "-",
                            "Описание": w.description if w.description else "-",
                        })
                    st.dataframe(pd.DataFrame(weights_data), use_container_width=True, hide_index=True)
                else:
                    st.warning("Файл весов не найден")
            
            # TAB 2: Формулы расчёта
            with tab2:
                st.subheader("📋 Формулы расчёта")
                if not position_rules_ready:
                    st.info("Для этой должности логика расчёта ещё не задана. Расчёт выполняется только по полям Репер / Фикс / Бонус из настроек.")
                else:
                    logic = loader.load_calculation_logic(
                        logic_profile=selected_logic_profile,
                        position=selected_position,
                    )
                    if logic:
                        for metric, formula in logic.items():
                            with st.expander(f"🔹 {metric}"):
                                st.write(formula)
                    else:
                        st.info("ℹ️ Файл логики расчета не найден")
            
            # TAB 3: Веса показателей
            with tab3:
                st.subheader("🎯 Веса показателей")
                if not position_rules_ready:
                    st.info("Для этой должности веса пока не заданы. До настройки отдельных правил KPI в расчёт не включаются.")
                elif current_weights:
                    weights_list = list(current_weights.items())
                    has_formula_weights = any(w.raw_weight for _, w in weights_list)

                    if has_formula_weights:
                        formula_rows = []
                        for name, w in weights_list:
                            amt = _compute_weight_amount(w, selected_reper, selected_fixed_salary, selected_bonus_pool)
                            formula_rows.append({
                                "Показатель": name,
                                "Формула/вес": (w.raw_weight if w.raw_weight else f"{w.weight * 100:.1f}%"),
                                "Сумма ₸": f"{amt:,.0f}" if amt is not None else "-",
                                "Описание": w.description if w.description else "-",
                                "Штрафы": w.penalty_flag if w.penalty_flag else "-",
                            })
                        formula_df = pd.DataFrame(formula_rows)
                        st.dataframe(formula_df, use_container_width=True, hide_index=True)
                        if selected_fixed_salary > 0:
                            st.info(f"База расчёта: Репер {selected_reper:,.0f} − Оклад {selected_fixed_salary:,.0f} = **{selected_reper - selected_fixed_salary:,.0f} ₸**")
                        else:
                            st.info("Для этой должности применяются формульные веса с листа Excel.")
                    else:
                        col1, col2 = st.columns([2, 1])
                        with col1:
                            df_weights = pd.DataFrame({
                                "Показатель": [name for name, _ in weights_list],
                                "Вес": [w.weight for _, w in weights_list]
                            })
                            fig = px.bar(df_weights, x="Показатель", y="Вес",
                                        color="Вес", color_continuous_scale="RdYlGn",
                                        title=f"Веса для должности: {selected_position}")
                            st.plotly_chart(fig, use_container_width=True)
                        with col2:
                            st.metric("Сумма всех весов", f"{sum(w.weight for _, w in weights_list):.2f}")
                else:
                    st.warning("Файл весов не найден")
            
            # TAB 4: Расчет бонуса
            with tab4:
                st.subheader("💰 Расчет бонуса за месяц")

                _MONTHS = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                           "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
                month = st.selectbox(
                    "Выберите месяц",
                    _MONTHS,
                    index=datetime.datetime.now().month - 1
                )

                try:
                    weights_file = loader.resolve_weights_file(
                        position=selected_position,
                        weights_profile=selected_weights_profile
                    )
                    logic_file = loader.resolve_logic_file(logic_profile=selected_logic_profile)
                    weights_sheet = loader.resolve_weights_sheet(
                        weights_file,
                        position=selected_position,
                        weights_profile=selected_weights_profile,
                    )
                    with st.expander("🔍 Источники правил для этого расчёта"):
                        st.write(f"**Файл весов:** {weights_file.name}")
                        st.write(f"**Лист весов:** {weights_sheet}")
                        st.write(f"**Файл логики:** {logic_file.name if logic_file.exists() else 'не найден'}")
                        st.write(f"**Профиль весов:** {selected_weights_profile or 'авто по должности'}")
                        st.write(f"**Профиль логики:** {selected_logic_profile or 'по умолчанию'}")
                except Exception as e:
                    st.warning(f"Не удалось определить источники правил: {e}")
                
                if st.button("🔄 Рассчитать", key="calc_bonus"):
                    with st.spinner(f"Расчет бонуса за {month} для {selected_dc} / {selected_position}..."):
                        try:
                            if is_service_consultant and consultant_mode == "По всем отдельно" and consultant_fio_options:
                                rows = []
                                for fio in consultant_fio_options:
                                    calc_data = calculate_bonus_cached(
                                        current_folder,
                                        selected_dc,
                                        selected_position,
                                        month,
                                        fio,
                                        int(staff_count_override),
                                    )
                                    r = calc_data["result"]
                                    rows.append({
                                        "ФИО": fio,
                                        "Оклад": r.base_salary_bonus,
                                        "Бонус": r.total_bonus,
                                        "Итого к начислению": r.base_salary_bonus + r.total_bonus,
                                        "План услуги": r.services_plan_value,
                                        "Факт услуги": r.services_fact_value,
                                        "План запчасти": r.spare_parts_plan_value,
                                        "Факт запчасти": r.spare_parts_fact_value,
                                    })

                                st.session_state['consultant_results'] = pd.DataFrame(rows)
                                st.session_state.pop('bonus_result', None)
                                st.session_state['bonus_calc_context'] = (
                                    current_folder, selected_dc, selected_position, month, consultant_mode, "ALL"
                                )
                                st.success(f"✅ Расчет выполнен по всем консультантам: {selected_dc} / {selected_position}")
                            else:
                                fio_for_calc = selected_consultant_fio if (is_service_consultant and consultant_mode == "По ФИО") else ""
                                calc_data = calculate_bonus_cached(
                                    current_folder,
                                    selected_dc,
                                    selected_position,
                                    month,
                                    fio_for_calc,
                                    int(staff_count_override) if is_service_consultant else 0,
                                )
                                result = calc_data["result"]
                                indicator_df = calc_data["indicator_df"]

                                st.session_state['bonus_result'] = result
                                st.session_state['indicator_df'] = indicator_df
                                st.session_state.pop('consultant_results', None)
                                st.session_state['rule_sources'] = {
                                    'weights_file': calc_data["weights_file"],
                                    'weights_sheet': calc_data["weights_sheet"],
                                    'logic_file': calc_data["logic_file"],
                                    'has_position_rules': calc_data["has_position_rules"],
                                    'has_explicit_weights': calc_data["has_explicit_weights"],
                                    'has_explicit_logic': calc_data["has_explicit_logic"],
                                    'weights_profile': selected_weights_profile or 'авто по должности',
                                    'logic_profile': selected_logic_profile or 'по умолчанию'
                                }
                                st.session_state['bonus_calc_context'] = (
                                    current_folder, selected_dc, selected_position, month, consultant_mode, fio_for_calc
                                )
                                suffix = f" / {fio_for_calc}" if fio_for_calc else ""
                                st.success(f"✅ Расчет выполнен: {selected_dc} / {selected_position}{suffix}")
                        except MissingColumnError as e:
                            st.error(f"🔧 Ошибка структуры файла: {e}")
                            st.info("💡 Проверьте, что названия колонок в Excel совпадают с ожидаемыми")
                            show_file_read_diagnostics(
                                e,
                                data_folder=current_folder,
                                required_files=ALL_FILES + ["Формула расчета.xlsx", "Сервис конультант факт.xlsx"],
                            )
                        except DataLoadingError as e:
                            st.error(f"⚠️ Ошибка данных: {e}")
                            show_file_read_diagnostics(
                                e,
                                data_folder=current_folder,
                                required_files=ALL_FILES + ["Формула расчета.xlsx", "Сервис конультант факт.xlsx"],
                            )
                        except Exception as e:
                            st.error(f"❌ Ошибка расчёта: {e}")
                            import traceback
                            st.code(traceback.format_exc())
                
                # Показываем результаты
                if 'consultant_results' in st.session_state:
                    _ctx = st.session_state.get('bonus_calc_context')
                    current_ctx = (current_folder, selected_dc, selected_position, month, consultant_mode, "ALL")
                    if _ctx and _ctx != current_ctx:
                        st.warning("⚠️ Выбранные параметры изменились — нажмите **Рассчитать** для обновления.")
                    st.markdown("### 👥 Расчеты по всем сервис-консультантам")
                    df_all = st.session_state['consultant_results'].copy()
                    if not df_all.empty:
                        totals = {
                            "ФИО": "ИТОГО",
                            "Оклад": df_all["Оклад"].sum(),
                            "Бонус": df_all["Бонус"].sum(),
                            "Итого к начислению": df_all["Итого к начислению"].sum(),
                            "План услуги": df_all["План услуги"].sum(),
                            "Факт услуги": df_all["Факт услуги"].sum(),
                            "План запчасти": df_all["План запчасти"].sum(),
                            "Факт запчасти": df_all["Факт запчасти"].sum(),
                        }
                        df_show = pd.concat([df_all, pd.DataFrame([totals])], ignore_index=True)
                        st.dataframe(df_show, use_container_width=True, hide_index=True)

                if 'bonus_result' in st.session_state:
                    _ctx = st.session_state.get('bonus_calc_context')
                    _fio_for_ctx = selected_consultant_fio if (is_service_consultant and consultant_mode == "По ФИО") else ""
                    if consultant_mode == "По всем отдельно":
                        _fio_for_ctx = "ALL"
                    current_ctx = (current_folder, selected_dc, selected_position, month, consultant_mode, _fio_for_ctx)
                    if _ctx and _ctx != current_ctx:
                        st.warning("⚠️ Выбранные параметры изменились — нажмите **Рассчитать** для обновления.")
                    result = st.session_state['bonus_result']
                    rule_sources = st.session_state.get('rule_sources', {})

                    if rule_sources:
                        with st.expander("🔍 Использованные источники правил", expanded=False):
                            st.write(f"Файл весов: {rule_sources.get('weights_file', '-')}")
                            st.write(f"Лист весов: {rule_sources.get('weights_sheet', '-')}")
                            st.write(f"Файл логики: {rule_sources.get('logic_file', '-')}")
                            st.write(f"Профиль весов: {rule_sources.get('weights_profile', '-')}")
                            st.write(f"Профиль логики: {rule_sources.get('logic_profile', '-')}")

                    # Итоговый бонус в красивой карточке
                    st.markdown("---")
                    col1, col2, col3 = st.columns([2, 1, 1])

                    with col1:
                        total_payout = result.base_salary_bonus + result.total_bonus
                        st.markdown(f"""
                        <div style="background-color: #f0f8ff; padding: 20px; border-radius: 10px; border-left: 5px solid #1f77b4;">
                            <h2 style="color: #1f77b4; margin: 0;">💰 Итого к начислению</h2>
                            <h1 style="color: #2e8b57; margin: 10px 0; font-size: 2.5em;">{total_payout:,.0f} тг</h1>
                            <p style="margin: 5px 0; color: #555;">🏦 Оклад (фикса): <b>{result.base_salary_bonus:,.0f} тг</b></p>
                            <p style="margin: 5px 0; color: #555;">🎯 Бонус: <b>{result.total_bonus:,.0f} тг</b></p>
                            <p style="margin: 5px 0; color: #555;">👤 ФИО сотрудника: <b>{selected_manager if selected_manager else '-'}</b></p>
                            <p style="margin: 0; color: #666;">за {month} 2026 г. для {selected_dc}</p>
                        </div>
                        """, unsafe_allow_html=True)

                    with col2:
                        st.metric("ДЦ", selected_dc)
                        st.metric("Должность", result.position)

                    with col3:
                        st.metric("Репер", f"{result.reper:,.0f} тг")
                        st.metric("Месяц", month)

                    # Детальная таблица показателей
                    st.markdown("### 📊 Детализация расчета")

                    has_position_rules_result = st.session_state.get('rule_sources', {}).get('has_position_rules', position_rules_ready)
                    has_formula_weights = any((w.raw_weight or "").strip() for w in current_weights.values())

                    if not has_position_rules_result:
                        details_df = pd.DataFrame([
                            {
                                "Показатель": "Репер",
                                "Тип": "Настройка",
                                "Вес": "-",
                                "План": "-",
                                "Факт": "-",
                                "Бонус": f"{result.reper:,.0f} тг",
                                "Статус": "ℹ️",
                            },
                            {
                                "Показатель": "Оклад (фикс)",
                                "Тип": "Фиксированный",
                                "Вес": "-",
                                "План": "-",
                                "Факт": "-",
                                "Бонус": f"{result.base_salary_bonus:,.0f} тг",
                                "Статус": "✅",
                            },
                            {
                                "Показатель": "Бонус по настройке",
                                "Тип": "Фиксированный",
                                "Вес": "-",
                                "План": "-",
                                "Факт": "-",
                                "Бонус": f"{result.total_bonus:,.0f} тг",
                                "Статус": "✅",
                            },
                        ])
                    elif has_formula_weights and current_weights:
                        other_plans = loader.load_other_plans()
                        other_facts = loader.load_other_facts(selected_dc)

                        def _metric_plan_fact(metric_name: str):
                            low = metric_name.lower()
                            if "услуг" in low:
                                return f"{result.services_plan_value:,.0f} тг", f"{result.services_fact_value:,.0f} тг"
                            if "запчаст" in low:
                                return f"{result.spare_parts_plan_value:,.0f} тг", f"{result.spare_parts_fact_value:,.0f} тг"
                            if "маржаналь" in low or "маржиналь" in low:
                                return "30%", f"{result.marginality_fact * 100:.1f}%"
                            if "неликвид" in low:
                                return "7%", f"{result.negliquidity_fact * 100:.1f}%"
                            plan_val = other_plans.get(metric_name, "-")
                            fact_val = other_facts.get(metric_name, "-")
                            return str(plan_val), str(fact_val)

                        def _metric_bonus(metric_name: str) -> float:
                            low = metric_name.lower()
                            if "оклад" in low and ("база" in low or "фикс" in low):
                                return result.base_salary_bonus
                            if "услуг" in low:
                                return result.services_plan_bonus
                            if "запчаст" in low:
                                return result.spare_parts_bonus
                            if "маржаналь" in low or "маржиналь" in low:
                                return result.marginality_bonus
                            if "неликвид" in low:
                                return result.negliquidity_deduction
                            return result.other_metric_bonuses.get(metric_name, 0.0)

                        formula_rows = []
                        for metric_name, w in current_weights.items():
                            plan_value, fact_value = _metric_plan_fact(metric_name)
                            metric_bonus = _metric_bonus(metric_name)
                            weight_view = w.raw_weight if w.raw_weight else (f"{w.weight * 100:.0f}%" if w.weight != 0 else "-")
                            row_type = "Штраф" if metric_bonus < 0 else "Бонус"
                            if "оклад" in metric_name.lower() and ("база" in metric_name.lower() or "фикс" in metric_name.lower()):
                                row_type = "Фиксированный"

                            formula_rows.append({
                                "Показатель": metric_name,
                                "Тип": row_type,
                                "Вес": weight_view,
                                "План": plan_value,
                                "Факт": fact_value,
                                "Бонус": f"{metric_bonus:,.0f} тг",
                                "Статус": "✅" if metric_bonus >= 0 else "❌",
                            })

                        details_df = pd.DataFrame(formula_rows)
                    else:
                        def _wt_fmt(metric_name, default="-"):
                            w = current_weights.get(metric_name)
                            return f"{w.weight * 100:.0f}%" if w else default

                        details_data = {
                            "Показатель": [
                                "Оклад (фикса)",
                                "Бонус: услуги",
                                "Бонус: запчасти",
                                "Бонус: маржинальность",
                                "Штраф: неликвид",
                                "Бонус: другие KPI"
                            ],
                            "Тип": [
                                "Фиксированный",
                                "Бонус",
                                "Бонус",
                                "Бонус",
                                "Штраф",
                                "Бонус"
                            ],
                            "Вес": [
                                _wt_fmt("Оклад (База)", "50%"),
                                _wt_fmt("Выполнение плана продаж по продажи услуг", "23%"),
                                _wt_fmt("Выполнение плана продаж по Запасным Частям", "23%"),
                                _wt_fmt("Выполнение  маржанальности", "4%"),
                                _wt_fmt("Соблюдение уровня неликвида на складе запасных частей", "-10%"),
                                "Разные"
                            ],
                            "План": [
                                "-",
                                f"{result.services_plan_value:,.0f} тг",
                                f"{result.spare_parts_plan_value:,.0f} тг",
                                "30%",
                                "7%",
                                "-"
                            ],
                            "Факт": [
                                "-",
                                f"{result.services_fact_value:,.0f} тг",
                                f"{result.spare_parts_fact_value:,.0f} тг",
                                f"{result.marginality_fact*100:.1f}%",
                                f"{result.negliquidity_fact*100:.1f}%",
                                "-"
                            ],
                            "Бонус": [
                                f"{result.base_salary_bonus:,.0f} тг",
                                f"{result.services_plan_bonus:,.0f} тг",
                                f"{result.spare_parts_bonus:,.0f} тг",
                                f"{result.marginality_bonus:,.0f} тг",
                                f"{result.negliquidity_deduction:,.0f} тг",
                                f"{sum(result.other_metric_bonuses.values()):,.0f} тг"
                            ],
                            "Статус": [
                                "✅ Базовый",
                                "✅" if result.services_execution >= 0.95 else "⚠️",
                                "✅" if result.spare_parts_execution >= 0.95 else "⚠️",
                                "✅" if result.marginality_fact >= 0.3 else "❌",
                                "✅" if result.negliquidity_fact <= 0.07 else "❌",
                                "✅"
                            ]
                        }

                        details_df = pd.DataFrame(details_data)

                    # Применяем цветовое форматирование
                    def color_status(val):
                        if val == "✅":
                            return 'background-color: #d4edda; color: #155724'
                        elif val == "⚠️":
                            return 'background-color: #fff3cd; color: #856404'
                        elif val == "❌":
                            return 'background-color: #f8d7da; color: #721c24'
                        return ''

                    styled_df = details_df.style.applymap(color_status, subset=['Статус'])
                    st.dataframe(styled_df, use_container_width=True, hide_index=True)

                    # Кнопки экспорта
                    col1, col2 = st.columns(2)

                    with col1:
                        if 'indicator_df' in st.session_state:
                            export_buffer = BytesIO()
                            with pd.ExcelWriter(export_buffer, engine='openpyxl') as writer:
                                # Лист с итогами
                                details_df.to_excel(writer, sheet_name='Итоги', index=False)

                                # Лист со всеми показателями
                                indicator_df = st.session_state['indicator_df']
                                indicator_df.to_excel(writer, sheet_name='Показатели', index=False)

                                # Лист с настройками
                                settings_data = {
                                    'Параметр': ['ДЦ', 'Должность', 'ФИО сотрудника', 'Репер', 'Месяц', 'Оклад', 'Бонус', 'Итого к начислению'],
                                    'Значение': [
                                        selected_dc,
                                        result.position,
                                        selected_manager if selected_manager else '-',
                                        f"{result.reper:,.0f} тг",
                                        month,
                                        f"{result.base_salary_bonus:,.0f} тг",
                                        f"{result.total_bonus:,.0f} тг",
                                        f"{(result.base_salary_bonus + result.total_bonus):,.0f} тг"
                                    ]
                                }
                                settings_df = pd.DataFrame(settings_data)
                                settings_df.to_excel(writer, sheet_name='Настройки', index=False)

                                # Лист аналитики по городам/ДЦ
                                city_export_df = pd.DataFrame(st.session_state.get('city_analytics_raw', []))
                                if not city_export_df.empty:
                                    # Добавляем строку итогов по суммовым колонкам.
                                    sum_columns = ['Бонус', 'Оклад', 'Итого к начислению', 'Репер']
                                    totals_row = {
                                        'ДЦ': 'ИТОГО',
                                        'Должность': '-',
                                        'ФИО сотрудника': '-',
                                        'Статус': '-',
                                        'Риск-факторов': city_export_df['Риск-факторов'].sum() if 'Риск-факторов' in city_export_df.columns else 0,
                                    }
                                    for col in sum_columns:
                                        totals_row[col] = city_export_df[col].sum() if col in city_export_df.columns else 0

                                    city_export_df = pd.concat([city_export_df, pd.DataFrame([totals_row])], ignore_index=True)

                                    city_export_df.to_excel(writer, sheet_name='Аналитика по ДЦ', index=False)

                                # Оформление листов
                                for sheet_name, title in {
                                    'Итоги': 'Итоговый расчет бонуса',
                                    'Показатели': 'Показатели и логика расчета',
                                    'Настройки': 'Параметры выбранного расчета',
                                    'Аналитика по ДЦ': f'Срез по городам за {month}'
                                }.items():
                                    if sheet_name in writer.book.sheetnames:
                                        integer_cols = None
                                        if sheet_name == 'Аналитика по ДЦ':
                                            integer_cols = ['Бонус', 'Оклад', 'Итого к начислению', 'Репер', 'Риск-факторов']
                                        style_report_sheet(writer.book[sheet_name], title, integer_columns=integer_cols)

                            st.download_button(
                                label='📊 Скачать полный отчет Excel',
                                data=export_buffer.getvalue(),
                                file_name=f'bonus_report_{selected_dc}_{month}.xlsx',
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                            )

                    with col2:
                        # Экспорт только показателей
                        if 'indicator_df' in st.session_state:
                            indicator_buffer = BytesIO()
                            with pd.ExcelWriter(indicator_buffer, engine='openpyxl') as writer:
                                st.session_state['indicator_df'].to_excel(writer, sheet_name='Показатели', index=False)

                            st.download_button(
                                label='📋 Скачать показатели по шаблону',
                                data=indicator_buffer.getvalue(),
                                file_name=f'indicators_{selected_dc}_{month}.xlsx',
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                            )
            
            # TAB 5: Аналитика
            with tab5:
                st.subheader("📈 Сводный отчёт по всем должностям и дилерским центрам")

                # Независимый выбор месяца для аналитики
                analytics_month = st.selectbox(
                    "Выберите месяц для сводного отчёта",
                    ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                     "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"],
                    index=datetime.datetime.now().month - 1,
                    key="analytics_month"
                )

                if st.button("🔄 Рассчитать всё", key="calc_all"):
                    all_dc_results = []
                    errors_list = []
                    _total = max(len(all_settings), 1)
                    _prog = st.progress(0, text="Инициализация...")
                    for _i, s in enumerate(all_settings):
                        _prog.progress((_i + 1) / _total,
                                       text=f"[{_i + 1}/{_total}] {s.dealer_center} / {s.position}")
                        try:
                            calc_data = calculate_bonus_cached(
                                current_folder,
                                s.dealer_center,
                                s.position,
                                analytics_month
                            )
                            temp_result = calc_data["result"]
                            temp_recs = build_recommendations(temp_result)
                            all_dc_results.append({
                                'ДЦ': s.dealer_center,
                                'Должность': s.position,
                                'ФИО сотрудника': s.manager_name if s.manager_name else '-',
                                'Бонус': temp_result.total_bonus,
                                'Оклад': temp_result.base_salary_bonus,
                                'Бонус по настройке': s.bonus_pool,
                                'Итого по настройке': s.fixed_salary + s.bonus_pool,
                                'Итого к начислению': temp_result.base_salary_bonus + temp_result.total_bonus,
                                'Услуги %': temp_result.services_execution * 100,
                                'Запчасти %': temp_result.spare_parts_execution * 100,
                                'Маржинальность %': temp_result.marginality_fact * 100,
                                'Неликвид %': temp_result.negliquidity_fact * 100,
                                'Репер': temp_result.reper,
                                'Риск-факторов': len(temp_recs),
                                'Правила по должности': 'Да' if calc_data.get('has_position_rules') else 'Нет',
                                'Статус': 'Отлично' if len(temp_recs) == 0 else ('Внимание' if len(temp_recs) <= 2 else 'Риск')
                            })
                        except Exception as e:
                            errors_list.append(f"{s.dealer_center} / {s.position}: {e}")
                            all_dc_results.append({
                                'ДЦ': s.dealer_center,
                                'Должность': s.position,
                                'ФИО сотрудника': s.manager_name if s.manager_name else '-',
                                'Бонус': 0, 'Оклад': 0, 'Итого к начислению': 0,
                                'Бонус по настройке': s.bonus_pool,
                                'Итого по настройке': s.fixed_salary + s.bonus_pool,
                                'Услуги %': 0, 'Запчасти %': 0, 'Маржинальность %': 0,
                                'Неликвид %': 0, 'Репер': s.reper, 'Риск-факторов': 0,
                                'Правила по должности': 'Нет',
                                'Статус': 'Нет данных'
                            })
                    _prog.empty()
                    st.session_state['all_analytics_raw'] = all_dc_results
                    st.session_state['all_analytics_month'] = analytics_month
                    st.session_state['city_analytics_raw'] = all_dc_results  # совместимость с экспортом
                    if errors_list:
                        with st.expander(f"⚠️ Ошибки при расчёте ({len(errors_list)})"):
                            for err in errors_list:
                                st.write(f"• {err}")
                    st.success(f"✅ Рассчитано {len(all_dc_results)} позиций")

                # Показываем результаты если есть
                if 'all_analytics_raw' in st.session_state:
                    all_analytics_raw = st.session_state['all_analytics_raw']
                    report_month = st.session_state.get('all_analytics_month', analytics_month)
                    comparison_raw = pd.DataFrame(all_analytics_raw)

                    # Итоговые KPI
                    st.markdown("---")
                    k1, k2, k3, k4 = st.columns(4)
                    with k1:
                        st.metric("Всего позиций", len(comparison_raw))
                    with k2:
                        st.metric("Дилерских центров", comparison_raw['ДЦ'].nunique())
                    with k3:
                        st.metric("Должностей", comparison_raw['Должность'].nunique())
                    with k4:
                        total_payout_all = int(comparison_raw['Итого к начислению'].sum())
                        st.metric("Итого к начислению (все)", f"{total_payout_all:,} тг".replace(",", " "))

                    # ТОП-5 по начислению
                    st.markdown(f"### 🏆 ТОП-5 по начислению за {report_month}")
                    top5 = comparison_raw.sort_values('Итого к начислению', ascending=False)[
                        ['ДЦ', 'Должность', 'ФИО сотрудника', 'Итого к начислению', 'Бонус', 'Риск-факторов', 'Статус']
                    ].head(5).copy()
                    st.dataframe(top5, use_container_width=True, hide_index=True)

                    # Группировка по должностям
                    all_positions = sorted(comparison_raw['Должность'].unique())
                    pos_filter = st.multiselect(
                        "Фильтр по должностям:",
                        options=all_positions,
                        default=all_positions,
                        key="pos_filter"
                    )
                    filtered_raw = comparison_raw[comparison_raw['Должность'].isin(pos_filter)]

                    no_rules_df = filtered_raw[filtered_raw['Правила по должности'] == 'Нет'][
                        ['ДЦ', 'Должность', 'ФИО сотрудника', 'Репер', 'Оклад', 'Бонус по настройке']
                    ].copy()
                    if not no_rules_df.empty:
                        st.warning("Для части должностей не заданы отдельные веса/логика. По ним в расчёте используются только Репер / Фикс / Бонус из настроек.")
                        st.dataframe(no_rules_df, use_container_width=True, hide_index=True)

                    st.markdown(f"### ⚖️ Сравнение Репер / Фикс / Бонус по настройкам за {report_month}")
                    compare_df = filtered_raw[[
                        'ДЦ', 'Должность', 'ФИО сотрудника', 'Репер', 'Оклад', 'Бонус по настройке', 'Итого по настройке', 'Правила по должности'
                    ]].copy()
                    compare_df['ДЦ / Должность'] = compare_df['ДЦ'] + ' / ' + compare_df['Должность']
                    st.dataframe(compare_df, use_container_width=True, hide_index=True)

                    compare_chart_df = compare_df.melt(
                        id_vars=['ДЦ / Должность', 'ФИО сотрудника', 'Правила по должности'],
                        value_vars=['Репер', 'Оклад', 'Бонус по настройке'],
                        var_name='Показатель',
                        value_name='Сумма'
                    )
                    fig_compare = px.bar(
                        compare_chart_df,
                        x='ДЦ / Должность', y='Сумма',
                        color='Показатель',
                        barmode='group',
                        hover_data=['ФИО сотрудника', 'Правила по должности'],
                        title=f'Сравнение настроек по ДЦ и должностям — {report_month}'
                    )
                    fig_compare.update_layout(xaxis_tickangle=-45)
                    st.plotly_chart(fig_compare, use_container_width=True)

                    st.markdown("---")
                    for pos in [p for p in all_positions if p in pos_filter]:
                        pos_df = filtered_raw[filtered_raw['Должность'] == pos].sort_values('Итого к начислению', ascending=False).copy()
                        pos_total = int(pos_df['Итого к начислению'].sum())
                        pos_avg_bonus = int(pos_df['Бонус'].mean())
                        with st.expander(f"👤 {pos}  —  ДЦ: {len(pos_df)}  |  Итого: {pos_total:,} тг  |  Средний бонус: {pos_avg_bonus:,} тг".replace(",", " "), expanded=True):
                            # Форматированная таблица
                            display_df = pos_df.copy()
                            display_df['Бонус'] = display_df['Бонус'].apply(lambda x: f"{x:,.0f}")
                            display_df['Оклад'] = display_df['Оклад'].apply(lambda x: f"{x:,.0f}")
                            display_df['Бонус по настройке'] = display_df['Бонус по настройке'].apply(lambda x: f"{x:,.0f}")
                            display_df['Итого по настройке'] = display_df['Итого по настройке'].apply(lambda x: f"{x:,.0f}")
                            display_df['Итого к начислению'] = display_df['Итого к начислению'].apply(lambda x: f"{x:,.0f}")
                            display_df['Репер'] = display_df['Репер'].apply(lambda x: f"{x:,.0f}")
                            display_df['Услуги %'] = display_df['Услуги %'].apply(lambda x: f"{x:.1f}%")
                            display_df['Запчасти %'] = display_df['Запчасти %'].apply(lambda x: f"{x:.1f}%")
                            display_df['Маржинальность %'] = display_df['Маржинальность %'].apply(lambda x: f"{x:.1f}%")
                            display_df['Неликвид %'] = display_df['Неликвид %'].apply(lambda x: f"{x:.1f}%")
                            st.dataframe(display_df.drop(columns=['Должность']), use_container_width=True, hide_index=True)

                    # Сводный график (с фильтром)
                    st.markdown(f"### 📊 Начисления по ДЦ и должностям за {report_month}")
                    fig_all = px.bar(
                        filtered_raw.sort_values('Итого к начислению', ascending=False),
                        x='ДЦ', y='Итого к начислению',
                        color='Должность',
                        barmode='group',
                        hover_data=['ФИО сотрудника', 'Бонус', 'Оклад', 'Правила по должности', 'Статус'],
                        title=f'Сводные начисления по всем должностям — {report_month}'
                    )
                    fig_all.update_layout(xaxis_tickangle=-45)
                    st.plotly_chart(fig_all, use_container_width=True)

                    # Экспорт — многолистовой Excel
                    st.markdown("---")
                    st.markdown("### 📥 Выгрузка сводного отчёта")
                    if st.button("📊 Сформировать сводный Excel", key="export_all_positions"):
                        export_buf = BytesIO()
                        with pd.ExcelWriter(export_buf, engine='openpyxl') as writer:
                            int_cols = ['Бонус', 'Оклад', 'Бонус по настройке', 'Итого по настройке', 'Итого к начислению', 'Репер', 'Риск-факторов']

                            # Сводный лист
                            summary_df = comparison_raw.copy()
                            summary_df.to_excel(writer, sheet_name='Сводный', index=False)
                            style_report_sheet(writer.book['Сводный'], f'Сводный отчёт за {report_month}', integer_columns=int_cols)

                            # Один лист на каждую должность
                            for pos in all_positions:
                                pos_df_export = comparison_raw[comparison_raw['Должность'] == pos].copy()
                                # Безопасное имя листа (макс 31 символ, без спецсимволов)
                                safe_name = pos[:28].replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace('[', '').replace(']', '').replace(':', '')
                                # Добавляем строку итогов
                                totals = {'ДЦ': 'ИТОГО', 'Должность': '-', 'ФИО сотрудника': '-', 'Статус': '-'}
                                for c in int_cols:
                                    totals[c] = pos_df_export[c].sum() if c in pos_df_export.columns else 0
                                for c in ['Услуги %', 'Запчасти %', 'Маржинальность %', 'Неликвид %']:
                                    if c in pos_df_export.columns:
                                        totals[c] = round(pos_df_export[c].mean(), 1)
                                pos_df_export = pd.concat([pos_df_export, pd.DataFrame([totals])], ignore_index=True)
                                pos_df_export.to_excel(writer, sheet_name=safe_name, index=False)
                                style_report_sheet(writer.book[safe_name], f'{pos} — {report_month}', integer_columns=int_cols)

                        st.download_button(
                            label='⬇️ Скачать сводный отчёт Excel',
                            data=export_buf.getvalue(),
                            file_name=f'сводный_отчёт_{report_month}.xlsx',
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            key='dl_all_positions'
                        )
                else:
                    st.info("ℹ️ Нажмите **Рассчитать всё** чтобы сформировать сводный отчёт по всем должностям и ДЦ")
        
        except MissingColumnError as e:
            st.error(f"🔧 Ошибка структуры файла: {e}")
            st.info("💡 Проверьте структуру Excel-файлов и названия колонок")
            show_file_read_diagnostics(
                e,
                data_folder=current_folder,
                required_files=ALL_FILES + ["Формула расчета.xlsx", "Сервис конультант факт.xlsx"],
            )
        except DataLoadingError as e:
            st.error(f"⚠️ Ошибка загрузки/валидации данных: {e}")
            show_file_read_diagnostics(
                e,
                data_folder=current_folder,
                required_files=ALL_FILES + ["Формула расчета.xlsx", "Сервис конультант факт.xlsx"],
            )
        except Exception as e:
            st.error(f"❌ Ошибка: {str(e)}")
            import traceback
            st.text(traceback.format_exc())
    
    else:
        st.error("❌ Не все файлы найдены")
        st.info(f"Убедитесь, что все {len(ALL_FILES)} файлов находятся в текущей папке")


# ============================================================================
# РЕЖИМ 2: РАСЧЕТ ДЛЯ МЕХАНИКОВ
# ============================================================================

elif mode == "👨‍🔧 Механики":
    st.header("👨‍🔧 Расчет бонусов механиков")

    # --- Поиск файла Механики.xlsx ---
    mechanics_file_path = None
    mapping_file_path = None
    app_dir = str(Path(__file__).parent)
    month_folder = st.session_state.get("data_folder", app_dir)
    candidates = [
        os.path.join(month_folder, "Механики.xlsx"),
        os.path.join(app_dir, "Механики.xlsx"),
    ]
    for _c in candidates:
        if os.path.exists(_c):
            mechanics_file_path = _c
            break

    uploaded_mech = st.file_uploader(
        "Загрузить файл Механики.xlsx (если не найден автоматически)",
        type=["xlsx"],
        key="mech_upload",
    )
    if uploaded_mech:
        _tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        _tmp.write(uploaded_mech.read())
        _tmp.flush()
        mechanics_file_path = _tmp.name

    # Маппинг фиксируем в постоянном месте (рядом с приложением)
    mapping_file_path = os.path.join(app_dir, "Типы Нарядов маппинг.xlsx")

    rates_file_path = os.path.join(app_dir, "Ставки механиков.xlsx")

    if not mechanics_file_path:
        st.warning("Файл Механики.xlsx не найден. Загрузите файл выше или положите его в папку с данными.")
    else:
        st.success(f"Файл: `{os.path.basename(mechanics_file_path)}`")
        if os.path.exists(mapping_file_path):
            st.success(f"Маппинг (фиксированный): `{os.path.basename(mapping_file_path)}`")

            # Отслеживаем обновления маппинга и сбрасываем устаревшие результаты
            map_signature = (os.path.getmtime(mapping_file_path), os.path.getsize(mapping_file_path))
            prev_signature = st.session_state.get("mapping_signature")
            if prev_signature and prev_signature != map_signature:
                for _k in ["mech_detail_df", "mech_summary_df", "mech_records", "mech_work_type_mapping"]:
                    st.session_state.pop(_k, None)
                st.info("Файл маппинга изменился. Старый расчет очищен, нажмите 'Рассчитать бонусы механиков'.")
            st.session_state["mapping_signature"] = map_signature
        else:
            st.error("Не найден фиксированный файл Типы Нарядов маппинг.xlsx рядом с приложением. Добавьте его в папку проекта.")

        if os.path.exists(rates_file_path):
            st.success(f"Ставки механиков: `{os.path.basename(rates_file_path)}`")
        else:
            st.info("Файл ставок пока не найден. Ожидаем файл 'Ставки механиков.xlsx' в папке проекта.")

        fuzzy_threshold = st.slider(
            "Порог схожести ФИО (для сопоставления со ставками)",
            min_value=0.0, max_value=1.0, value=0.78, step=0.01,
            help="Чем ниже порог, тем больше ФИО будут сопоставлены, но выше риск ошибок. 0.78 - хороший баланс."
        )
        # --- Ставки (пока фиксированные, позже — редактируемые) ---
        with st.expander("⚙️ Ставки по категориям (тг/Н/Ч)", expanded=False):
            col_r1, col_r2, col_r3 = st.columns(3)
            rate_warranty = col_r1.number_input("Гарантийные", value=3915, step=1, key="rate_warranty")
            rate_internal = col_r2.number_input("Внутренние ОП/Сервис + ППП", value=1895, step=1, key="rate_internal")
            col_r3.info("**Сервис / Доп сервис (коммерческие)**\nСтавка берётся индивидуально из файла «Ставки механиков.xlsx» для каждого механика.")

        user_rates = {
            "warranty": float(rate_warranty),
            "internal": float(rate_internal),
            "commercial": 0.0,
            "exclude": 0.0,
        }

        if st.button("🔄 Рассчитать бонусы механиков", key="calc_mechanics"):
            with st.spinner("Загружаем данные..."):
                try:
                    calc = MechanicsCalculator()
                    records = calc.load_from_mechanics_xlsx(mechanics_file_path)
                    work_type_mapping = None
                    work_type_labels = None
                    original_mapped_types = set()
                    approx_mapping_df = pd.DataFrame()

                    if os.path.exists(mapping_file_path):
                        work_type_mapping, work_type_labels = calc.load_work_type_mapping_xlsx(mapping_file_path)
                        original_mapped_types = set(work_type_mapping.keys()) if work_type_mapping else set()

                    # Автоприближение маппинга: если тип не найден в файле,
                    # пробуем сопоставить с наиболее похожим типом из маппинга.
                    if records and work_type_mapping:
                        all_types = sorted({str(r.get("тип_наряда", "")).strip() for r in records if r.get("тип_наряда")})
                        base_mapping = dict(work_type_mapping)
                        base_labels = dict(work_type_labels or {})
                        normalized_base = {}
                        for k in base_mapping.keys():
                            norm_k = re.sub(r"[^a-zа-я0-9]+", " ", str(k).lower().replace("ё", "е"))
                            norm_k = re.sub(r"\s+", " ", norm_k).strip()
                            normalized_base[k] = norm_k

                        approx_rows = []
                        similarity_threshold = 0.74

                        for raw_type in all_types:
                            if raw_type in base_mapping:
                                continue

                            raw_norm = re.sub(r"[^a-zа-я0-9]+", " ", str(raw_type).lower().replace("ё", "е"))
                            raw_norm = re.sub(r"\s+", " ", raw_norm).strip()

                            # Бизнес-правило: любые гарантийные работы всегда в warranty.
                            if "гарант" in raw_norm:
                                base_mapping[raw_type] = "warranty"
                                base_labels[raw_type] = "Гарантия (авто-правило)"
                                approx_rows.append(
                                    {
                                        "Новый тип": raw_type,
                                        "Похожий тип из маппинга": "(авто-правило: гарантия)",
                                        "Сходство": 1.0,
                                        "Категория": "warranty",
                                        "Группа маппинга": "Гарантия (авто-правило)",
                                    }
                                )
                                continue

                            best_key = ""
                            best_score = 0.0

                            for candidate, candidate_norm in normalized_base.items():
                                score = SequenceMatcher(None, raw_norm, candidate_norm).ratio()
                                if score > best_score:
                                    best_score = score
                                    best_key = candidate

                            if best_key and best_score >= similarity_threshold:
                                base_mapping[raw_type] = work_type_mapping[best_key]
                                base_labels[raw_type] = (work_type_labels or {}).get(best_key, "")
                                approx_rows.append(
                                    {
                                        "Новый тип": raw_type,
                                        "Похожий тип из маппинга": best_key,
                                        "Сходство": round(best_score, 3),
                                        "Категория": base_mapping[raw_type],
                                        "Группа маппинга": base_labels[raw_type],
                                    }
                                )

                        work_type_mapping = base_mapping
                        work_type_labels = base_labels
                        approx_mapping_df = pd.DataFrame(approx_rows)

                    if not records:
                        st.error("Не удалось извлечь данные из файла. Проверьте структуру файла.")
                        with st.expander("🔎 Диагностика структуры Механики.xlsx", expanded=True):
                            st.write(f"Файл: {mechanics_file_path}")
                            try:
                                import openpyxl
                                wb_dbg = openpyxl.load_workbook(mechanics_file_path, data_only=True)
                                ws_dbg = wb_dbg.active

                                indent_stats = {}
                                non_empty_a = 0
                                for row_dbg in ws_dbg.iter_rows(min_row=1, max_row=min(ws_dbg.max_row, 3000)):
                                    if not row_dbg:
                                        continue
                                    cell_a_dbg = row_dbg[0]
                                    val_dbg = cell_a_dbg.value
                                    if val_dbg is None or str(val_dbg).strip() == "":
                                        continue
                                    non_empty_a += 1
                                    indent_dbg = 0
                                    if cell_a_dbg.alignment and cell_a_dbg.alignment.indent is not None:
                                        indent_dbg = cell_a_dbg.alignment.indent
                                    indent_stats[indent_dbg] = indent_stats.get(indent_dbg, 0) + 1

                                st.write(f"Непустых строк в колонке A (первые 3000): {non_empty_a}")
                                st.write(f"Распределение отступов: {indent_stats}")

                                preview_rows = []
                                for i, row_dbg in enumerate(ws_dbg.iter_rows(min_row=1, max_row=20), start=1):
                                    a = row_dbg[0].value if len(row_dbg) > 0 else None
                                    f = row_dbg[5].value if len(row_dbg) > 5 else None
                                    g = row_dbg[6].value if len(row_dbg) > 6 else None
                                    if a is None and f is None and g is None:
                                        continue
                                    preview_rows.append({
                                        "Строка": i,
                                        "A": str(a) if a is not None else "",
                                        "F": str(f) if f is not None else "",
                                        "G(Н/Ч?)": g if g is not None else "",
                                    })
                                if preview_rows:
                                    st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)
                            except Exception as dbg_e:
                                st.code(f"Ошибка доп. диагностики: {dbg_e}")
                    else:
                        st.session_state["mech_records"] = records
                        st.session_state["mech_work_type_mapping"] = work_type_mapping or {}
                        st.session_state["mech_original_mapped_types"] = sorted(original_mapped_types)
                        st.session_state["mech_approx_mapping_df"] = approx_mapping_df
                        detail_df, summary_df = calc.build_summary_from_records(
                            records,
                            work_type_mapping=work_type_mapping,
                            work_type_labels=work_type_labels,
                            category_rates=user_rates,
                        )

                        # Если есть файл ставок, пробуем fuzzy-сопоставление ФИО
                        rates_df = pd.DataFrame()
                        rates_match_df = pd.DataFrame()
                        rates_mismatch_df = pd.DataFrame()
                        if os.path.exists(rates_file_path) and not summary_df.empty: # Добавлена проверка на пустой summary_df
                            rates_df = calc.load_rates_workbook(rates_file_path)
                            rates_match_df, rates_mismatch_df = calc.apply_rates_with_fuzzy_matching(summary_df, rates_df, fuzzy_threshold=fuzzy_threshold)

                        # Применяем персональные ставки из файла ставок для коммерческих З/Н.
                        if not rates_match_df.empty and not detail_df.empty:
                            matched_rates = rates_match_df[
                                rates_match_df["Статус сопоставления"].isin(["exact", "fuzzy"])
                            ][["ФИО", "ДЦ", "ДЦ в ставках", "Ставка (ставки)", "Должность (ставки)"]].copy()
                            matched_rates = matched_rates.drop_duplicates(subset=["ФИО", "ДЦ"], keep="first")

                            detail_df = detail_df.merge(
                                matched_rates,
                                on=["ФИО", "ДЦ"],
                                how="left",
                            )

                            # Для отображения используем ДЦ из файла ставок, если он найден.
                            detail_df["ДЦ"] = np.where(
                                detail_df["ДЦ в ставках"].notna() & (detail_df["ДЦ в ставках"].astype(str).str.strip() != ""),
                                detail_df["ДЦ в ставках"],
                                detail_df["ДЦ"],
                            )

                            unresolved_dc_mask = (
                                detail_df["ДЦ в ставках"].isna()
                                | (detail_df["ДЦ в ставках"].astype(str).str.strip() == "")
                            ) & (
                                detail_df["ДЦ"].astype(str).str.lower().str.startswith("too")
                                | detail_df["ДЦ"].astype(str).str.lower().str.startswith("тоо")
                                | detail_df["ДЦ"].astype(str).str.lower().str.contains("эвакуатор")
                            )
                            detail_df.loc[unresolved_dc_mask, "ДЦ"] = "Не сопоставлен в ставках"

                            commercial_mask = (
                                (detail_df["Категория"] == "commercial")
                                & detail_df["Ставка (ставки)"].notna()
                                & (detail_df["Ставка (ставки)"] > 0)
                            )
                            detail_df.loc[commercial_mask, "Ставка"] = detail_df.loc[commercial_mask, "Ставка (ставки)"]

                            # Пересчитываем начисление по финальной ставке.
                            detail_df["Начисление"] = (detail_df["Н/Ч"] * detail_df["Ставка"]).round(2)

                            detail_df = detail_df.drop(columns=["ДЦ в ставках", "Ставка (ставки)"], errors="ignore")
                            summary_df = (
                                detail_df.groupby(["ФИО", "ДЦ"], as_index=False)
                                .agg({"Н/Ч": "sum", "Начисление": "sum"})
                                .rename(columns={"Н/Ч": "Итого Н/Ч", "Начисление": "Итого начисление"})
                                .sort_values(["ДЦ", "ФИО"])
                            )
                            summary_df["Итого Н/Ч"] = summary_df["Итого Н/Ч"].round(2)
                            summary_df["Итого начисление"] = summary_df["Итого начисление"].round(2)

                            # Синхронизируем таблицу сопоставления с финальным итогом по механикам.
                            if not rates_df.empty:
                                rates_match_df, rates_mismatch_df = calc.apply_rates_with_fuzzy_matching(
                                    summary_df,
                                    rates_df,
                                    fuzzy_threshold=fuzzy_threshold,
                                )

                        st.session_state["mech_detail_df"] = detail_df
                        st.session_state["mech_summary_df"] = summary_df
                        st.session_state["mech_rates_df"] = rates_df
                        st.session_state["mech_rates_match_df"] = rates_match_df
                        st.session_state["mech_rates_mismatch_df"] = rates_mismatch_df
                        st.success(f"Загружено записей: {len(records)} (механиков: {summary_df.shape[0]})")
                except Exception as _e:
                    st.error(f"Ошибка при загрузке: {_e}")
                    show_file_read_diagnostics(
                        _e,
                        data_folder=app_dir,
                        required_files=["Механики.xlsx", "Типы Нарядов маппинг.xlsx", "Ставки механиков.xlsx"],
                    )
                    import traceback
                    st.code(traceback.format_exc())

        # --- Вывод результатов ---
        if "mech_summary_df" in st.session_state and st.session_state["mech_summary_df"] is not None:
            detail_df = st.session_state.get("mech_detail_df", pd.DataFrame())
            summary_df = st.session_state["mech_summary_df"]
            dc_values = sorted([str(x).strip() for x in detail_df.get("ДЦ", pd.Series(dtype=str)).dropna().unique().tolist() if str(x).strip()])
            dc_options = ["— все ДЦ —"] + dc_values
            selected_dc = st.selectbox("🏢 Фильтр по ДЦ", dc_options, key="mech_dc_filter")

            if selected_dc != "— все ДЦ —":
                detail_scope = detail_df[detail_df["ДЦ"].astype(str) == selected_dc].copy()
                summary_scope = summary_df[summary_df["ДЦ"].astype(str) == selected_dc].copy()
            else:
                detail_scope = detail_df.copy()
                summary_scope = summary_df.copy()

            st.subheader("📋 Итог по механикам")
            st.dataframe(
                summary_scope.style.format({"Итого Н/Ч": "{:.2f}", "Итого начисление": "{:,.0f}", "Реализация": "{:,.0f}"}),
                use_container_width=True,
            )

            export_detail_quick = detail_scope.copy() if not detail_scope.empty else detail_df.copy()
            export_summary_quick = summary_scope.copy() if not summary_scope.empty else summary_df.copy()
            if not export_summary_quick.empty:
                quick_excel_data = build_mechanics_cards_export(export_detail_quick, export_summary_quick)
                st.download_button(
                    "⬇️ Скачать карточки и итог (быстро)",
                    data=quick_excel_data,
                    file_name="механики_карточки_и_итог.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="mech_download_quick",
                )

            # --- Разбивка по 5 группам маппинга ---
            if not detail_scope.empty and "Группа маппинга" in detail_scope.columns:
                with st.expander("📊 Норма-часы и начисление по группам (5 строк)", expanded=True):
                    _grp_pivot = (
                        detail_scope.groupby(["ФИО", "ДЦ", "Группа маппинга"], as_index=False)
                        .agg(**{"Н/Ч": ("Н/Ч", "sum"), "Начисление": ("Начисление", "sum")})
                    )
                    _all_groups = sorted(_grp_pivot["Группа маппинга"].dropna().unique())
                    _pivot_nh = _grp_pivot.pivot_table(
                        index=["ФИО", "ДЦ"],
                        columns="Группа маппинга",
                        values="Н/Ч",
                        aggfunc="sum",
                        fill_value=0,
                    ).reset_index()
                    _pivot_sum = _grp_pivot.pivot_table(
                        index=["ФИО", "ДЦ"],
                        columns="Группа маппинга",
                        values="Начисление",
                        aggfunc="sum",
                        fill_value=0,
                    ).reset_index()
                    # Переименуем колонки групп с суффиксами
                    _nh_rename = {g: f"{g} (н/ч)" for g in _all_groups if g in _pivot_nh.columns}
                    _sum_rename = {g: f"{g} (₸)" for g in _all_groups if g in _pivot_sum.columns}
                    _pivot_nh = _pivot_nh.rename(columns=_nh_rename)
                    _pivot_sum = _pivot_sum.rename(columns=_sum_rename)
                    _group_breakdown = _pivot_nh.merge(_pivot_sum, on=["ФИО", "ДЦ"], how="outer")
                    # Порядок колонок: ФИО, ДЦ, затем чередуем н/ч / ₸ по каждой группе
                    _ordered_cols = ["ФИО", "ДЦ"]
                    for _g in _all_groups:
                        if f"{_g} (н/ч)" in _group_breakdown.columns:
                            _ordered_cols.append(f"{_g} (н/ч)")
                        if f"{_g} (₸)" in _group_breakdown.columns:
                            _ordered_cols.append(f"{_g} (₸)")
                    _group_breakdown = _group_breakdown[[c for c in _ordered_cols if c in _group_breakdown.columns]]
                    _fmt_dict = {c: ("{:.2f}" if "(н/ч)" in c else "{:,.0f}") for c in _group_breakdown.columns if c not in ["ФИО", "ДЦ"]}
                    st.dataframe(
                        _group_breakdown.style.format(_fmt_dict),
                        use_container_width=True,
                    )

            rates_match_df = st.session_state.get("mech_rates_match_df", pd.DataFrame())
            rates_mismatch_df = st.session_state.get("mech_rates_mismatch_df", pd.DataFrame())
            if selected_dc != "— все ДЦ —" and not rates_match_df.empty:
                dc_match_col = "ДЦ в ставках" if "ДЦ в ставках" in rates_match_df.columns else "ДЦ"
                rates_match_scope = rates_match_df[rates_match_df[dc_match_col].astype(str) == selected_dc].copy()
                rates_mismatch_scope = rates_mismatch_df[rates_mismatch_df[dc_match_col].astype(str) == selected_dc].copy() if not rates_mismatch_df.empty else rates_mismatch_df
            else:
                rates_match_scope = rates_match_df
                rates_mismatch_scope = rates_mismatch_df

            if not rates_match_scope.empty:
                st.subheader("🧩 Сопоставление с файлом ставок")
                # Если в исходном отчете мало ДЦ, показываем управленческую логику источника ДЦ.
                if "ДЦ" in summary_scope.columns and summary_scope["ДЦ"].nunique() <= 2 and "ДЦ в ставках" in rates_match_scope.columns:
                    st.info("В исходном файле Механики ДЦ агрегирован и может показываться как 1-2 значения. Для аналитики по городам используем 'ДЦ в ставках' из файла ставок.")
                st.dataframe(
                    rates_match_scope.style.format({"Совпадение": "{:.3f}", "Оклад (ставки)": "{:,.0f}", "Ставка (ставки)": "{:,.0f}"}),
                    use_container_width=True,
                )

            if not rates_mismatch_scope.empty:
                st.warning(f"Есть ФИО без уверенного сопоставления: {len(rates_mismatch_scope)}")
                with st.expander("⚠️ Проверить ФИО с сильным расхождением"):
                    mismatch_cols = ["ФИО", "ДЦ", "ФИО в ставках", "ДЦ в ставках", "Совпадение", "Статус сопоставления"]
                    mismatch_cols = [c for c in mismatch_cols if c in rates_mismatch_scope.columns]
                    mismatch_view = rates_mismatch_scope[mismatch_cols].copy()
                    mismatch_view["Статус сопоставления"] = mismatch_view["Статус сопоставления"].replace(
                        {"unmatched": "нет уверенного совпадения", "fuzzy": "похожее совпадение", "exact": "точное совпадение"}
                    )
                    st.dataframe(mismatch_view, use_container_width=True)

            # --- Механики без выработки ---
            _rates_df_diag = st.session_state.get("mech_rates_df", pd.DataFrame())
            if not _rates_df_diag.empty and not rates_match_scope.empty:
                if selected_dc != "— все ДЦ —" and "ДЦ" in _rates_df_diag.columns:
                    _rates_df_diag = _rates_df_diag[_rates_df_diag["ДЦ"].astype(str) == selected_dc].copy()
                _matched_in_rates = set(
                    rates_match_scope[rates_match_scope["Статус сопоставления"].isin(["exact", "fuzzy"])]
                    ["ФИО в ставках"].dropna().astype(str).str.strip()
                )
                _rate_fio_col = "ФИО" if "ФИО" in _rates_df_diag.columns else _rates_df_diag.columns[0]
                _no_output_df = _rates_df_diag[
                    ~_rates_df_diag[_rate_fio_col].astype(str).str.strip().isin(_matched_in_rates)
                ].copy()
                _keep_cols = [c for c in ["ФИО", "ДЦ", "Должность", "Ставка", "Оклад"] if c in _no_output_df.columns]
                _no_output_df = _no_output_df[_keep_cols].rename(columns={"ФИО": "ФИО (из ставок)"})
                _no_output_df = _no_output_df.sort_values(["ДЦ", "ФИО (из ставок)"] if "ДЦ" in _no_output_df.columns else ["ФИО (из ставок)"])

                st.subheader(f"🚫 Механики без выработки: {len(_no_output_df)}")
                if not _no_output_df.empty:
                    st.warning(
                        f"Эти {len(_no_output_df)} механиков есть в файле «Ставки механиков.xlsx», "
                        "но не найдены в выгрузке нарядов (нет записей в Механики.xlsx)."
                    )
                    _fmt_cols = {c: "{:,.0f}" for c in ["Ставка", "Оклад"] if c in _no_output_df.columns}
                    st.dataframe(
                        _no_output_df.style.format(_fmt_cols) if _fmt_cols else _no_output_df,
                        use_container_width=True,
                    )
                else:
                    st.success("✅ Все механики из файла ставок имеют записи в выгрузке нарядов.")

            # --- Аналитика для руководителя (вся группа) ---
            if not detail_scope.empty:
                st.subheader("📈 Аналитика для руководителя (вся группа)")

                category_labels = {
                    "warranty": "Гарантийные",
                    "internal": "Внутренние / ППП",
                    "commercial": "Коммерческие",
                    "exclude": "Исключено",
                }
                category_colors = {
                    "Гарантийные": "#2E8B57",
                    "Внутренние / ППП": "#F28C28",
                    "Коммерческие": "#1F77B4",
                    "Исключено": "#8C8C8C",
                }

                def _detect_brand(raw_type: str) -> str:
                    txt = str(raw_type).lower()
                    brand_map = [
                        ("li xiang", "Li Xiang"),
                        ("lixiang", "Li Xiang"),
                        ("chevrolet", "Chevrolet"),
                        ("kia", "KIA"),
                        ("jac", "JAC"),
                        ("jetour", "Jetour"),
                        ("hongqi", "Hongqi"),
                        ("rox", "ROX"),
                        ("skoda", "Skoda"),
                        ("мультибренд", "Мультибренд"),
                    ]
                    for key, name in brand_map:
                        if key in txt:
                            return name
                    return "Прочее"

                manager_df = detail_scope.copy()
                manager_df["Категория"] = manager_df["Категория"].map(category_labels).fillna(manager_df["Категория"])
                manager_df["Бренд"] = manager_df["Тип наряда"].apply(_detect_brand)

                brand_category_df = (
                    manager_df.groupby(["Бренд", "Категория"], as_index=False)
                    .agg({"Н/Ч": "sum", "Начисление": "sum"})
                    .sort_values(["Бренд", "Категория"])
                )
                brand_category_df["Н/Ч"] = brand_category_df["Н/Ч"].round(2)
                brand_category_df["Начисление"] = brand_category_df["Начисление"].round(2)

                category_total_df = (
                    manager_df.groupby("Категория", as_index=False)
                    .agg({"Н/Ч": "sum", "Начисление": "sum"})
                    .sort_values("Начисление", ascending=False)
                )
                category_total_df["Доля выплат, %"] = np.where(
                    category_total_df["Начисление"].sum() > 0,
                    (category_total_df["Начисление"] / category_total_df["Начисление"].sum()) * 100,
                    0,
                )
                category_total_df["Н/Ч"] = category_total_df["Н/Ч"].round(2)
                category_total_df["Начисление"] = category_total_df["Начисление"].round(2)
                category_total_df["Доля выплат, %"] = category_total_df["Доля выплат, %"].round(2)

                brand_total_df = (
                    manager_df.groupby("Бренд", as_index=False)
                    .agg({"Н/Ч": "sum", "Начисление": "sum"})
                    .sort_values("Начисление", ascending=False)
                )
                brand_total_df["Н/Ч"] = brand_total_df["Н/Ч"].round(2)
                brand_total_df["Начисление"] = brand_total_df["Начисление"].round(2)

                pivot_brand_category = (
                    brand_category_df.pivot(index="Бренд", columns="Категория", values="Начисление")
                    .fillna(0)
                    .reset_index()
                )

                total_payout = float(manager_df["Начисление"].sum())
                total_hours = float(manager_df["Н/Ч"].sum())
                rates_df_local = st.session_state.get("mech_rates_df", pd.DataFrame())
                rates_match_local = st.session_state.get("mech_rates_match_df", pd.DataFrame())

                # Для статистики по ставкам/окладам берем только тех, кто реально попал в расчет
                # (есть в отчете и сопоставлен exact/fuzzy).
                rates_for_stats = rates_df_local.copy()
                if (
                    not rates_for_stats.empty
                    and not rates_match_local.empty
                    and {"ФИО", "ДЦ"}.issubset(rates_for_stats.columns)
                    and {"ФИО в ставках", "ДЦ в ставках", "Статус сопоставления"}.issubset(rates_match_local.columns)
                ):
                    matched_people = (
                        rates_match_local[rates_match_local["Статус сопоставления"].isin(["exact", "fuzzy"])][["ФИО в ставках", "ДЦ в ставках"]]
                        .dropna()
                        .drop_duplicates()
                        .rename(columns={"ФИО в ставках": "ФИО", "ДЦ в ставках": "ДЦ"})
                    )
                    if not matched_people.empty:
                        matched_people["ФИО"] = matched_people["ФИО"].astype(str).str.replace(r"\s+", " ", regex=True).str.strip()
                        matched_people["ДЦ"] = matched_people["ДЦ"].astype(str).str.strip()

                        rates_for_stats["ФИО"] = rates_for_stats["ФИО"].astype(str).str.replace(r"\s+", " ", regex=True).str.strip()
                        rates_for_stats["ДЦ"] = rates_for_stats["ДЦ"].astype(str).str.strip()

                        rates_for_stats = rates_for_stats.merge(
                            matched_people,
                            on=["ФИО", "ДЦ"],
                            how="inner",
                        )
                _commercial_df = manager_df[manager_df["Категория"] == "commercial"] if "Категория" in manager_df.columns else pd.DataFrame()
                _commercial_payout = float(_commercial_df["Начисление"].sum()) if not _commercial_df.empty else 0.0
                _commercial_hours = float(_commercial_df["Н/Ч"].sum()) if not _commercial_df.empty else 0.0
                _commercial_share = (_commercial_payout / total_payout * 100) if total_payout > 0 else 0.0
                _total_реализация = float(manager_df["Реализация"].sum()) if "Реализация" in manager_df.columns else 0.0

                col_m1, col_m2, col_m3, col_m4, col_m5 = st.columns(5)
                col_m1.metric("Общая выплата по группе", f"{total_payout:,.0f} ₸")
                col_m2.metric("Общие Н/Ч по группе", f"{total_hours:,.2f}")
                col_m3.metric("Реализовано работ (сумма)", f"{_total_реализация:,.0f} ₸")
                col_m4.metric("Выплата по коммерции", f"{_commercial_payout:,.0f} ₸", f"{_commercial_share:.1f}% от общей")
                col_m5.metric("Н/Ч по коммерции", f"{_commercial_hours:,.2f}")

                chart_col1, chart_col2 = st.columns(2)

                with chart_col1:
                    category_chart_df = category_total_df[category_total_df["Начисление"] > 0].copy()
                    if not category_chart_df.empty:
                        fig_category_share = px.pie(
                            category_chart_df,
                            names="Категория",
                            values="Начисление",
                            hole=0.55,
                            color="Категория",
                            color_discrete_map=category_colors,
                        )
                        fig_category_share.update_traces(
                            textposition="inside",
                            textinfo="percent+label",
                            hovertemplate="Категория: %{label}<br>Начисление: %{value:,.0f}<br>Доля: %{percent}<extra></extra>",
                        )
                        fig_category_share.update_layout(
                            title="Структура выплат по категориям",
                            margin=dict(l=10, r=10, t=50, b=10),
                            legend_title_text="",
                        )
                        st.plotly_chart(fig_category_share, use_container_width=True)

                with chart_col2:
                    top_brand_chart_df = brand_total_df.head(8).sort_values("Начисление", ascending=True)
                    if not top_brand_chart_df.empty:
                        fig_brand_total = px.bar(
                            top_brand_chart_df,
                            x="Начисление",
                            y="Бренд",
                            orientation="h",
                            text="Начисление",
                            color="Начисление",
                            color_continuous_scale=[[0, "#CFE8FF"], [1, "#1F77B4"]],
                        )
                        fig_brand_total.update_traces(
                            texttemplate="%{text:,.0f}",
                            textposition="outside",
                            hovertemplate="Бренд: %{y}<br>Начисление: %{x:,.0f}<extra></extra>",
                        )
                        fig_brand_total.update_layout(
                            title="Топ брендов по выплатам",
                            margin=dict(l=10, r=10, t=50, b=10),
                            coloraxis_showscale=False,
                            xaxis_title="Начисление, тг",
                            yaxis_title="",
                        )
                        st.plotly_chart(fig_brand_total, use_container_width=True)

                stacked_chart_df = brand_category_df[brand_category_df["Начисление"] > 0].copy()
                if not stacked_chart_df.empty:
                    fig_brand_category = px.bar(
                        stacked_chart_df,
                        x="Бренд",
                        y="Начисление",
                        color="Категория",
                        barmode="stack",
                        color_discrete_map=category_colors,
                        text="Начисление",
                    )
                    fig_brand_category.update_traces(
                        texttemplate="%{text:,.0f}",
                        textposition="inside",
                        hovertemplate="Бренд: %{x}<br>Категория: %{legendgroup}<br>Начисление: %{y:,.0f}<extra></extra>",
                    )
                    fig_brand_category.update_layout(
                        title="Выплаты по брендам с разбивкой по категориям",
                        margin=dict(l=10, r=10, t=50, b=10),
                        xaxis_title="",
                        yaxis_title="Начисление, тг",
                        legend_title_text="",
                    )
                    st.plotly_chart(fig_brand_category, use_container_width=True)

                efficiency_df = brand_category_df.groupby("Бренд", as_index=False).agg(
                    {"Н/Ч": "sum", "Начисление": "sum"}
                )
                efficiency_df = efficiency_df[(efficiency_df["Н/Ч"] > 0) & (efficiency_df["Начисление"] > 0)].copy()
                if not efficiency_df.empty:
                    efficiency_df["Средняя ставка за Н/Ч"] = (efficiency_df["Начисление"] / efficiency_df["Н/Ч"]).round(0)
                    fig_efficiency = px.scatter(
                        efficiency_df,
                        x="Н/Ч",
                        y="Начисление",
                        size="Средняя ставка за Н/Ч",
                        color="Бренд",
                        text="Бренд",
                        size_max=55,
                    )
                    fig_efficiency.update_traces(
                        textposition="top center",
                        hovertemplate=(
                            "Бренд: %{text}<br>Н/Ч: %{x:.2f}<br>Начисление: %{y:,.0f}"
                            "<br>Средняя ставка за Н/Ч: %{marker.size:,.0f}<extra></extra>"
                        ),
                    )
                    fig_efficiency.update_layout(
                        title="Карта эффективности брендов",
                        margin=dict(l=10, r=10, t=50, b=10),
                        xaxis_title="Н/Ч",
                        yaxis_title="Начисление, тг",
                        legend_title_text="",
                    )
                    st.plotly_chart(fig_efficiency, use_container_width=True)

                top_mechanics_df = (
                    summary_scope.sort_values(["Итого начисление", "Итого Н/Ч"], ascending=[False, False])
                    .head(10)
                    .copy()
                )
                if not top_mechanics_df.empty:
                    st.markdown("**Топ-10 механиков по начислению**")
                    fig_top_mechanics = px.bar(
                        top_mechanics_df.sort_values("Итого начисление", ascending=True),
                        x="Итого начисление",
                        y="ФИО",
                        orientation="h",
                        color="ДЦ",
                        text="Итого начисление",
                    )
                    fig_top_mechanics.update_traces(
                        texttemplate="%{text:,.0f}",
                        textposition="outside",
                        hovertemplate="Механик: %{y}<br>ДЦ: %{legendgroup}<br>Начисление: %{x:,.0f}<extra></extra>",
                    )
                    fig_top_mechanics.update_layout(
                        margin=dict(l=10, r=10, t=30, b=10),
                        xaxis_title="Начисление, тг",
                        yaxis_title="",
                        legend_title_text="ДЦ",
                    )
                    st.plotly_chart(fig_top_mechanics, use_container_width=True)
                    st.dataframe(
                        top_mechanics_df.style.format({"Итого Н/Ч": "{:.2f}", "Итого начисление": "{:,.0f}"}),
                        use_container_width=True,
                    )
                st.session_state["mech_top_mechanics_df"] = top_mechanics_df

                dc_salary_compare_df = pd.DataFrame()
                dc_salary_fund_df = pd.DataFrame()
                if not rates_for_stats.empty and {"ДЦ", "Оклад", "Ставка"}.issubset(rates_for_stats.columns):
                    _stats_src = rates_for_stats.copy()
                    _stats_src["Оклад"] = pd.to_numeric(_stats_src["Оклад"], errors="coerce")
                    _stats_src["Ставка"] = pd.to_numeric(_stats_src["Ставка"], errors="coerce")
                    _salary_src = _stats_src[_stats_src["Оклад"] > 0].copy()

                    dc_salary_compare_df = (
                        _stats_src.groupby("ДЦ", as_index=False)
                        .agg(
                            Количество_механиков=("ФИО", "count"),
                            Средний_оклад=("Оклад", lambda s: s[s > 0].mean()),
                            Медианный_оклад=("Оклад", lambda s: s[s > 0].median()),
                            Мин_оклад=("Оклад", lambda s: s[s > 0].min()),
                            Макс_оклад=("Оклад", lambda s: s[s > 0].max()),
                            Средняя_ставка=("Ставка", lambda s: s[s > 0].mean()),
                        )
                        .sort_values("Средний_оклад", ascending=False)
                    )
                    dc_salary_compare_df[["Средний_оклад", "Медианный_оклад", "Мин_оклад", "Макс_оклад", "Средняя_ставка"]] = (
                        dc_salary_compare_df[["Средний_оклад", "Медианный_оклад", "Мин_оклад", "Макс_оклад", "Средняя_ставка"]].round(0)
                    )

                    if not _salary_src.empty:
                        dc_salary_fund_df = (
                            _salary_src.groupby("ДЦ", as_index=False)
                            .agg(
                                Количество_с_окладом=("ФИО", "count"),
                                Фонд_окладов=("Оклад", "sum"),
                                Средний_оклад=("Оклад", "mean"),
                                Медианный_оклад=("Оклад", "median"),
                                Мин_оклад=("Оклад", "min"),
                                Макс_оклад=("Оклад", "max"),
                            )
                            .sort_values("Фонд_окладов", ascending=False)
                        )
                        total_salary_fund = dc_salary_fund_df["Фонд_окладов"].sum()
                        dc_salary_fund_df["Доля фонда, %"] = np.where(
                            total_salary_fund > 0,
                            dc_salary_fund_df["Фонд_окладов"] / total_salary_fund * 100,
                            0,
                        )
                        for _c in ["Фонд_окладов", "Средний_оклад", "Медианный_оклад", "Мин_оклад", "Макс_оклад"]:
                            dc_salary_fund_df[_c] = dc_salary_fund_df[_c].round(0)
                        dc_salary_fund_df["Доля фонда, %"] = dc_salary_fund_df["Доля фонда, %"].round(2)

                if not dc_salary_compare_df.empty:
                    st.markdown("**Сравнение окладов и ставок по ДЦ**")
                    salary_chart_col1, salary_chart_col2 = st.columns(2)

                    with salary_chart_col1:
                        fig_dc_salary = px.bar(
                            dc_salary_compare_df.sort_values("Средний_оклад", ascending=True),
                            x="Средний_оклад",
                            y="ДЦ",
                            orientation="h",
                            text="Средний_оклад",
                            color="Средний_оклад",
                            color_continuous_scale=[[0, "#F6D6AD"], [1, "#D97706"]],
                        )
                        fig_dc_salary.update_traces(
                            texttemplate="%{text:,.0f}",
                            textposition="outside",
                            hovertemplate="ДЦ: %{y}<br>Средний оклад: %{x:,.0f}<extra></extra>",
                        )
                        fig_dc_salary.update_layout(
                            title="Где оклад выше / ниже",
                            margin=dict(l=10, r=10, t=50, b=10),
                            coloraxis_showscale=False,
                            xaxis_title="Средний оклад, тг",
                            yaxis_title="",
                        )
                        st.plotly_chart(fig_dc_salary, use_container_width=True)

                    with salary_chart_col2:
                        fig_dc_rate = px.bar(
                            dc_salary_compare_df.sort_values("Средняя_ставка", ascending=True),
                            x="Средняя_ставка",
                            y="ДЦ",
                            orientation="h",
                            text="Средняя_ставка",
                            color="Средняя_ставка",
                            color_continuous_scale=[[0, "#D7ECFF"], [1, "#2563EB"]],
                        )
                        fig_dc_rate.update_traces(
                            texttemplate="%{text:,.0f}",
                            textposition="outside",
                            hovertemplate="ДЦ: %{y}<br>Средняя ставка: %{x:,.0f}<extra></extra>",
                        )
                        fig_dc_rate.update_layout(
                            title="Средняя ставка по ДЦ",
                            margin=dict(l=10, r=10, t=50, b=10),
                            coloraxis_showscale=False,
                            xaxis_title="Средняя ставка, тг",
                            yaxis_title="",
                        )
                        st.plotly_chart(fig_dc_rate, use_container_width=True)

                    st.dataframe(
                        dc_salary_compare_df.rename(columns={
                            "Количество_механиков": "Кол-во механиков",
                            "Средний_оклад": "Средний оклад",
                            "Медианный_оклад": "Медианный оклад",
                            "Мин_оклад": "Мин. оклад",
                            "Макс_оклад": "Макс. оклад",
                            "Средняя_ставка": "Средняя ставка",
                        }).style.format({
                            "Средний оклад": "{:,.0f}",
                            "Медианный оклад": "{:,.0f}",
                            "Мин. оклад": "{:,.0f}",
                            "Макс. оклад": "{:,.0f}",
                            "Средняя ставка": "{:,.0f}",
                        }),
                        use_container_width=True,
                    )

                    if not dc_salary_fund_df.empty:
                        st.markdown("**Фонд окладов по ДЦ**")
                        st.dataframe(
                            dc_salary_fund_df.rename(columns={
                                "Количество_с_окладом": "Кол-во с окладом",
                                "Фонд_окладов": "Сумма окладов",
                                "Средний_оклад": "Средний оклад",
                                "Медианный_оклад": "Медианный оклад",
                                "Мин_оклад": "Мин. оклад",
                                "Макс_оклад": "Макс. оклад",
                            }).style.format({
                                "Сумма окладов": "{:,.0f}",
                                "Средний оклад": "{:,.0f}",
                                "Медианный оклад": "{:,.0f}",
                                "Мин. оклад": "{:,.0f}",
                                "Макс. оклад": "{:,.0f}",
                                "Доля фонда, %": "{:.2f}%",
                            }),
                            use_container_width=True,
                        )
                st.session_state["mech_dc_salary_compare_df"] = dc_salary_compare_df
                st.session_state["mech_dc_salary_fund_df"] = dc_salary_fund_df

                # Полноценная аналитика по ДЦ
                dc_analytics_df = (
                    detail_df.groupby("ДЦ", as_index=False)
                    .agg({
                        "ФИО": "nunique",
                        "Н/Ч": "sum",
                        "Начисление": "sum",
                        **( {"Реализация": "sum"} if "Реализация" in detail_df.columns else {} ),
                    })
                    .rename(columns={"ФИО": "Кол-во механиков"})
                )
                # Добавляем суммы по коммерции в разрезе ДЦ
                if "Категория" in detail_df.columns:
                    _dc_commercial = (
                        detail_df[detail_df["Категория"] == "commercial"]
                        .groupby("ДЦ", as_index=False)
                        .agg(**{
                            "Выплата коммерция": ("Начисление", "sum"),
                            "Н/Ч коммерция": ("Н/Ч", "sum"),
                        })
                    )
                    dc_analytics_df = dc_analytics_df.merge(_dc_commercial, on="ДЦ", how="left").fillna({"Выплата коммерция": 0, "Н/Ч коммерция": 0})
                    dc_analytics_df["Доля коммерции, %"] = np.where(
                        dc_analytics_df["Начисление"] > 0,
                        dc_analytics_df["Выплата коммерция"] / dc_analytics_df["Начисление"] * 100,
                        0,
                    ).round(1)

                dc_analytics_df["Средняя ставка за Н/Ч"] = np.where(
                    dc_analytics_df["Н/Ч"] > 0,
                    dc_analytics_df["Начисление"] / dc_analytics_df["Н/Ч"],
                    0,
                )

                if not rates_for_stats.empty and {"ДЦ", "Оклад", "Ставка"}.issubset(rates_for_stats.columns):
                    # Средние считаем только по реальным значениям (пустые/нулевые исключаем).
                    rates_for_avg = rates_for_stats.copy()
                    rates_for_avg["Оклад"] = pd.to_numeric(rates_for_avg["Оклад"], errors="coerce")
                    rates_for_avg["Ставка"] = pd.to_numeric(rates_for_avg["Ставка"], errors="coerce")

                    dc_salary_avg = (
                        rates_for_avg[rates_for_avg["Оклад"] > 0]
                        .groupby("ДЦ")["Оклад"]
                        .mean()
                        .rename("Средний оклад (ставки)")
                    )
                    dc_bonus_rate_avg = (
                        rates_for_avg[rates_for_avg["Ставка"] > 0]
                        .groupby("ДЦ")["Ставка"]
                        .mean()
                        .rename("Средняя ставка (ставки)")
                    )
                    dc_rates_rollup = pd.concat([dc_salary_avg, dc_bonus_rate_avg], axis=1).reset_index()
                    dc_analytics_df = dc_analytics_df.merge(dc_rates_rollup, on="ДЦ", how="left")

                dc_analytics_df["Н/Ч"] = dc_analytics_df["Н/Ч"].round(2)
                for c in ["Начисление", "Средняя ставка за Н/Ч", "Средний оклад (ставки)", "Средняя ставка (ставки)"]:
                    if c in dc_analytics_df.columns:
                        dc_analytics_df[c] = dc_analytics_df[c].round(0)

                if not dc_analytics_df.empty:
                    st.markdown("**Полноценная аналитика по ДЦ**")

                    dc_top = dc_analytics_df.sort_values("Начисление", ascending=False).iloc[0]
                    dc_bottom = dc_analytics_df.sort_values("Начисление", ascending=True).iloc[0]
                    col_dc1, col_dc2 = st.columns(2)
                    col_dc1.metric("Лидер ДЦ по выплатам", f"{dc_top['ДЦ']} ({dc_top['Начисление']:,.0f})")
                    col_dc2.metric("Минимум ДЦ по выплатам", f"{dc_bottom['ДЦ']} ({dc_bottom['Начисление']:,.0f})")

                    dc_chart_col1, dc_chart_col2 = st.columns(2)
                    with dc_chart_col1:
                        fig_dc_payout = px.bar(
                            dc_analytics_df.sort_values("Начисление", ascending=True),
                            x="Начисление",
                            y="ДЦ",
                            orientation="h",
                            text="Начисление",
                            color="Начисление",
                            color_continuous_scale=[[0, "#DBEAFE"], [1, "#1D4ED8"]],
                        )
                        fig_dc_payout.update_traces(texttemplate="%{text:,.0f}", textposition="outside")
                        fig_dc_payout.update_layout(
                            title="Выплаты по ДЦ",
                            margin=dict(l=10, r=10, t=50, b=10),
                            coloraxis_showscale=False,
                            xaxis_title="Начисление, тг",
                            yaxis_title="",
                        )
                        st.plotly_chart(fig_dc_payout, use_container_width=True)

                    with dc_chart_col2:
                        fig_dc_hours = px.bar(
                            dc_analytics_df.sort_values("Н/Ч", ascending=True),
                            x="Н/Ч",
                            y="ДЦ",
                            orientation="h",
                            text="Н/Ч",
                            color="Н/Ч",
                            color_continuous_scale=[[0, "#DCFCE7"], [1, "#15803D"]],
                        )
                        fig_dc_hours.update_traces(texttemplate="%{text:.2f}", textposition="outside")
                        fig_dc_hours.update_layout(
                            title="Выработка Н/Ч по ДЦ",
                            margin=dict(l=10, r=10, t=50, b=10),
                            coloraxis_showscale=False,
                            xaxis_title="Н/Ч",
                            yaxis_title="",
                        )
                        st.plotly_chart(fig_dc_hours, use_container_width=True)

                    st.dataframe(
                        dc_analytics_df.style.format({
                            "Н/Ч": "{:.2f}",
                            "Начисление": "{:,.0f}",
                            "Реализация": "{:,.0f}",
                            "Выплата коммерция": "{:,.0f}",
                            "Н/Ч коммерция": "{:.2f}",
                            "Доля коммерции, %": "{:.1f}%",
                            "Средняя ставка за Н/Ч": "{:,.0f}",
                            "Средний оклад (ставки)": "{:,.0f}",
                            "Средняя ставка (ставки)": "{:,.0f}",
                        }),
                        use_container_width=True,
                    )

                    # --- Доля 5 групп маппинга по ДЦ ---
                    if "Группа маппинга" in detail_df.columns:
                        st.markdown("**Доля Н/Ч по группам маппинга в разрезе ДЦ**")

                        _dc_grp = (
                            detail_df.groupby(["ДЦ", "Группа маппинга"], as_index=False)
                            .agg(**{"Н/Ч": ("Н/Ч", "sum"), "Начисление": ("Начисление", "sum")})
                        )
                        _dc_total = detail_df.groupby("ДЦ")["Н/Ч"].sum().rename("Итого Н/Ч")
                        _dc_grp = _dc_grp.merge(_dc_total, on="ДЦ", how="left")
                        _dc_grp["Доля Н/Ч, %"] = (_dc_grp["Н/Ч"] / _dc_grp["Итого Н/Ч"] * 100).round(1)

                        # Stacked bar 100% по Н/Ч
                        _all_grps_dc = sorted(_dc_grp["Группа маппинга"].dropna().unique())
                        _color_map = {
                            g: c for g, c in zip(
                                _all_grps_dc,
                                ["#3B82F6", "#F59E0B", "#10B981", "#8B5CF6", "#EF4444",
                                 "#06B6D4", "#EC4899", "#84CC16", "#F97316", "#6366F1"],
                            )
                        }
                        fig_dc_grp_nh = px.bar(
                            _dc_grp,
                            x="ДЦ",
                            y="Доля Н/Ч, %",
                            color="Группа маппинга",
                            barmode="stack",
                            text="Доля Н/Ч, %",
                            color_discrete_map=_color_map,
                        )
                        fig_dc_grp_nh.update_traces(texttemplate="%{text:.1f}%", textposition="inside")
                        fig_dc_grp_nh.update_layout(
                            title="Доля Н/Ч по группам маппинга (100% по каждому ДЦ)",
                            yaxis_title="Доля, %",
                            xaxis_title="",
                            margin=dict(l=10, r=10, t=50, b=10),
                            legend_title="Группа",
                        )
                        st.plotly_chart(fig_dc_grp_nh, use_container_width=True)

                        # Stacked bar абсолютные суммы начисления
                        fig_dc_grp_sum = px.bar(
                            _dc_grp,
                            x="ДЦ",
                            y="Начисление",
                            color="Группа маппинга",
                            barmode="stack",
                            text="Начисление",
                            color_discrete_map=_color_map,
                        )
                        fig_dc_grp_sum.update_traces(texttemplate="%{text:,.0f}", textposition="inside")
                        fig_dc_grp_sum.update_layout(
                            title="Начисление по группам маппинга в разрезе ДЦ",
                            yaxis_title="Начисление, ₸",
                            xaxis_title="",
                            margin=dict(l=10, r=10, t=50, b=10),
                            legend_title="Группа",
                        )
                        st.plotly_chart(fig_dc_grp_sum, use_container_width=True)

                        # Сводная таблица: ДЦ × группа → Н/Ч и доля
                        _dc_pivot_nh = _dc_grp.pivot_table(
                            index="ДЦ", columns="Группа маппинга", values="Н/Ч", aggfunc="sum", fill_value=0
                        )
                        _dc_pivot_pct = _dc_grp.pivot_table(
                            index="ДЦ", columns="Группа маппинга", values="Доля Н/Ч, %", aggfunc="sum", fill_value=0
                        )
                        _dc_pivot_nh.columns = [f"{c} (н/ч)" for c in _dc_pivot_nh.columns]
                        _dc_pivot_pct.columns = [f"{c} (%)" for c in _dc_pivot_pct.columns]
                        _dc_grp_table = _dc_pivot_nh.join(_dc_pivot_pct).reset_index()
                        # Чередуем колонки: н/ч + % на каждую группу
                        _reordered = ["ДЦ"]
                        for _g in _all_grps_dc:
                            if f"{_g} (н/ч)" in _dc_grp_table.columns:
                                _reordered.append(f"{_g} (н/ч)")
                            if f"{_g} (%)" in _dc_grp_table.columns:
                                _reordered.append(f"{_g} (%)")
                        _dc_grp_table = _dc_grp_table[[c for c in _reordered if c in _dc_grp_table.columns]]
                        _fmt_grp = {c: ("{:.2f}" if "(н/ч)" in c else "{:.1f}%") for c in _dc_grp_table.columns if c != "ДЦ"}
                        st.dataframe(
                            _dc_grp_table.style.format(_fmt_grp),
                            use_container_width=True,
                        )

                st.session_state["mech_dc_analytics_df"] = dc_analytics_df

                st.markdown("**Выплаты по брендам и категориям**")
                st.dataframe(
                    brand_category_df.style.format({"Н/Ч": "{:.2f}", "Начисление": "{:,.0f}"}),
                    use_container_width=True,
                )

                st.markdown("**Итоги по категориям**")
                st.dataframe(
                    category_total_df.style.format(
                        {"Н/Ч": "{:.2f}", "Начисление": "{:,.0f}", "Доля выплат, %": "{:.2f}%"}
                    ),
                    use_container_width=True,
                )

                st.markdown("**Итоги по брендам**")
                st.dataframe(
                    brand_total_df.style.format({"Н/Ч": "{:.2f}", "Начисление": "{:,.0f}"}),
                    use_container_width=True,
                )

                with st.expander("Сводная матрица выплат (бренд × категория)"):
                    st.dataframe(
                        pivot_brand_category.style.format("{:,.0f}", subset=[c for c in pivot_brand_category.columns if c != "Бренд"]),
                        use_container_width=True,
                    )

                st.session_state["mech_brand_category_df"] = brand_category_df
                st.session_state["mech_category_total_df"] = category_total_df
                st.session_state["mech_brand_total_df"] = brand_total_df
                st.session_state["mech_brand_category_pivot_df"] = pivot_brand_category

            # --- Фильтр по ФИО ---
            if not detail_scope.empty:
                all_fios = sorted(detail_scope["ФИО"].unique().tolist())
                selected_fio = st.selectbox("🔍 Детализация по механику", ["— все —"] + all_fios, key="mech_fio_filter")

                show_df = detail_scope if selected_fio == "— все —" else detail_scope[detail_scope["ФИО"] == selected_fio]

                st.subheader("📊 Детализация по типам нарядов")

                category_labels = {
                    "warranty": "Гарантийные",
                    "internal": "Внутренние / ППП",
                    "commercial": "Коммерческие",
                    "exclude": "Исключено",
                }
                detail_view = show_df.copy()

                # Подставляем ДЦ из файла ставок (если сопоставление найдено), чтобы убрать неоднозначность исходного ДЦ.
                if not rates_match_df.empty and "ДЦ в ставках" in rates_match_df.columns:
                    dc_map = (
                        rates_match_df[rates_match_df["Статус сопоставления"].isin(["exact", "fuzzy"])]
                        .sort_values(["Совпадение"], ascending=False)
                        .drop_duplicates(subset=["ФИО"], keep="first")
                        .set_index("ФИО")["ДЦ в ставках"]
                        .to_dict()
                    )
                    detail_view["ДЦ"] = detail_view["ФИО"].map(dc_map).fillna(detail_view["ДЦ"])

                detail_view["Категория"] = detail_view["Категория"].map(category_labels).fillna(detail_view["Категория"])
                detail_view["Статус"] = np.where(
                    (detail_view["Ставка"] == 0) & (detail_view["Категория"] == "Коммерческие"),
                    "нет ставки",
                    "OK",
                )

                display_columns = [
                    "ФИО",
                    "ДЦ",
                    "Тип наряда",
                    "Группа маппинга",
                    "Категория",
                    "Статус",
                    "Ставка",
                    "Н/Ч",
                    "Начисление",
                ]
                
                def highlight_errors(s):
                    # Подсвечиваем строки, где ставка 0, но категория не "Исключено"
                    # и не "Коммерческие" (если для коммерческих ставка 0 - это нормально, если нет индивидуальной)
                    return ['background-color: #ffcccc' if s.Ставка == 0 and s.Категория not in ['Исключено', 'Коммерческие'] else '' for _ in s]

                styled_detail = detail_view[display_columns].style.apply(
                    highlight_errors,
                    axis=1,
                ).format({"Н/Ч": "{:.2f}", "Начисление": "{:,.0f}", "Ставка": "{:,.0f}"})

                st.dataframe(
                    styled_detail,
                    use_container_width=True,
                )

                # Диагностика по маппингу
                records_all = st.session_state.get("mech_records", [])
                all_types = set(r["тип_наряда"] for r in records_all)
                mapping_dict = st.session_state.get("mech_work_type_mapping", {})
                mapped_types = set(mapping_dict.keys()) if mapping_dict else set()
                original_mapped_types = set(st.session_state.get("mech_original_mapped_types", []))
                excluded_by_result = all_types - set(detail_df["Тип наряда"].unique())

                approx_mapping_df = st.session_state.get("mech_approx_mapping_df", pd.DataFrame())
                approx_types = (
                    set(approx_mapping_df["Новый тип"].astype(str))
                    if (not approx_mapping_df.empty and "Новый тип" in approx_mapping_df.columns)
                    else set()
                )
                if original_mapped_types:
                    missing_in_mapping = all_types - (original_mapped_types | approx_types)
                else:
                    missing_in_mapping = all_types - mapped_types if mapping_dict else set()
                if not approx_mapping_df.empty:
                    with st.expander(f"🪄 Автоприближенное сопоставление типов ({len(approx_mapping_df)})"):
                        st.write("Эти типы автоматически сопоставлены с похожими строками из вашего маппинга.")
                        st.dataframe(
                            approx_mapping_df.style.format({"Сходство": "{:.3f}"}),
                            use_container_width=True,
                        )

                if mapping_dict and missing_in_mapping:
                    with st.expander(f"⚠️ Типы нарядов отсутствуют в файле маппинга ({len(missing_in_mapping)})"):
                        st.write(sorted(missing_in_mapping))

                if excluded_by_result:
                    with st.expander(f"⚠️ Типы нарядов без начисления / исключённые ({len(excluded_by_result)})"):
                        st.write("Эти типы не попали в начисление (исключены маппингом или авто-классификацией):")
                        st.write(sorted(excluded_by_result))

                # --- Выгрузка ---
                export_detail_df = detail_scope.copy()
                export_summary_df = summary_scope.copy()
                _excel_buf = BytesIO()
                with pd.ExcelWriter(_excel_buf, engine="openpyxl") as _wr:
                    export_summary_df.to_excel(_wr, sheet_name="Итог", index=False)
                    export_detail_df.to_excel(_wr, sheet_name="Детализация", index=False)
                    style_report_sheet(_wr.book["Итог"], "Итог по механикам", integer_columns=["Итого начисление"])
                    style_report_sheet(_wr.book["Детализация"], "Детализация по механикам", integer_columns=["Ставка", "Начисление"])
                    if not rates_match_df.empty:
                        rates_match_df.to_excel(_wr, sheet_name="Сопоставление ставок", index=False)
                        style_report_sheet(_wr.book["Сопоставление ставок"], "Сопоставление с файлом ставок", integer_columns=["Оклад (ставки)", "Ставка (ставки)"])
                    if not rates_mismatch_df.empty:
                        rates_mismatch_df.to_excel(_wr, sheet_name="Проблемные ФИО", index=False)
                        style_report_sheet(_wr.book["Проблемные ФИО"], "ФИО без уверенного сопоставления")
                    if "mech_approx_mapping_df" in st.session_state and not st.session_state["mech_approx_mapping_df"].empty:
                        st.session_state["mech_approx_mapping_df"].to_excel(_wr, sheet_name="Авто-маппинг типов", index=False)
                        style_report_sheet(_wr.book["Авто-маппинг типов"], "Автоприближенное сопоставление типов")
                    if "mech_brand_category_df" in st.session_state:
                        st.session_state["mech_brand_category_df"].to_excel(_wr, sheet_name="Аналитика Бренд-Кат", index=False)
                        style_report_sheet(_wr.book["Аналитика Бренд-Кат"], "Выплаты по брендам и категориям", integer_columns=["Начисление"])
                    if "mech_category_total_df" in st.session_state:
                        st.session_state["mech_category_total_df"].to_excel(_wr, sheet_name="Итоги Категории", index=False)
                        style_report_sheet(_wr.book["Итоги Категории"], "Итоги по категориям", integer_columns=["Начисление"])
                    if "mech_brand_total_df" in st.session_state:
                        st.session_state["mech_brand_total_df"].to_excel(_wr, sheet_name="Итоги Бренды", index=False)
                        style_report_sheet(_wr.book["Итоги Бренды"], "Итоги по брендам", integer_columns=["Начисление"])
                    if "mech_brand_category_pivot_df" in st.session_state:
                        st.session_state["mech_brand_category_pivot_df"].to_excel(_wr, sheet_name="Матрица БрендКат", index=False)
                        style_report_sheet(_wr.book["Матрица БрендКат"], "Матрица выплат бренд × категория")
                    if "mech_top_mechanics_df" in st.session_state and not st.session_state["mech_top_mechanics_df"].empty:
                        st.session_state["mech_top_mechanics_df"].to_excel(_wr, sheet_name="Топ 10 механиков", index=False)
                        style_report_sheet(_wr.book["Топ 10 механиков"], "Топ-10 механиков по начислению", integer_columns=["Итого начисление"])
                    if "mech_dc_salary_compare_df" in st.session_state and not st.session_state["mech_dc_salary_compare_df"].empty:
                        st.session_state["mech_dc_salary_compare_df"].rename(columns={
                            "Количество_механиков": "Кол-во механиков",
                            "Средний_оклад": "Средний оклад",
                            "Медианный_оклад": "Медианный оклад",
                            "Мин_оклад": "Мин. оклад",
                            "Макс_оклад": "Макс. оклад",
                            "Средняя_ставка": "Средняя ставка",
                        }).to_excel(_wr, sheet_name="Оклады по ДЦ", index=False)
                        style_report_sheet(
                            _wr.book["Оклады по ДЦ"],
                            "Сравнение окладов и ставок по ДЦ",
                            integer_columns=["Средний оклад", "Медианный оклад", "Мин. оклад", "Макс. оклад", "Средняя ставка"],
                        )
                    if "mech_dc_analytics_df" in st.session_state and not st.session_state["mech_dc_analytics_df"].empty:
                        st.session_state["mech_dc_analytics_df"].to_excel(_wr, sheet_name="Аналитика по ДЦ", index=False)
                        style_report_sheet(
                            _wr.book["Аналитика по ДЦ"],
                            "Полноценная аналитика по ДЦ",
                            integer_columns=["Начисление", "Средняя ставка за Н/Ч", "Средний оклад (ставки)", "Средняя ставка (ставки)"],
                        )

                    cards_sheet = _wr.book.create_sheet("Карточки механиков")
                    write_mechanic_cards_sheet(cards_sheet, export_detail_df)

                    used_sheet_names = set(_wr.book.sheetnames)
                    for dc_name, dc_slice in export_detail_df.groupby("ДЦ", sort=True):
                        if dc_slice.empty:
                            continue
                        dc_title = str(dc_name).strip() if str(dc_name).strip() else "Без ДЦ"
                        safe_dc_sheet = make_safe_sheet_name(f"Карточки {dc_title}", used_sheet_names)
                        used_sheet_names.add(safe_dc_sheet)
                        dc_sheet = _wr.book.create_sheet(safe_dc_sheet)
                        write_mechanic_cards_sheet(dc_sheet, dc_slice)
                _excel_buf.seek(0)
                st.download_button(
                    "⬇️ Скачать Excel (аналитика + карточки механиков)",
                    data=_excel_buf,
                    file_name="механики_бонусы.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="mech_download",
                )


# ============================================================================
# РЕЖИМ 3: ОСТАЛЬНЫЕ СОТРУДНИКИ
# ============================================================================

elif mode == "📊 Остальные сотрудники":
    st.header("📊 Расчет бонусов остальных сотрудников")
    st.info("Функционал для остальных сотрудников будет добавлен позже")
