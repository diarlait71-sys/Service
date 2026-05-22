import glob
import os
import re

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = r'c:\Users\D.Muldabaev\Desktop\Приложения для сервиса\Ндс'
OUTPUT_PATH = os.path.join(BASE_DIR, 'Taraz_NDS_Full_Analysis.xlsx')


def parse_quarter_forms():
    quarters = {}
    pattern = os.path.join(BASE_DIR, 'Форма 300 * лист.xlsx')

    for path in sorted(glob.glob(pattern)):
        name = os.path.basename(path)
        match = re.search(r'Форма 300 (\d) кв (\d{4}) (\d) лист', name)
        if not match:
            continue

        quarter, year, sheet_no = match.groups()
        key = f'{year} Q{quarter}'
        workbook = openpyxl.load_workbook(path, data_only=True)
        sheet = workbook.active

        data = quarters.setdefault(
            key,
            {
                'year': int(year),
                'quarter': int(quarter),
                'taxable_turnover': 0,
                'exempt_turnover': 0,
                'total_turnover': 0,
                'output_vat': 0,
                'input_vat_total': 0,
                'input_vat_allowed': 0,
                'correction': 0,
            },
        )

        for row in range(1, sheet.max_row + 1):
            code = str(sheet.cell(row, 1).value or '').strip()
            if code == '300.00.001':
                data['taxable_turnover'] = sheet.cell(row, 16).value or 0
                data['output_vat'] = sheet.cell(row, 30).value or 0
            elif code == '300.00.003':
                data['correction'] = sheet.cell(row, 30).value or 0
            elif code == '300.00.005':
                data['exempt_turnover'] = sheet.cell(row, 16).value or 0
            elif code == '300.00.006':
                data['total_turnover'] = sheet.cell(row, 16).value or 0
            elif code == '300.00.023':
                data['input_vat_total'] = sheet.cell(row, 30).value or 0

        if sheet_no == '2':
            allowed_vat = sheet.cell(84, 30).value
            if isinstance(allowed_vat, (int, float)):
                data['input_vat_allowed'] = allowed_vat

    result = []
    for key in sorted(quarters, key=lambda item: (quarters[item]['year'], quarters[item]['quarter'])):
        item = quarters[key]
        taxable_turnover = item['taxable_turnover'] or 0
        exempt_turnover = item['exempt_turnover'] or 0
        total_turnover = item['total_turnover'] or (taxable_turnover + exempt_turnover)
        output_vat = item['output_vat'] or 0
        correction = item['correction'] or 0
        input_vat_total = item['input_vat_total'] or 0
        input_vat_allowed = item['input_vat_allowed'] or 0

        taxable_share = (taxable_turnover / total_turnover) if total_turnover else 0
        exempt_share = (exempt_turnover / total_turnover) if total_turnover else 0
        input_vat_not_allowed = input_vat_total - input_vat_allowed
        net_vat = output_vat + correction - input_vat_allowed

        result.append(
            {
                'period': key,
                'taxable_turnover': taxable_turnover,
                'exempt_turnover': exempt_turnover,
                'total_turnover': total_turnover,
                'taxable_share': taxable_share,
                'exempt_share': exempt_share,
                'output_vat': output_vat,
                'correction': correction,
                'input_vat_total': input_vat_total,
                'input_vat_allowed': input_vat_allowed,
                'input_vat_not_allowed': input_vat_not_allowed,
                'net_vat': net_vat,
            }
        )

    return result


def autosize(sheet):
    for col in range(1, sheet.max_column + 1):
        max_len = 0
        for row in range(1, sheet.max_row + 1):
            value = sheet.cell(row, col).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        sheet.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 14), 40)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if isinstance(cell.value, (int, float)):
                if 0 < abs(cell.value) < 1:
                    cell.number_format = '0.00%'
                else:
                    cell.number_format = '#,##0'


def apply_table_borders(sheet, min_row, max_row, min_col, max_col):
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            sheet.cell(row, col).border = border


def build_report(rows):
    workbook = Workbook()
    header_fill = PatternFill('solid', fgColor='123B5D')
    sub_fill = PatternFill('solid', fgColor='E9F2FB')
    accent_fill = PatternFill('solid', fgColor='DDEFD7')
    zebra_fill = PatternFill('solid', fgColor='F7FBFF')
    title_font = Font(size=15, bold=True, color='123B5D')
    header_font_white = Font(size=10, bold=True, color='FFFFFF')
    section_font = Font(size=12, bold=True, color='123B5D')
    base_font = Font(size=10, color='1F1F1F')

    summary = workbook.active
    summary.title = 'Вывод по Ф300'

    totals_2025 = [row for row in rows if row['period'].startswith('2025')]
    total_taxable = sum(row['taxable_turnover'] for row in totals_2025)
    total_exempt = sum(row['exempt_turnover'] for row in totals_2025)
    total_turnover = sum(row['total_turnover'] for row in totals_2025)
    total_output_vat = sum(row['output_vat'] for row in totals_2025)
    total_correction = sum(row['correction'] for row in totals_2025)
    total_input = sum(row['input_vat_total'] for row in totals_2025)
    total_allowed = sum(row['input_vat_allowed'] for row in totals_2025)
    total_not_allowed = sum(row['input_vat_not_allowed'] for row in totals_2025)
    total_net = sum(row['net_vat'] for row in totals_2025)
    taxable_share = total_taxable / total_turnover if total_turnover else 0
    exempt_share = total_exempt / total_turnover if total_turnover else 0

    summary_rows = [
        ['Анализ форм 300: ДЦ Тараз'],
        ['Источник', 'Только формы 300 из папки НДС'],
        ['Период', '2025 полностью + 1 кв 2026'],
        [],
        ['Вывод по факту 2025'],
        ['Общий оборот', total_turnover],
        ['Облагаемый оборот', total_taxable],
        ['Освобожденный оборот', total_exempt],
        ['Доля облагаемого оборота', taxable_share],
        ['Доля освобожденного оборота', exempt_share],
        ['Начисленный НДС', total_output_vat],
        ['Корректировка НДС', total_correction],
        ['Входной НДС всего', total_input],
        ['Разрешенный к зачету НДС', total_allowed],
        ['НДС не принят к зачету', total_not_allowed],
        ['Чистый НДС к уплате', total_net],
        [],
        ['Ключевой смысл'],
        ['По формам 300 видно, что компания уже применяет пропорциональный метод зачета НДС.'],
        ['Это означает, что формы уже подтверждают смешанный характер оборота: облагаемый и освобожденный одновременно.'],
        ['Если доля освобожденного оборота вырастет, сумма НДС к зачету будет снижаться пропорционально.'],
        ['Файл намеренно не использует ОПУ и ОСВ: только факты из деклараций 300.'],
    ]

    for r_idx, row in enumerate(summary_rows, 1):
        for c_idx, value in enumerate(row, 1):
            summary.cell(r_idx, c_idx, value)

    summary.merge_cells('A1:B1')
    summary['A1'].font = title_font
    summary['A1'].alignment = Alignment(horizontal='left', vertical='center')

    summary['A5'].font = section_font
    summary['A5'].fill = accent_fill
    summary['A18'].font = section_font
    summary['A18'].fill = sub_fill

    for row in range(2, 4):
        for col in range(1, 3):
            summary.cell(row, col).fill = sub_fill
            summary.cell(row, col).font = base_font

    for row in range(6, 17):
        summary.cell(row, 1).font = Font(size=10, bold=True, color='2F2F2F')
        summary.cell(row, 2).font = base_font
        if row % 2 == 0:
            summary.cell(row, 1).fill = zebra_fill
            summary.cell(row, 2).fill = zebra_fill

    for row in range(19, 23):
        summary.cell(row, 1).font = base_font

    summary.freeze_panes = 'A6'
    apply_table_borders(summary, 2, 3, 1, 2)
    apply_table_borders(summary, 6, 16, 1, 2)
    apply_table_borders(summary, 19, 22, 1, 1)

    details = workbook.create_sheet('Поквартально')
    headers = [
        'Период',
        'Облагаемый оборот',
        'Освобожденный оборот',
        'Общий оборот',
        'Доля облагаемого',
        'Доля освобожденного',
        'Начисленный НДС',
        'Корректировка',
        'Входной НДС всего',
        'Разрешенный к зачету',
        'НДС не принят к зачету',
        'Чистый НДС к уплате',
    ]
    for c_idx, value in enumerate(headers, 1):
        cell = details.cell(1, c_idx, value)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for r_idx, row in enumerate(rows, 2):
        details.cell(r_idx, 1, row['period'])
        details.cell(r_idx, 2, row['taxable_turnover'])
        details.cell(r_idx, 3, row['exempt_turnover'])
        details.cell(r_idx, 4, row['total_turnover'])
        details.cell(r_idx, 5, row['taxable_share'])
        details.cell(r_idx, 6, row['exempt_share'])
        details.cell(r_idx, 7, row['output_vat'])
        details.cell(r_idx, 8, row['correction'])
        details.cell(r_idx, 9, row['input_vat_total'])
        details.cell(r_idx, 10, row['input_vat_allowed'])
        details.cell(r_idx, 11, row['input_vat_not_allowed'])
        details.cell(r_idx, 12, row['net_vat'])
        if r_idx % 2 == 0:
            for col in range(1, 13):
                details.cell(r_idx, col).fill = zebra_fill

    details.freeze_panes = 'A2'
    apply_table_borders(details, 1, max(details.max_row, 2), 1, 12)

    source = workbook.create_sheet('Источник строк')
    source_rows = [
        ['Какие строки формы 300 использованы'],
        ['300.00.001', 'Облагаемый оборот и начисленный НДС'],
        ['300.00.003', 'Корректировка размера облагаемого оборота'],
        ['300.00.005', 'Освобожденный от НДС оборот'],
        ['300.00.006', 'Общий оборот'],
        ['300.00.023', 'Общая сумма НДС, относимого в зачет'],
        ['Строка 84 на 2 листе', 'Разрешенный к зачету НДС при пропорциональном методе'],
        [],
        ['Примечание'],
        ['В 2026 Q1 в формах 300 нет заполненного разрешенного к зачету НДС в той же структуре, поэтому в отчете он оставлен как 0 по факту файла.'],
    ]
    for r_idx, row in enumerate(source_rows, 1):
        for c_idx, value in enumerate(row, 1):
            source.cell(r_idx, c_idx, value)
    source['A1'].font = section_font
    source['A9'].font = section_font
    source['A1'].fill = sub_fill
    source['A9'].fill = sub_fill

    for row in range(2, 8):
        source.cell(row, 1).font = Font(size=10, bold=True, color='2F2F2F')
        source.cell(row, 2).font = base_font
        if row % 2 == 0:
            source.cell(row, 1).fill = zebra_fill
            source.cell(row, 2).fill = zebra_fill

    source.cell(10, 1).font = base_font
    source.freeze_panes = 'A2'
    apply_table_borders(source, 2, 7, 1, 2)

    for sheet in workbook.worksheets:
        autosize(sheet)

    workbook.save(OUTPUT_PATH)


def main():
    rows = parse_quarter_forms()
    build_report(rows)
    print(OUTPUT_PATH)


if __name__ == '__main__':
    main()