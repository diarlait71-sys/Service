import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

path = r'C:\Users\D.Muldabaev\Desktop\Приложения для сервиса\Монит 28.04.2026) 2 — копия.xlsx'

wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print('=== ЛИСТЫ ===')
for name in wb.sheetnames:
    print(' -', name)

print()
for sheet_name in wb.sheetnames:
    if 'свод' in sheet_name.lower():
        print(f'[пропускаю лист Свод: {sheet_name}]')
        continue
    ws = wb[sheet_name]
    print(f'\n=== ЛИСТ: {sheet_name} ===')
    rows = list(ws.iter_rows(values_only=True))
    print(f'Строк: {len(rows)}, Столбцов: {ws.max_column}')
    print('--- Первые 10 строк ---')
    for i, row in enumerate(rows[:10]):
        print(f'  {i+1}: {row}')
    print('--- Строки 11-20 ---')
    for i, row in enumerate(rows[10:20], start=11):
        print(f'  {i}: {row}')
