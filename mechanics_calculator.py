"""
Модуль для расчета бонусов механиков
Логика: Норма часов × Ставка по типу наряда = Сумма
"""

import pandas as pd
from dataclasses import dataclass, field
from typing import Dict, List, Optional
from enum import Enum
from difflib import SequenceMatcher
import re


class WorkOrderType(Enum):
    """Типы заказ-нарядов"""
    PPP = "ППП"                          # Внутренний отдел Продаж, ППП
    ADDITIONAL = "Доп"                   # Дополнительные услуги
    INTERNAL_SERVICE = "Вн сервис"      # Внутренний сервис
    SERVICE_ADDITIONAL = "Сервис/Доп"   # Сервис, Доп сервис
    WARRANTY = "Гарантия"                # Гарантия


# Единая фиксированная ставка для ППП и Внутреннего сервиса (одна для всех механиков и городов)
FLAT_RATE = 1895
FLAT_RATE_TYPES = {WorkOrderType.PPP, WorkOrderType.INTERNAL_SERVICE}

# Фиксированные ставки для укрупненного расчета по типам нарядов
WARRANTY_RATE = 3915
INTERNAL_RATE = 1895

# Базовый маппинг категорий для нового файла "Механики.xlsx"
# Позже можно расширить/переопределить бизнес-маппингом из внешнего источника
DEFAULT_WORK_TYPE_CATEGORY_KEYWORDS: Dict[str, List[str]] = {
    "warranty": ["гарантия", "гарантий"],
    "internal": [
        "внутрен",
        "вн",
        "оп сервис",
        "ппп",
        "предпродаж",
        "сервис ремзона",
    ],
    "commercial": ["сервис", "доп", "коммерч", "вэк"],
    "exclude": ["банковский продукт", "прочее"],
}

DEFAULT_CATEGORY_RATES: Dict[str, float] = {
    "warranty": WARRANTY_RATE,
    "internal": INTERNAL_RATE,
    "commercial": 0,
    "exclude": 0,
}


# Ключевые слова для определения типа наряда
WORK_ORDER_KEYWORDS = {
    WorkOrderType.PPP: ["ппп", "внутренний отдел продаж"],
    WorkOrderType.ADDITIONAL: ["доп", "дополнительный"],
    WorkOrderType.INTERNAL_SERVICE: ["вн сервис", "внутренний сервис"],
    WorkOrderType.SERVICE_ADDITIONAL: ["сервис", "сервиса"],
    WorkOrderType.WARRANTY: ["гарантия", "гарант"],
}


@dataclass
class MechanicRates:
    """Ставки механика для разных типов работ"""
    fio: str
    city: str = ""                     # Город / Дилерский центр
    # Индивидуальные утверждённые ставки (разные для каждого механика)
    service_additional: float = 0      # Ставка Сервис/Доп
    warranty: float = 0                # Ставка Гарантия
    additional: float = 0             # Ставка Доп

    def get_rate_by_type(self, work_type: WorkOrderType) -> float:
        """Получить ставку по типу наряда.
        ППП и Вн сервис — единая фиксированная ставка FLAT_RATE для всех.
        Остальные — индивидуальная утверждённая ставка механика.
        """
        if work_type in FLAT_RATE_TYPES:
            return FLAT_RATE
        individual = {
            WorkOrderType.SERVICE_ADDITIONAL: self.service_additional,
            WorkOrderType.WARRANTY: self.warranty,
            WorkOrderType.ADDITIONAL: self.additional,
        }
        return individual.get(work_type, 0)


@dataclass
class WorkOrder:
    """Заказ-наряд механика"""
    mechanic_fio: str
    description: str                    # Описание работы
    work_type: WorkOrderType           # Тип наряда
    norm_hours: float                  # Норма часов (н/ч)
    dealer_center: str = ""            # ДЦ/город (если известен)
    current_sum: float = 0             # Текущая сумма из 1С
    calculated_sum: float = field(default=0, init=False)  # Пересчитанная сумма
    
    def calculate_sum(self, rate: float) -> float:
        """Рассчитать сумму: Норма часов × Ставка"""
        self.calculated_sum = self.norm_hours * rate
        return self.calculated_sum


def detect_work_order_type(description: str) -> Optional[WorkOrderType]:
    """
    Определить тип наряда по описанию
    Ищет ключевые слова в тексте
    """
    if not description:
        return None
    
    description_lower = str(description).lower().strip()
    
    # Более точное определение с русским языком
    if 'гарантия' in description_lower or 'гарант' in description_lower:
        return WorkOrderType.WARRANTY
    
    if 'доп' in description_lower or 'дополнительный' in description_lower:
        if 'сервис' not in description_lower:  # Чтобы не перепутать с Сервис/Доп
            return WorkOrderType.ADDITIONAL
    
    if 'ппп' in description_lower:
        return WorkOrderType.PPP
    
    if 'вн сервис' in description_lower or 'внутренний сервис' in description_lower:
        return WorkOrderType.INTERNAL_SERVICE
    
    if 'сервис' in description_lower:
        return WorkOrderType.SERVICE_ADDITIONAL
    
    return None


class MechanicsCalculator:
    """Калькулятор бонусов для механиков"""
    
    def __init__(self):
        self.mechanics_rates: Dict[str, MechanicRates] = {}
        self.work_orders: List[WorkOrder] = []
    
    def load_mechanics_rates(self, df: pd.DataFrame, dealer_center: str = "") -> bool:
        """
        Загрузить ставки механиков из DataFrame.

        Ожидаемые колонки:
        - Город / ДЦ / Дилерский центр  (фильтр по dealer_center)
        - ФИО (или Механик, Имя)
        - Сервис/Доп  — индивидуальная ставка
        - Гарантия    — индивидуальная ставка
        - Доп         — индивидуальная ставка

        ППП и Вн сервис не нужны в файле — ставка FLAT_RATE применяется автоматически.
        """
        try:
            # Колонка города/ДЦ
            city_col = self._find_column(df, ['город', 'дц', 'дилерский центр', 'диллерский центр', 'dealer center', 'название дц'])

            # Колонка ФИО
            fio_col = self._find_column(df, ['фио', 'механик', 'имя', 'name'])
            if not fio_col:
                print("Не найдена колонка с ФИО механика")
                return False

            # Индивидуальные ставки
            service_additional_col = self._find_column(df, ['сервис/доп', 'сервис доп'])
            warranty_col = self._find_column(df, ['гарантия', 'warranty'])
            additional_col = self._find_column(df, ['доп', 'дополнительные'])

            dc_lower = dealer_center.lower().strip()

            for _, row in df.iterrows():
                fio = str(row[fio_col]).strip()
                if not fio or fio.lower() == 'nan':
                    continue

                city = str(row[city_col]).strip() if city_col else ""

                # Фильтруем по городу/ДЦ если указан
                if dc_lower and city_col and city.lower() != dc_lower:
                    continue

                rates = MechanicRates(
                    fio=fio,
                    city=city,
                    service_additional=self._get_numeric_value(row, service_additional_col, 0),
                    warranty=self._get_numeric_value(row, warranty_col, 0),
                    additional=self._get_numeric_value(row, additional_col, 0),
                )
                self.mechanics_rates[fio] = rates

            loaded = len(self.mechanics_rates)
            if dealer_center:
                print(f"Загружено {loaded} механиков для ДЦ '{dealer_center}'")
            else:
                print(f"Загружено {loaded} механиков (все города)")
            return loaded > 0

        except Exception as e:
            print(f"Ошибка при загрузке ставок механиков: {e}")
            return False
    
    def load_work_orders(self, df: pd.DataFrame) -> bool:
        """
        Загрузить заказ-наряды из DataFrame
        
        Ожидаемые колонки:
        - Механик (ФИО)
        - Описание (работа)
        - Тип наряда (опционально - будет определен по описанию)
        - Норма часов (н/ч)
        - Сумма (текущая из 1С, опционально)
        """
        try:
            self.work_orders = []
            
            # Находим нужные колонки
            mechanic_col = self._find_column(df, ['механик', 'фио', 'мастер', 'имя'])
            description_col = self._find_column(df, ['описание', 'работа', 'наряд', 'description'])
            type_col = self._find_column(df, ['тип наряда', 'тип', 'type'])
            norm_hours_col = self._find_column(df, ['норма часов', 'н/ч', 'часы', 'hours', 'норма'])
            sum_col = self._find_column(df, ['сумма', 'sum', 'amount'])
            
            if not (mechanic_col and description_col and norm_hours_col):
                print(f"Не найдены обязательные колонки")
                print(f"  Найдено - Механик: {mechanic_col}, Описание: {description_col}, Норма часов: {norm_hours_col}")
                print(f"  Доступные колонки: {list(df.columns)}")
                return False
            
            for idx, row in df.iterrows():
                mechanic = str(row[mechanic_col]).strip()
                description = str(row[description_col]).strip()
                
                if not mechanic or mechanic.lower() == 'nan':
                    continue
                if not description or description.lower() == 'nan':
                    continue
                
                # Определяем тип наряда
                if type_col and pd.notna(row[type_col]):
                    work_type_str = str(row[type_col]).strip()
                    work_type = self._parse_work_type(work_type_str)
                else:
                    work_type = detect_work_order_type(description)
                
                if not work_type:
                    # Логирование для отладки
                    print(f"⚠️ Не удалось определить тип наряда для: {description}")
                    work_type = WorkOrderType.ADDITIONAL  # По умолчанию
                
                norm_hours = self._get_numeric_value(row, norm_hours_col, 0)
                current_sum = self._get_numeric_value(row, sum_col, 0) if sum_col else 0
                
                if norm_hours <= 0:
                    # Пропускаем нулевые часы
                    continue
                
                work_order = WorkOrder(
                    mechanic_fio=mechanic,
                    description=description,
                    work_type=work_type,
                    norm_hours=norm_hours,
                    dealer_center="",
                    current_sum=current_sum
                )
                self.work_orders.append(work_order)
                print(f"✓ {mechanic}: {description} -> {work_type.value}, {norm_hours} часов")
            
            return len(self.work_orders) > 0
        
        except Exception as e:
            print(f"Ошибка при загрузке заказ-нарядов: {e}")
            import traceback
            traceback.print_exc()
            return False

    def load_work_orders_from_mechanics_report(self, df: pd.DataFrame) -> bool:
        """
        Загрузить наряды из формата файла "Механики.xlsx" (лист TDSheet).

        Ожидаемый шаблон:
        - строка 5 (индекс 4): заголовки, где колонка A = "Механик", E = "Н/Ч"
        - строка 6 (индекс 5): строка-описание "Заказ наряд.Тип наряда"
        - далее в колонке A: либо ДЦ (например, DCG ALMATY), либо тип наряда
        - в колонке E: норма-часы
        """
        try:
            self.work_orders = []
            current_dc = ""

            if df.shape[1] < 5:
                print("Недостаточно колонок в отчете механиков")
                return False

            # Пропускаем служебные строки до данных. Обычно данные начинаются с индекса 6.
            for idx in range(6, len(df)):
                raw_name = df.iat[idx, 0] if 0 in df.columns else None
                raw_hours = df.iat[idx, 4] if 4 in df.columns else None

                if pd.isna(raw_name):
                    continue

                name = str(raw_name).strip()
                if not name:
                    continue

                # Строка ДЦ: запоминаем контекст города/ДЦ, но в расчет как наряд не берем.
                if self._is_dealer_center_row(name):
                    current_dc = name
                    continue

                norm_hours = self._to_float(raw_hours)
                if norm_hours <= 0:
                    continue

                work_type = detect_work_order_type(name)
                if not work_type:
                    # Для нового формата точная классификация идет отдельным маппингом по description,
                    # поэтому сохраняем как ADDITIONAL, чтобы объект был валидным.
                    work_type = WorkOrderType.ADDITIONAL

                self.work_orders.append(
                    WorkOrder(
                        mechanic_fio="",
                        dealer_center=current_dc,
                        description=name,
                        work_type=work_type,
                        norm_hours=norm_hours,
                        current_sum=0,
                    )
                )

            print(f"Загружено {len(self.work_orders)} строк нарядов из отчета механиков")
            return len(self.work_orders) > 0

        except Exception as e:
            print(f"Ошибка при загрузке отчета механиков: {e}")
            return False

    def calculate_bonus_by_work_type(
        self,
        category_keywords: Optional[Dict[str, List[str]]] = None,
        category_rates: Optional[Dict[str, float]] = None,
        exclude_categories: Optional[List[str]] = None,
    ) -> pd.DataFrame:
        """
        Расчет бонусов в разрезе типа наряда.

        По умолчанию:
        - гарантийные: 3915 тг
        - внутренние/ОП сервис/ППП/предпродажные: 1895 тг
        - категории из exclude не попадают в расчет и не отражают норма-часы

        Можно передать пользовательский маппинг category_keywords и ставки category_rates.
        """
        category_keywords = category_keywords or DEFAULT_WORK_TYPE_CATEGORY_KEYWORDS
        category_rates = category_rates or DEFAULT_CATEGORY_RATES
        exclude_categories = exclude_categories or ["exclude"]

        rows = []
        for wo in self.work_orders:
            category = self._classify_category(wo.description, category_keywords)
            if category in exclude_categories:
                continue

            rate = float(category_rates.get(category, 0) or 0)
            bonus = wo.norm_hours * rate
            rows.append(
                {
                    "ДЦ/Город": wo.dealer_center,
                    "Тип наряда (сырой)": wo.description,
                    "Категория": category,
                    "Норма часов": wo.norm_hours,
                    "Ставка": rate,
                    "Начисление": bonus,
                }
            )

        if not rows:
            return pd.DataFrame(
                columns=[
                    "ДЦ/Город",
                    "Тип наряда (сырой)",
                    "Категория",
                    "Норма часов",
                    "Ставка",
                    "Начисление",
                ]
            )

        df = pd.DataFrame(rows)
        grouped = (
            df.groupby(["ДЦ/Город", "Тип наряда (сырой)", "Категория", "Ставка"], as_index=False)
            .agg({"Норма часов": "sum", "Начисление": "sum"})
            .sort_values(["ДЦ/Город", "Категория", "Тип наряда (сырой)"])
        )

        grouped["Норма часов"] = grouped["Норма часов"].round(2)
        grouped["Начисление"] = grouped["Начисление"].round(2)
        return grouped
    
    def calculate_mechanics_bonuses(self) -> pd.DataFrame:
        """
        Рассчитать бонусы для всех механиков
        
        Returns:
            DataFrame с результатами
        """
        results = {}
        
        for work_order in self.work_orders:
            mechanic_fio = work_order.mechanic_fio
            
            if mechanic_fio not in results:
                results[mechanic_fio] = {
                    'ФИО': mechanic_fio,
                    'Всего_нарядов': 0,
                    'Норма_часов': 0,
                    'Сумма_1С': 0,
                    'Пересчитано': 0,
                    'Разница': 0,
                    'Детали': []
                }
            
            # Получаем ставку механика
            if mechanic_fio in self.mechanics_rates:
                rates = self.mechanics_rates[mechanic_fio]
                rate = rates.get_rate_by_type(work_order.work_type)
                work_order.calculate_sum(rate)
            else:
                work_order.calculated_sum = 0
            
            # Накапливаем данные
            results[mechanic_fio]['Всего_нарядов'] += 1
            results[mechanic_fio]['Норма_часов'] += work_order.norm_hours
            results[mechanic_fio]['Сумма_1С'] += work_order.current_sum
            results[mechanic_fio]['Пересчитано'] += work_order.calculated_sum
            results[mechanic_fio]['Разница'] += work_order.calculated_sum - work_order.current_sum
            
            results[mechanic_fio]['Детали'].append({
                'Описание': work_order.description,
                'Тип': work_order.work_type.value,
                'Тип ставки': 'Единая' if work_order.work_type in FLAT_RATE_TYPES else 'Индивидуальная',
                'Норма_часов': work_order.norm_hours,
                'Ставка': rates.get_rate_by_type(work_order.work_type) if mechanic_fio in self.mechanics_rates else 0,
                'Сумма_1С': work_order.current_sum,
                'Пересчитано': work_order.calculated_sum,
            })
        
        # Создаем DataFrame результатов
        result_list = []
        for mechanic, data in results.items():
            result_list.append({
                'ФИО': data['ФИО'],
                'Нарядов': data['Всего_нарядов'],
                'Норма часов': round(data['Норма_часов'], 2),
                'Сумма 1С': round(data['Сумма_1С'], 2),
                'Пересчитано': round(data['Пересчитано'], 2),
                'Разница': round(data['Разница'], 2),
            })
        
        return pd.DataFrame(result_list)
    
    def get_detailed_report(self) -> pd.DataFrame:
        """Получить детальный отчет по каждому наряду"""
        data = []
        for work_order in self.work_orders:
            if work_order.mechanic_fio in self.mechanics_rates:
                rates = self.mechanics_rates[work_order.mechanic_fio]
                rate = rates.get_rate_by_type(work_order.work_type)
            else:
                rate = 0
            
            data.append({
                'Механик': work_order.mechanic_fio,
                'ДЦ/Город': rates.city if work_order.mechanic_fio in self.mechanics_rates else '',
                'Описание': work_order.description[:50],  # Первые 50 символов
                'Тип наряда': work_order.work_type.value,
                'Тип ставки': 'Единая' if work_order.work_type in FLAT_RATE_TYPES else 'Индивидуальная',
                'Норма часов': work_order.norm_hours,
                'Ставка': rate,
                'Сумма 1С': work_order.current_sum,
                'Пересчитано': round(work_order.calculated_sum, 2),
            })
        
        return pd.DataFrame(data)
    
    # === ВСПОМОГАТЕЛЬНЫЕ МЕТОДЫ ===
    
    @staticmethod
    def _find_column(df: pd.DataFrame, keywords: List[str]) -> Optional[str]:
        """Найти колонку по ключевым словам"""
        def _norm(x: str) -> str:
            txt = str(x).lower().strip().replace("ё", "е")
            txt = re.sub(r"[^a-zа-я0-9]+", "", txt)
            return txt

        columns = list(df.columns)
        normalized_cols = {col: _norm(col) for col in columns}

        for keyword in keywords:
            k = _norm(keyword)
            if not k:
                continue

            # Сначала точное совпадение по нормализованному виду
            for col, norm_col in normalized_cols.items():
                if norm_col == k:
                    return col

            # Затем частичное совпадение
            for col, norm_col in normalized_cols.items():
                if k in norm_col or norm_col in k:
                    return col

        return None
    
    @staticmethod
    def _get_numeric_value(row, col: Optional[str], default: float = 0) -> float:
        """Получить числовое значение из ячейки"""
        if col is None:
            return default
        
        try:
            val = row[col]
            if pd.isna(val):
                return default
            return float(val)
        except:
            return default
    
    @staticmethod
    def _parse_work_type(work_type_str: str) -> Optional[WorkOrderType]:
        """Преобразовать строку в WorkOrderType"""
        work_type_str = work_type_str.lower().strip()
        
        for work_type in WorkOrderType:
            if work_type.value.lower() in work_type_str:
                return work_type
        
        # Пытаемся определить по ключевым словам
        return detect_work_order_type(work_type_str)

    @staticmethod
    def _is_dealer_center_row(text: str) -> bool:
        """Проверить, является ли строка заголовком ДЦ/города (например: DCG ALMATY)."""
        txt = str(text).strip().lower()
        return txt.startswith("dcg") or txt.startswith("too") or txt.startswith("тоо")

    def load_from_mechanics_xlsx(self, file_path: str) -> list:
        """
        Парсит Механики.xlsx через openpyxl с учётом отступов (indent):
          indent=0 + НЕ ДЦ  → ФИО механика
          indent=2           → тип наряда под текущим механиком

        Первая секция (ДЦ-итоги, где indent=0 строки — "DCG XXX"/"TOO XXX") пропускается.

        Возвращает список dict: {fio, тип_наряда, norm_hours}
        """
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        records = []
        current_mechanic = None
        in_mechanics_section = False
        current_dc = ""

        # Динамически определяем колонки по заголовкам, т.к. в разных выгрузках
        # Н/Ч/Тип/Работа могут сдвигаться.
        norm_col_idx = 6   # fallback: G
        realization_col_idx = 8  # fallback: I
        type_col_idx = 4   # fallback: E (Тип наряда)
        work_col_idx = 5   # fallback: F (Работа)

        for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row)):
            for idx, cell in enumerate(row):
                val = str(cell.value).strip().lower() if cell.value is not None else ""
                if not val:
                    continue
                if val in {"н/ч", "нч", "норма часов"}:
                    norm_col_idx = idx
                elif "реализовано" in val and "сумм" in val:
                    realization_col_idx = idx
                elif "тип" in val and "наряд" in val:
                    type_col_idx = idx
                elif val == "работа" or "работа" in val:
                    work_col_idx = idx

        for row in ws.iter_rows():
            if not row:
                continue

            cell_a = row[0]
            val = cell_a.value
            if val is None:
                continue

            name = str(val).strip()
            if not name:
                continue

            # Получаем отступ.
            try:
                indent = cell_a.alignment.indent if cell_a.alignment and cell_a.alignment.indent is not None else 0
            except Exception:
                indent = 0

            # Пропускаем служебные заголовки.
            low_name = name.lower()
            if low_name in {"механик", "заказ наряд", "заказ-наряд", "параметры:"}:
                continue

            # Н/Ч
            norm_hours = 0.0
            if len(row) > norm_col_idx and row[norm_col_idx].value is not None:
                norm_hours = self._to_float(row[norm_col_idx].value)

            # Реализовано работ на сумму.
            реализация = 0.0
            if len(row) > realization_col_idx and row[realization_col_idx].value is not None:
                реализация = self._to_float(row[realization_col_idx].value)

            if indent == 0:
                clean_name = re.sub(r"^,+", "", name)
                clean_name = re.sub(r"\s+", " ", clean_name).strip()

                if self._is_dealer_center_row(clean_name):
                    current_dc = clean_name
                    current_mechanic = None
                    in_mechanics_section = False
                    continue

                # Строки вида ",Имя Фамилия" и обычные ФИО считаем текущим механиком.
                if clean_name and "заказ-наряд" not in clean_name.lower():
                    in_mechanics_section = True
                    current_mechanic = clean_name
                continue

            if indent == 2 and in_mechanics_section and current_mechanic:
                if norm_hours <= 0:
                    continue

                raw_type = ""
                if len(row) > type_col_idx and row[type_col_idx].value is not None:
                    raw_type = str(row[type_col_idx].value).strip()
                if not raw_type and len(row) > work_col_idx and row[work_col_idx].value is not None:
                    raw_type = str(row[work_col_idx].value).strip()
                if not raw_type:
                    raw_type = name

                # Не сохраняем служебные строки вместо типа работ.
                if raw_type.lower() in {"тип наряда", "работа", "заказ наряд", "заказ-наряд"}:
                    continue

                records.append({
                    "fio": current_mechanic,
                    "dc": current_dc,
                    "тип_наряда": raw_type,
                    "norm_hours": norm_hours,
                    "реализация": реализация,
                })

        return records

    def load_work_type_mapping_xlsx(self, file_path: str) -> tuple[dict, dict]:
        """
        Загружает маппинг типов нарядов из файла "Типы Нарядов маппинг.xlsx".

        Ожидаемые колонки:
        - "ТипНаряда" (или похожее имя)
        - "Маппинг для бонуса" (или похожее имя)

        Возвращает:
        - work_type_mapping: {raw_type: category_key}
        - work_type_labels:  {raw_type: mapping_label}
        """
        df = pd.read_excel(file_path)
        if df.empty:
            return {}, {}

        def _norm(x: str) -> str:
            txt = str(x).lower().strip().replace("ё", "е")
            txt = re.sub(r"[^a-zа-я0-9]+", "", txt)
            return txt

        norm_cols = {_norm(col): col for col in df.columns}

        # Важно: сначала ищем максимально точные имена,
        # чтобы не выбирать колонку "Тип" вместо "ТипНаряда".
        type_col = (
            norm_cols.get("типнаряда")
            or norm_cols.get("типнаряд")
            or self._find_column(df, ["тип наряда", "наряд", "type"])
        )
        mapping_col = (
            norm_cols.get("маппингдлябонуса")
            or self._find_column(df, ["маппинг для бонуса", "маппинг", "группа", "категория"])
        )

        if not type_col or not mapping_col:
            raise ValueError(
                "Не найдены колонки маппинга. Ожидались: 'ТипНаряда' и 'Маппинг для бонуса'."
            )

        work_type_mapping: dict = {}
        work_type_labels: dict = {}

        for _, row in df.iterrows():
            raw_type = str(row[type_col]).strip() if pd.notna(row[type_col]) else ""
            raw_label = str(row[mapping_col]).strip() if pd.notna(row[mapping_col]) else ""
            if not raw_type:
                continue

            category_key = self._map_bonus_label_to_category(raw_label)
            work_type_mapping[raw_type] = category_key
            work_type_labels[raw_type] = raw_label

        return work_type_mapping, work_type_labels

    def build_summary_from_records(
        self,
        records: list,
        work_type_mapping: dict | None = None,
        work_type_labels: dict | None = None,
        category_rates: dict | None = None,
        exclude_types: set | None = None,
    ) -> tuple:
        """
        Строит сводные таблицы из записей load_from_mechanics_xlsx.

        work_type_mapping: {raw_type_name: category}  — пользовательский маппинг.
            Если не задан — используется keyword-based автоклассификация.
        work_type_labels: {raw_type_name: исходная группа из маппинга}.
        category_rates: {category: rate}  — ставки по категориям.
        exclude_types: множество raw_type_name, которые полностью исключаются из отчёта.

        Возвращает (detail_df, summary_df):
          detail_df  — детальный: ФИО, Тип наряда, Категория, Ставка, Н/Ч, Начисление
          summary_df — итог по механику: ФИО, Гарантия НЧ, Внутр НЧ, ... Итого начисление
        """
        _rates = category_rates or DEFAULT_CATEGORY_RATES
        _keywords = DEFAULT_WORK_TYPE_CATEGORY_KEYWORDS
        exclude_types = exclude_types or set()

        rows = []
        for r in records:
            raw_type = r["тип_наряда"]
            if raw_type in exclude_types:
                continue

            # Категория: сначала пользовательский маппинг, потом авто
            if work_type_mapping and raw_type in work_type_mapping:
                category = work_type_mapping[raw_type]
                mapping_label = (work_type_labels or {}).get(raw_type, "")
            else:
                category = self._classify_category(raw_type, _keywords)
                mapping_label = ""

            if category == "exclude":
                continue

            rate = float(_rates.get(category, 0) or 0)
            bonus = round(r["norm_hours"] * rate, 2)
            rows.append({
                "ФИО": r["fio"],
                "ДЦ": r.get("dc", ""),
                "Тип наряда": raw_type,
                "Группа маппинга": mapping_label,
                "Категория": category,
                "Ставка": rate,
                "Н/Ч": r["norm_hours"],
                "Начисление": bonus,
                "Источник ставки": "Авто-маппинг", # Добавлено для диагностики
                "Реализация": r.get("реализация", 0.0),
            })

        if not rows:
            empty_detail = pd.DataFrame(columns=["ФИО", "ДЦ", "Тип наряда", "Группа маппинга", "Категория", "Ставка", "Н/Ч", "Начисление", "Реализация"])
            empty_summary = pd.DataFrame(columns=["ФИО", "ДЦ", "Итого Н/Ч", "Итого начисление", "Реализация"])
            return empty_detail, empty_summary

        detail_df = pd.DataFrame(rows)

        summary_df = (
            detail_df.groupby(["ФИО", "ДЦ"], as_index=False)
            .agg(
                **{"Итого Н/Ч": ("Н/Ч", "sum"), "Итого начисление": ("Начисление", "sum"), "Реализация": ("Реализация", "sum")}
            )
            .sort_values(["ДЦ", "ФИО"])
        )
        summary_df["Итого Н/Ч"] = summary_df["Итого Н/Ч"].round(2)
        summary_df["Итого начисление"] = summary_df["Итого начисление"].round(2)
        summary_df["Реализация"] = summary_df["Реализация"].round(2)

        return detail_df, summary_df

    @staticmethod
    def _to_float(value) -> float:
        """Безопасное преобразование в float (поддержка строк с запятой)."""
        if value is None or pd.isna(value):
            return 0.0
        try:
            if isinstance(value, str):
                return float(value.replace(" ", "").replace(",", "."))
            return float(value)
        except Exception:
            return 0.0

    @staticmethod
    def _classify_category(description: str, category_keywords: Dict[str, List[str]]) -> str:
        """Классифицировать сырой тип наряда в бизнес-категорию по ключевым словам."""
        txt = str(description).lower().strip()

        # Важно: сначала исключаем не-бонусные типы, чтобы они не попали в общие категории.
        for keyword in category_keywords.get("exclude", []):
            if keyword in txt:
                return "exclude"

        for category, keywords in category_keywords.items():
            if category == "exclude":
                continue
            for keyword in keywords:
                if keyword in txt:
                    return category

        # Если не нашли в маппинге, и не исключено, то по умолчанию это коммерческая работа.
        # Это предотвращает исключение работ, которые просто не были явно классифицированы.
        if "сервис" in txt or "ремонт" in txt or "работы" in txt or "наряд" in txt:
            return "commercial"
        return "exclude" # Если даже общих слов нет, то действительно исключаем.

    @staticmethod
    def _map_bonus_label_to_category(label: str) -> str:
        """Преобразует текст группы из маппинга в ключ категории ставки."""
        txt = str(label).lower().strip()
        if not txt or txt == "nan":
            return "exclude"

        if "гаран" in txt:
            return "warranty"

        # Приоритетное бизнес-правило:
        # если группа содержит одновременно сервис и доп,
        # это коммерческие работы (ставка механика из файла ставок).
        if "сервис" in txt and "доп" in txt:
            return "commercial"

        if (
            "внутрен" in txt
            or "вн сервис" in txt
            or "ппп" in txt
            or "предпродаж" in txt
            or "оп сервис" in txt
            or "оп-сервис" in txt
        ):
            return "internal"

        if "сервис" in txt or "доп" in txt or "коммерч" in txt or "вэк" in txt:
            return "commercial"

        # Если не нашли в маппинге, проверяем базовые слова, чтобы не исключать лишнего
        if "сервис" in txt or "ремонт" in txt:
            return "commercial"
            
        return "exclude"

    @staticmethod
    def normalize_fio(value: str) -> str:
        """Нормализация ФИО для устойчивого сопоставления (регистр/пробелы/символы)."""
        txt = str(value or "").lower().strip().replace("ё", "е")
        # Нормализация частых казахских/кириллических вариантов написания.
        char_map = {
            "қ": "к", "ғ": "г", "ң": "н", "ә": "а", "ө": "о", "ұ": "у", "ү": "у", "һ": "х", "і": "и",
            "й": "и", "ъ": "", "ь": "",
        }
        for src, dst in char_map.items():
            txt = txt.replace(src, dst)
        txt = re.sub(r"[^a-zа-я0-9\s]", " ", txt)
        txt = re.sub(r"\s+", " ", txt).strip()
        return txt

    @staticmethod
    def _fio_similarity(a: str, b: str) -> float:
        """Оценка схожести ФИО с учетом пробелов/порядка/опечаток."""
        a_norm = MechanicsCalculator.normalize_fio(a)
        b_norm = MechanicsCalculator.normalize_fio(b)
        if not a_norm or not b_norm:
            return 0.0

        a_compact = a_norm.replace(" ", "")
        b_compact = b_norm.replace(" ", "")
        a_sorted = " ".join(sorted(a_norm.split()))
        b_sorted = " ".join(sorted(b_norm.split()))

        s1 = SequenceMatcher(None, a_norm, b_norm).ratio()
        s2 = SequenceMatcher(None, a_compact, b_compact).ratio()
        s3 = SequenceMatcher(None, a_sorted, b_sorted).ratio()

        score = max(s1, s2, s3)

        # Бонус, если совпадает фамилия (первый токен).
        a_tokens = a_norm.split()
        b_tokens = b_norm.split()
        if a_tokens and b_tokens and a_tokens[0] == b_tokens[0]:
            score = min(1.0, score + 0.05)

        return score

    @staticmethod
    def normalize_dc(value: str) -> str:
        """Нормализация названия ДЦ/города для сопоставления между файлами."""
        txt = str(value or "").lower().strip().replace("ё", "е")
        txt = re.sub(r"[^a-zа-я0-9\s]", " ", txt)
        txt = re.sub(r"\s+", " ", txt).strip()
        return txt

    @staticmethod
    def _read_sheet_with_detected_header(file_path: str, sheet_name: str) -> pd.DataFrame:
        """Читает лист ставок и пытается найти строку-заголовок, если стандартный header сломан."""
        raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        max_scan = min(12, len(raw))

        def _row_text(row_values) -> str:
            vals = [str(v).strip().lower() for v in row_values if pd.notna(v)]
            return " | ".join(vals)

        header_idx = None
        for i in range(max_scan):
            txt = _row_text(raw.iloc[i].tolist())
            has_fio = ("фио" in txt) or ("ф и о" in txt) or ("ф.и.о" in txt)
            has_rate_or_salary = ("ставк" in txt) or ("оклад" in txt) or ("фикс" in txt)
            if has_fio and has_rate_or_salary:
                header_idx = i
                break

        if header_idx is None:
            # Обычное чтение, если заголовок в первой строке
            return pd.read_excel(file_path, sheet_name=sheet_name)

        headers = []
        for j, v in enumerate(raw.iloc[header_idx].tolist()):
            if pd.isna(v):
                headers.append(f"col_{j}")
            else:
                headers.append(str(v).strip())

        data = raw.iloc[header_idx + 1 :].copy()
        data.columns = headers
        data = data.reset_index(drop=True)
        return data

    def load_rates_workbook(self, file_path: str) -> pd.DataFrame:
        """
        Загружает файл ставок механиков из всех листов.
        Каждому листу присваивается ДЦ = имя листа.
        """
        xls = pd.ExcelFile(file_path)
        rows = []

        for sheet in xls.sheet_names:
            df = self._read_sheet_with_detected_header(file_path, sheet)
            if df is None or df.empty:
                continue

            fio_col = self._find_column(df, ["фио", "ф и о", "ф.и.о", "фио сотрудника", "сотрудник", "механик", "name"])
            position_col = self._find_column(df, ["должность", "позиция", "position"])
            salary_col = self._find_column(df, ["оклад", "фикс", "фиксированныйоклад", "фиксироклад", "salary"])
            rate_col = self._find_column(df, ["ставка", "ставкакначислению", "сервисдопсервис", "rate"])

            # Если лист не табличный (например, просто список ДЦ), пропускаем.
            if not fio_col:
                continue

            for _, r in df.iterrows():
                fio = str(r[fio_col]).strip() if pd.notna(r[fio_col]) else ""
                if not fio or fio.lower() in {"nan", "none", "фио"}:
                    continue

                position = str(r[position_col]).strip() if position_col and pd.notna(r[position_col]) else ""
                salary = self._to_float(r[salary_col]) if salary_col else 0.0
                rate = self._to_float(r[rate_col]) if rate_col else 0.0

                rows.append(
                    {
                        "ДЦ": sheet,
                        "ФИО": fio,
                        "Должность": position,
                        "Оклад": salary,
                        "Ставка": rate,
                    }
                )

        if not rows:
            return pd.DataFrame(columns=["ДЦ", "ФИО", "Должность", "Оклад", "Ставка"])

        out_df = pd.DataFrame(rows)
        out_df["_dc_norm"] = out_df["ДЦ"].apply(self.normalize_dc)
        out_df["_fio_norm"] = out_df["ФИО"].apply(self.normalize_fio)
        out_df = out_df.drop_duplicates(subset=["_dc_norm", "_fio_norm"], keep="first")
        return out_df.reset_index(drop=True)

    @staticmethod
    def _best_fio_match(source_fio: str, target_norm_map: Dict[str, str]) -> tuple[str, float]:
        """Возвращает лучшее совпадение ФИО и confidence [0..1]."""
        src_norm = MechanicsCalculator.normalize_fio(source_fio)
        if not src_norm or not target_norm_map:
            return "", 0.0

        # 1) Точное совпадение после нормализации
        for original, norm in target_norm_map.items():
            if norm == src_norm:
                return original, 1.0

        # 2) Fuzzy best match
        best_fio = ""
        best_score = 0.0
        for original, norm in target_norm_map.items():
            score = MechanicsCalculator._fio_similarity(src_norm, norm)
            if score > best_score:
                best_score = score
                best_fio = original

        return best_fio, best_score

    def apply_rates_with_fuzzy_matching(
        self,
        summary_df: pd.DataFrame,
        rates_df: pd.DataFrame,
        fuzzy_threshold: float = 0.78,
    ) -> tuple[pd.DataFrame, pd.DataFrame]:
        """
        Сопоставляет итог механиков с файлом ставок по ФИО.

        Возвращает:
        - merged_df: итог с колонками сопоставления/статуса
        - mismatches_df: проблемные строки (нет совпадения или слабое fuzzy)
        """
        if summary_df is None or summary_df.empty:
            return pd.DataFrame(), pd.DataFrame()
        if rates_df is None or rates_df.empty:
            return summary_df.copy(), summary_df.copy()

        dc_col = self._find_column(rates_df, ["дц", "город", "дилерский центр", "dealer center"])
        fio_col = self._find_column(rates_df, ["фио", "механик", "сотрудник", "name"])
        position_col = self._find_column(rates_df, ["должность", "позиция", "position"])
        salary_col = self._find_column(rates_df, ["оклад", "фикс", "salary"])
        rate_col = self._find_column(rates_df, ["ставка", "rate"])

        if not fio_col:
            raise ValueError("В файле ставок не найдена колонка ФИО")

        rates_local = rates_df.copy()
        rates_local[fio_col] = rates_local[fio_col].astype(str).str.strip()
        rates_local = rates_local[rates_local[fio_col] != ""]
        if dc_col:
            rates_local[dc_col] = rates_local[dc_col].astype(str).str.strip()
            rates_local = rates_local.drop_duplicates(subset=[dc_col, fio_col], keep="first")
        else:
            rates_local = rates_local.drop_duplicates(subset=[fio_col], keep="first")

        target_norm_map: Dict[str, str] = {
            fio: self.normalize_fio(fio) for fio in rates_local[fio_col].tolist()
        }

        enriched_rows = []
        for _, row in summary_df.iterrows():
            src_fio = str(row.get("ФИО", "")).strip()
            src_dc = str(row.get("ДЦ", "")).strip()

            # Сначала пытаемся матчить в рамках того же ДЦ.
            rates_scope = rates_local
            if dc_col and src_dc:
                src_dc_norm = self.normalize_dc(src_dc)
                same_dc = rates_local[rates_local[dc_col].apply(self.normalize_dc) == src_dc_norm]
                if not same_dc.empty:
                    rates_scope = same_dc

            target_norm_map_scope: Dict[str, str] = {
                fio: self.normalize_fio(fio) for fio in rates_scope[fio_col].tolist()
            }
            best_fio, score = self._best_fio_match(src_fio, target_norm_map_scope)

            # Если в рамках ДЦ не нашли — fallback по всему файлу.
            if score < fuzzy_threshold and len(target_norm_map_scope) != len(target_norm_map):
                best_fio_global, score_global = self._best_fio_match(src_fio, target_norm_map)
                if score_global > score:
                    best_fio, score = best_fio_global, score_global
                    rates_scope = rates_local

            status = "unmatched"
            if score == 1.0:
                status = "exact"
            elif score >= fuzzy_threshold:
                status = "fuzzy"

            out = dict(row)
            out["ФИО в ставках"] = best_fio if status in ("exact", "fuzzy") else ""
            out["Совпадение"] = round(score, 3)
            out["Статус сопоставления"] = status

            if status in ("exact", "fuzzy"):
                rate_row = rates_scope[rates_scope[fio_col] == best_fio].iloc[0]
                out["ДЦ в ставках"] = str(rate_row[dc_col]).strip() if dc_col and pd.notna(rate_row[dc_col]) else ""
                out["Должность (ставки)"] = str(rate_row[position_col]).strip() if position_col else ""
                out["Источник ставки"] = f"Файл ставок ({status})" # Добавлено для диагностики
                try:
                    out["Оклад (ставки)"] = float(rate_row[salary_col]) if salary_col and pd.notna(rate_row[salary_col]) else 0.0
                except Exception:
                    out["Оклад (ставки)"] = 0.0
                try:
                    out["Ставка (ставки)"] = float(rate_row[rate_col]) if rate_col and pd.notna(rate_row[rate_col]) else 0.0
                except Exception:
                    out["Ставка (ставки)"] = 0.0
            else:
                out["ДЦ в ставках"] = ""
                out["Должность (ставки)"] = ""
                out["Оклад (ставки)"] = 0.0
                out["Ставка (ставки)"] = 0.0

            enriched_rows.append(out)

        merged_df = pd.DataFrame(enriched_rows)
        mismatches_df = merged_df[merged_df["Статус сопоставления"] == "unmatched"].copy()
        return merged_df, mismatches_df
