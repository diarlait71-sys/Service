"""
==========================================================================
АНАЛИЗ ПЛАН/ФАКТ РАСХОДА ЛЕКАРСТВЕННЫХ СРЕДСТВ — 2026 год
==========================================================================
Источник: Монит 28.04.2026) 2 — копия.xlsx, лист «ЛС»
Дата расчёта: 04.05.2026

Структура колонок (row 1 = настоящий заголовок, данные с row 3):
  col[0]   № п/п
  col[6]   Тип (ЛС/МИ/РРМ)
  col[8]   Наименование ЛС
  col[10]  Единица измерения
  col[11]  Утверждённый бюджет, кол-во (ПЛАН)
  col[12]  Цена за ед., тенге
  col[13]  Сумма план, тенге
  col[14]  Заявка отделений общая
  col[48]  Уточнённый бюджет, кол-во
  col[52]  Расход за 2025 год
  col[55]  Остаток общий на 01.01.2026
  col[58]  Остаток общий на 31.03.2026
  col[59]  Расход за I-квартал 2026 (ФАКТ)
  col[60]  Среднемесячный расход I-кв. 2026
  col[61]  Прогноз на 14 мес
  col[86]  Приход Q1 (кол-во)
  col[91]  Приход Q2 (кол-во, апрель+)
  col[103] Приход 2026 всего (кол-во)
  col[105] Остаток по договору (кол-во)
==========================================================================
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import warnings
warnings.filterwarnings('ignore')

# ── ПАРАМЕТРЫ ─────────────────────────────────────────────────────────────
SRC = r'C:\Users\D.Muldabaev\Desktop\Приложения для сервиса\Монит 28.04.2026) 2 — копия.xlsx'
SHEET = 'ЛС'
OUT  = r'C:\Users\D.Muldabaev\Desktop\Приложения для сервиса\Анализ_ПланФакт_ЛС_2026.xlsx'

# Текущее положение дел (04.05.2026)
MONTHS_ELAPSED = 4          # Прошло месяцев (янв–апр)
MONTHS_REMAINING = 8        # Осталось (май–дек)
MONTHS_TOTAL = 12

# ── ЗАГРУЗКА ДАННЫХ ────────────────────────────────────────────────────────
print("Загружаю данные...")
# Данные начинаются с 3-й строки (skiprows=2), нет заголовков
df = pd.read_excel(SRC, sheet_name=SHEET, header=None, skiprows=2, engine='openpyxl')

# Переименуем нужные колонки
COL_MAP = {
    0:  'num',
    6:  'type',
    8:  'name',
    10: 'unit',
    11: 'plan_qty',
    12: 'price',
    13: 'plan_sum',
    14: 'dept_request',
    48: 'plan_adj_qty',
    52: 'spend_2025',
    55: 'stock_01jan',
    58: 'stock_31mar',
    59: 'fact_q1',
    60: 'avg_monthly_q1',
    61: 'forecast_14m',
    83: 'jan', 84: 'feb', 85: 'mar',
    86: 'income_q1_qty',
    88: 'apr', 89: 'may_col', 90: 'jun',
    91: 'income_q2_qty',
    103: 'income_2026_qty',
    105: 'contract_rest_qty',
}

df = df.rename(columns=COL_MAP)

# Убираем служебные строки (итоги, пустые, нечисловые num)
df['num'] = pd.to_numeric(df['num'], errors='coerce')
df = df[df['num'].notna()].copy()
df = df[df['name'].notna() & (df['name'].astype(str).str.strip() != '')].copy()
print(f"  Строк с данными: {len(df)}")

# Числовые преобразования
NUM_COLS = ['plan_qty', 'price', 'plan_sum', 'dept_request', 'plan_adj_qty',
            'spend_2025', 'stock_01jan', 'stock_31mar', 'fact_q1',
            'avg_monthly_q1', 'forecast_14m',
            'jan', 'feb', 'mar', 'income_q1_qty',
            'apr', 'may_col', 'jun', 'income_q2_qty',
            'income_2026_qty', 'contract_rest_qty']
for c in NUM_COLS:
    if c in df.columns:
        df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)

# ── РАСЧЁТЫ ────────────────────────────────────────────────────────────────
print("Выполняю расчёты...")

# 1. Среднемесячный расход (если не заполнен — считаем сами из Q1)
df['avg_monthly_calc'] = np.where(
    df['avg_monthly_q1'] > 0,
    df['avg_monthly_q1'],
    df['fact_q1'] / 3   # Q1 = 3 месяца
)

# 2. Прогноз расхода до конца года
#    Уже прошло 4 месяца (Q1=факт, апрель≈1 месяц по среднему)
#    Оцениваем фактический расход за апрель как avg_monthly (нет точных данных)
df['fact_apr_est'] = df['avg_monthly_calc']          # оценка апреля
df['fact_ytd'] = df['fact_q1'] + df['fact_apr_est']  # факт янв-апр (оценка)

df['forecast_remain'] = df['avg_monthly_calc'] * MONTHS_REMAINING  # май-дек
df['forecast_year_total'] = df['fact_ytd'] + df['forecast_remain']  # прогноз год

# 3. ПЛАН (берём уточнённый, если есть; иначе утверждённый)
df['plan_final'] = np.where(df['plan_adj_qty'] > 0, df['plan_adj_qty'], df['plan_qty'])

# 4. Отклонение план vs прогноз
df['deviation_qty'] = df['forecast_year_total'] - df['plan_final']
df['deviation_pct'] = np.where(
    df['plan_final'] > 0,
    (df['deviation_qty'] / df['plan_final'] * 100).round(1),
    np.nan
)

# 5. Выполнение плана на текущий момент (план на 4 мес = plan/12*4)
df['plan_4m'] = df['plan_final'] / MONTHS_TOTAL * MONTHS_ELAPSED
df['exec_pct'] = np.where(
    df['plan_4m'] > 0,
    (df['fact_ytd'] / df['plan_4m'] * 100).round(1),
    np.nan
)

# 6. Потребность к закупке до конца года
#    Нужно: forecast_remain; Есть: stock_31mar + (приход Q2 к этому моменту)
#    В базе приход Q2 = данные апреля (income_q2_qty)
df['available_now'] = df['stock_31mar'] + df['income_q2_qty']  # оценка запаса на 04.05
df['need_to_buy'] = np.maximum(0, df['forecast_remain'] - df['available_now'])
df['need_to_buy_sum'] = df['need_to_buy'] * df['price']

# 7. ВЫЯВЛЕНИЕ «СКРЫТОГО» РАСХОДА
#    Скрытый расход = есть фактический расход (fact_q1 > 0), НО:
#    - заявка отделений = 0 (dept_request == 0), или
#    - разница (заявка >> приход): dept_request > 1.5 * income_q1_qty
df['hidden_flag'] = 'Нет'
mask_no_request = (df['fact_q1'] > 0) & (df['dept_request'] == 0)
mask_excess_req  = (df['dept_request'] > 0) & (df['income_q1_qty'] > 0) & \
                   (df['dept_request'] > 1.5 * df['income_q1_qty'])
mask_no_income   = (df['fact_q1'] > 0) & (df['income_q1_qty'] == 0) & \
                   (df['contract_rest_qty'] == 0)
df.loc[mask_no_request, 'hidden_flag'] = '⚠ Расход без заявки'
df.loc[mask_excess_req,  'hidden_flag'] = '⚠ Заявка >> Поступление'
df.loc[mask_no_income,   'hidden_flag'] = '⚠ Расход без поступления'

# 8. СТАТУС ПОЗИЦИИ
def status(row):
    if row['forecast_year_total'] == 0 and row['plan_final'] == 0:
        return '—'
    if row['need_to_buy'] > 0 and row['need_to_buy'] > row['plan_final'] * 0.3:
        return '🔴 НЕХВАТКА'
    if row['deviation_pct'] is not None and not np.isnan(row['deviation_pct']):
        if row['deviation_pct'] > 20:
            return '🟠 Перерасход'
        if row['deviation_pct'] < -20:
            return '🟢 Экономия'
    return '✅ В норме'

df['status'] = df.apply(status, axis=1)

# ── ФОРМИРОВАНИЕ ОТЧЁТА (Excel) ────────────────────────────────────────────
print("Формирую Excel-отчёт...")

wb = Workbook()

# --- Цвета и стили -------------------------
RED    = 'FFCC0000'
ORANGE = 'FFFF6600'
GREEN  = 'FF00AA44'
YELLOW = 'FFFFF2CC'
LIGHT_BLUE  = 'FFDCE6F1'
LIGHT_GREEN = 'FFE2EFDA'
LIGHT_RED   = 'FFFCE4D6'
HEADER_FILL = PatternFill("solid", fgColor="FF1F3864")
SUB_FILL    = PatternFill("solid", fgColor="FF2E75B6")
ROW_ALT     = PatternFill("solid", fgColor="FFF2F7FF")
THIN = Side(style='thin', color='FFBFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WHITE_FONT  = Font(name='Calibri', color='FFFFFFFF', bold=True, size=11)
HEADER_FONT = Font(name='Calibri', color='FFFFFFFF', bold=True, size=10)
NORMAL_FONT = Font(name='Calibri', size=10)
BOLD_FONT   = Font(name='Calibri', bold=True, size=10)

def hdr_cell(ws, row, col, val, fill=None, font=None, align='center', wrap=True):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill or HEADER_FILL
    c.font = font or HEADER_FONT
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border = BORDER
    return c

def data_cell(ws, row, col, val, fmt=None, fill=None, font=None, align='center'):
    c = ws.cell(row=row, column=col, value=val)
    if fmt:   c.number_format = fmt
    if fill:  c.fill = fill
    c.font   = font or NORMAL_FONT
    c.border = BORDER
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=False)
    return c

# ═══════════════════════════════════════════════════════════════════════════
# ЛИСТ 1: СВОДНЫЙ АНАЛИЗ
# ═══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Сводный анализ'
ws1.freeze_panes = 'C4'
ws1.sheet_view.showGridLines = False

# Строка 1: Заголовок отчёта
ws1.merge_cells('A1:S1')
t = ws1['A1']
t.value = '📊 АНАЛИЗ ПЛАН / ФАКТ РАСХОДА ЛЕКАРСТВЕННЫХ СРЕДСТВ — 2026 год'
t.font = Font(name='Calibri', bold=True, size=14, color='FF1F3864')
t.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 28

ws1.merge_cells('A2:S2')
t2 = ws1['A2']
t2.value = (f'По состоянию на 04.05.2026 | Прошло месяцев: {MONTHS_ELAPSED} '
            f'| Осталось: {MONTHS_REMAINING} | Источник: Монит 28.04.2026')
t2.font = Font(name='Calibri', size=10, italic=True, color='FF444444')
t2.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[2].height = 18

# Строка 3: Заголовки колонок
COLS = [
    ('№', 5), ('Наименование ЛС', 42), ('Ед.', 7),
    ('ПЛАН\n(год, шт)', 12), ('Факт Q1\n(шт)', 12),
    ('Факт янв-апр\n(оценка)', 12), ('Ср.мес.\nрасход', 10),
    ('Прогноз\nдо конца года', 14), ('Прогноз\nгод итого', 14),
    ('Откл. от\nплана, шт', 12), ('Откл.\n%', 8),
    ('Исп. плана\n4 мес, %', 11), ('Заявка\nотделений', 12),
    ('Приход\nQ1, шт', 11), ('Приход\nQ2 (апр), шт', 12),
    ('Запас\nтекущий', 11), ('Докупить\nдо конца года', 14),
    ('Скрытый\nрасход', 20), ('Статус', 14),
]
for i, (hdr, w) in enumerate(COLS, start=1):
    hdr_cell(ws1, 3, i, hdr)
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.row_dimensions[3].height = 42

# Данные
df_sorted = df.sort_values('name').reset_index(drop=True)
for r_idx, (_, row) in enumerate(df_sorted.iterrows(), start=4):
    alt = (r_idx % 2 == 0)
    bg  = ROW_ALT if alt else None

    # Окраска статуса
    stat = str(row['status'])
    if '🔴' in stat:
        row_fill = PatternFill("solid", fgColor=LIGHT_RED)
    elif '🟠' in stat:
        row_fill = PatternFill("solid", fgColor="FFFCE4D6")
    elif '🟢' in stat:
        row_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    elif alt:
        row_fill = ROW_ALT
    else:
        row_fill = None

    def dc(col, val, fmt=None, align='center'):
        data_cell(ws1, r_idx, col, val, fmt=fmt, fill=row_fill, align=align)

    dc(1,  int(row['num']) if row['num'] else '')
    dc(2,  str(row['name']).strip(), align='left')
    dc(3,  str(row['unit']).strip() if pd.notna(row['unit']) else '')
    dc(4,  row['plan_final'] or '',           fmt='#,##0')
    dc(5,  row['fact_q1'] or '',              fmt='#,##0')
    dc(6,  round(row['fact_ytd'], 1) or '',   fmt='#,##0.0')
    dc(7,  round(row['avg_monthly_calc'], 1) if row['avg_monthly_calc'] > 0 else '', fmt='#,##0.0')
    dc(8,  round(row['forecast_remain'], 1) if row['forecast_remain'] > 0 else '', fmt='#,##0.0')
    dc(9,  round(row['forecast_year_total'], 1) if row['forecast_year_total'] > 0 else '', fmt='#,##0.0')
    dc(10, round(row['deviation_qty'], 1) if row['deviation_qty'] != 0 else '', fmt='+#,##0.0;-#,##0.0')
    dc(11, row['deviation_pct'] if pd.notna(row['deviation_pct']) else '', fmt='+0.0%;-0.0%' if False else '0.0')
    dc(12, row['exec_pct'] if pd.notna(row['exec_pct']) else '',           fmt='0.0')
    dc(13, row['dept_request'] or '',         fmt='#,##0')
    dc(14, row['income_q1_qty'] or '',        fmt='#,##0')
    dc(15, row['income_q2_qty'] or '',        fmt='#,##0')
    dc(16, round(row['available_now'], 1) or '', fmt='#,##0.0')
    dc(17, round(row['need_to_buy'], 1) if row['need_to_buy'] > 0 else '', fmt='#,##0.0')
    dc(18, row['hidden_flag'], align='left')
    dc(19, row['status'], align='center')

total_rows = len(df_sorted)
print(f"  Заполнено строк: {total_rows}")

# ═══════════════════════════════════════════════════════════════════════════
# ЛИСТ 2: КРИТИЧЕСКИЕ ПОЗИЦИИ (нехватка + перерасход)
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('🔴 Нехватка и перерасход')
ws2.sheet_view.showGridLines = False

ws2.merge_cells('A1:M1')
c = ws2['A1']
c.value = '🔴 КРИТИЧЕСКИЕ ПОЗИЦИИ — нехватка и значительный перерасход'
c.font = Font(name='Calibri', bold=True, size=13, color='FFCC0000')
c.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 26

COLS2 = [
    ('№', 5), ('Наименование ЛС', 45), ('Ед.', 7),
    ('ПЛАН год', 11), ('Факт Q1', 11), ('Прогноз год', 12),
    ('Откл. шт', 12), ('Откл. %', 9), ('Заявка', 10),
    ('Приход Q1', 10), ('Запас', 10),
    ('ДОКУПИТЬ, шт', 14), ('Статус', 15),
]
for i, (h, w) in enumerate(COLS2, start=1):
    hdr_cell(ws2, 2, i, h, fill=PatternFill("solid", fgColor="FF922B21"))
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[2].height = 36

df_crit = df_sorted[
    df_sorted['status'].str.contains('🔴|🟠', na=False)
].copy()

r_idx = 3
for _, row in df_crit.iterrows():
    fill_crit = PatternFill("solid", fgColor=LIGHT_RED)
    def dc2(col, val, fmt=None, align='center'):
        data_cell(ws2, r_idx, col, val, fmt=fmt, fill=fill_crit, align=align)
    dc2(1,  int(row['num']) if row['num'] else '')
    dc2(2,  str(row['name']).strip(), align='left')
    dc2(3,  str(row['unit']).strip() if pd.notna(row['unit']) else '')
    dc2(4,  row['plan_final'] or '', fmt='#,##0')
    dc2(5,  row['fact_q1'] or '', fmt='#,##0')
    dc2(6,  round(row['forecast_year_total'], 1), fmt='#,##0.0')
    dc2(7,  round(row['deviation_qty'], 1), fmt='+#,##0.0;-#,##0.0')
    dc2(8,  row['deviation_pct'] if pd.notna(row['deviation_pct']) else '', fmt='0.0')
    dc2(9,  row['dept_request'] or '', fmt='#,##0')
    dc2(10, row['income_q1_qty'] or '', fmt='#,##0')
    dc2(11, round(row['available_now'], 1), fmt='#,##0.0')
    dc2(12, round(row['need_to_buy'], 1) if row['need_to_buy'] > 0 else 0, fmt='#,##0.0')
    dc2(13, row['status'])
    r_idx += 1

print(f"  Критических позиций: {len(df_crit)}")

# Итог по докупке
ws2.cell(r_idx, 1, 'ИТОГО к докупке').font = BOLD_FONT
ws2.cell(r_idx, 12, df_crit['need_to_buy'].sum().round(1))
ws2.cell(r_idx, 12).number_format = '#,##0.0'
ws2.cell(r_idx, 12).font = Font(bold=True, color='FFCC0000')

# ═══════════════════════════════════════════════════════════════════════════
# ЛИСТ 3: СКРЫТЫЙ РАСХОД
# ═══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('⚠ Скрытый расход')
ws3.sheet_view.showGridLines = False

ws3.merge_cells('A1:J1')
c = ws3['A1']
c.value = '⚠ ПОЗИЦИИ СО "СКРЫТЫМ" РАСХОДОМ (расход есть, но заявки/списания не отражены)'
c.font = Font(name='Calibri', bold=True, size=12, color='FFCC6600')
c.alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 26

COLS3 = [
    ('№', 5), ('Наименование ЛС', 45), ('Ед.', 7),
    ('ПЛАН год', 11), ('Факт Q1\n(расход)', 11),
    ('Заявка\nотделений', 12), ('Приход Q1', 11),
    ('Остаток\n01.01', 11), ('Остаток\n31.03', 11),
    ('Флаг скрытого\nрасхода', 28),
]
for i, (h, w) in enumerate(COLS3, start=1):
    hdr_cell(ws3, 2, i, h, fill=PatternFill("solid", fgColor="FFCC6600"))
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.row_dimensions[2].height = 42

df_hidden = df_sorted[df_sorted['hidden_flag'] != 'Нет'].copy()
HIDDEN_FILL = PatternFill("solid", fgColor="FFFFF2CC")
r_idx = 3
for _, row in df_hidden.iterrows():
    def dc3(col, val, fmt=None, align='center'):
        data_cell(ws3, r_idx, col, val, fmt=fmt, fill=HIDDEN_FILL, align=align)
    dc3(1,  int(row['num']) if row['num'] else '')
    dc3(2,  str(row['name']).strip(), align='left')
    dc3(3,  str(row['unit']).strip() if pd.notna(row['unit']) else '')
    dc3(4,  row['plan_final'] or '', fmt='#,##0')
    dc3(5,  row['fact_q1'] or '', fmt='#,##0')
    dc3(6,  row['dept_request'] or '', fmt='#,##0')
    dc3(7,  row['income_q1_qty'] or '', fmt='#,##0')
    dc3(8,  row['stock_01jan'] or '', fmt='#,##0.0')
    dc3(9,  row['stock_31mar'] or '', fmt='#,##0.0')
    dc3(10, row['hidden_flag'], align='left')
    r_idx += 1

print(f"  Позиций со скрытым расходом: {len(df_hidden)}")

# ═══════════════════════════════════════════════════════════════════════════
# ЛИСТ 4: ЧТО И СКОЛЬКО ДОКУПИТЬ
# ═══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('📦 К закупке')
ws4.sheet_view.showGridLines = False

ws4.merge_cells('A1:H1')
c = ws4['A1']
c.value = '📦 ПЕРЕЧЕНЬ ЛЕКАРСТВЕННЫХ СРЕДСТВ К ДОПОЛНИТЕЛЬНОЙ ЗАКУПКЕ (до конца 2026 года)'
c.font = Font(name='Calibri', bold=True, size=13, color='FF1F3864')
c.alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 26

COLS4 = [
    ('№', 5), ('Наименование ЛС', 48), ('Ед.', 7),
    ('Текущий запас\n(оценка)', 14), ('Прогноз расхода\nмай-дек', 15),
    ('К ЗАКУПКЕ, шт', 14), ('Цена за ед.,\nтенге', 14),
    ('Сумма к закупке,\nтенге', 18),
]
for i, (h, w) in enumerate(COLS4, start=1):
    hdr_cell(ws4, 2, i, h)
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.row_dimensions[2].height = 42

df_buy = df_sorted[df_sorted['need_to_buy'] > 0].sort_values('need_to_buy_sum', ascending=False).copy()
BUY_FILL = PatternFill("solid", fgColor="FFFCE4D6")
r_idx = 3
for _, row in df_buy.iterrows():
    alt = (r_idx % 2 == 0)
    fill = BUY_FILL if row['need_to_buy'] > 100 else (ROW_ALT if alt else None)
    def dc4(col, val, fmt=None, align='center'):
        data_cell(ws4, r_idx, col, val, fmt=fmt, fill=fill, align=align)
    dc4(1, int(row['num']) if row['num'] else '')
    dc4(2, str(row['name']).strip(), align='left')
    dc4(3, str(row['unit']).strip() if pd.notna(row['unit']) else '')
    dc4(4, round(row['available_now'], 1), fmt='#,##0.0')
    dc4(5, round(row['forecast_remain'], 1), fmt='#,##0.0')
    dc4(6, round(row['need_to_buy'], 1), fmt='#,##0.0')
    dc4(7, row['price'] if row['price'] > 0 else '', fmt='#,##0.00')
    dc4(8, round(row['need_to_buy_sum'], 2) if row['need_to_buy_sum'] > 0 else '', fmt='#,##0.00')
    r_idx += 1

# Итоговая строка
total_sum = df_buy['need_to_buy_sum'].sum()
for col in range(1, 9):
    c = ws4.cell(r_idx, col)
    c.border = BORDER
    c.fill = PatternFill("solid", fgColor="FF1F3864")
    c.font = WHITE_FONT
ws4.cell(r_idx, 2, 'ИТОГО К ЗАКУПКЕ')
ws4.cell(r_idx, 2).font = WHITE_FONT
ws4.cell(r_idx, 2).fill = PatternFill("solid", fgColor="FF1F3864")
ws4.cell(r_idx, 2).border = BORDER
ws4.cell(r_idx, 8, round(total_sum, 2))
ws4.cell(r_idx, 8).number_format = '#,##0.00'
ws4.cell(r_idx, 8).font = WHITE_FONT
ws4.cell(r_idx, 8).fill = PatternFill("solid", fgColor="FF1F3864")
ws4.cell(r_idx, 8).border = BORDER
print(f"  Позиций к закупке: {len(df_buy)}, сумма: {total_sum:,.0f} тенге")

# ═══════════════════════════════════════════════════════════════════════════
# ЛИСТ 5: СВОДНЫЕ ПОКАЗАТЕЛИ (дашборд)
# ═══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('📈 Дашборд')
ws5.sheet_view.showGridLines = False
ws5.column_dimensions['A'].width = 35
ws5.column_dimensions['B'].width = 18
ws5.column_dimensions['C'].width = 18
ws5.column_dimensions['D'].width = 35

def kpi_row(ws, row, label, value, fmt='#,##0', note=''):
    c1 = ws.cell(row, 1, label)
    c1.font = Font(name='Calibri', bold=True, size=11)
    c1.alignment = Alignment(horizontal='left', vertical='center')
    c2 = ws.cell(row, 2, value)
    c2.number_format = fmt
    c2.font = Font(name='Calibri', bold=True, size=12, color='FF1F3864')
    c2.alignment = Alignment(horizontal='right', vertical='center')
    if note:
        c3 = ws.cell(row, 4, note)
        c3.font = Font(name='Calibri', size=10, italic=True, color='FF666666')
    ws.row_dimensions[row].height = 22

ws5.merge_cells('A1:D1')
t = ws5['A1']
t.value = '📈 СВОДНЫЕ ПОКАЗАТЕЛИ — МОНИТОРИНГ ЛС 2026'
t.font = Font(name='Calibri', bold=True, size=15, color='FFFFFFFF')
t.fill = PatternFill("solid", fgColor="FF1F3864")
t.alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[1].height = 32

ws5.cell(2, 1, '').fill = PatternFill("solid", fgColor="FF2E75B6")
r = 3
kpi_row(ws5, r, '📅 Дата расчёта', '04.05.2026', fmt='@'); r+=1
kpi_row(ws5, r, '📅 Прошло месяцев (янв–апр)', MONTHS_ELAPSED); r+=1
kpi_row(ws5, r, '📅 Осталось месяцев (май–дек)', MONTHS_REMAINING); r+=1
r += 1

ws5.cell(r, 1, '--- ПОЗИЦИИ ---').font = Font(bold=True, color='FF1F3864'); r+=1
kpi_row(ws5, r, 'Всего позиций ЛС в мониторинге', len(df_sorted)); r+=1
kpi_row(ws5, r, 'Позиций с фактическим расходом Q1', int((df_sorted['fact_q1'] > 0).sum())); r+=1
kpi_row(ws5, r, 'Позиций БЕЗ расхода (не используются)', int((df_sorted['fact_q1'] == 0).sum())); r+=1
r += 1

ws5.cell(r, 1, '--- ПЛАН vs ФАКТ ---').font = Font(bold=True, color='FF1F3864'); r+=1
plan_total = df_sorted['plan_final'].sum()
fact_q1_total = df_sorted['fact_q1'].sum()
fact_ytd_total = df_sorted['fact_ytd'].sum()
forecast_total = df_sorted['forecast_year_total'].sum()
kpi_row(ws5, r, 'Суммарный план на год (шт)', plan_total); r+=1
kpi_row(ws5, r, 'Факт расход Q1 (шт)', fact_q1_total); r+=1
kpi_row(ws5, r, 'Факт янв–апр оценка (шт)', fact_ytd_total); r+=1
kpi_row(ws5, r, 'Прогноз расход за год (шт)', forecast_total); r+=1
kpi_row(ws5, r, 'Общее отклонение план vs прогноз (шт)', forecast_total - plan_total, fmt='+#,##0;-#,##0'); r+=1
r += 1

ws5.cell(r, 1, '--- СТАТУСЫ ---').font = Font(bold=True, color='FF1F3864'); r+=1
for stat_val, stat_name in [('🔴 НЕХВАТКА', '🔴 Нехватка (критично)'),
                              ('🟠 Перерасход', '🟠 Перерасход (>20%)'),
                              ('🟢 Экономия', '🟢 Экономия (<-20%)'),
                              ('✅ В норме', '✅ В норме')]:
    cnt = int(df_sorted['status'].str.contains(stat_val.split()[1], regex=False, na=False).sum())
    kpi_row(ws5, r, stat_name, cnt); r+=1
r += 1

ws5.cell(r, 1, '--- ЗАКУПКА ---').font = Font(bold=True, color='FF1F3864'); r+=1
kpi_row(ws5, r, 'Позиций требует дозакупки', len(df_buy)); r+=1
kpi_row(ws5, r, 'Итого к дозакупке (шт)', df_buy['need_to_buy'].sum()); r+=1
kpi_row(ws5, r, 'Итого к дозакупке (тенге)', total_sum, fmt='#,##0.00'); r+=1
r += 1

ws5.cell(r, 1, '--- СКРЫТЫЙ РАСХОД ---').font = Font(bold=True, color='FF1F3864'); r+=1
kpi_row(ws5, r, 'Позиций со скрытым расходом', len(df_hidden)); r+=1

# ── СОХРАНЯЕМ ──────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"\n✅ Отчёт сохранён: {OUT}")
print(f"\n=== КРАТКИЕ ИТОГИ ===")
print(f"  Всего позиций:         {len(df_sorted)}")
print(f"  Критических:           {len(df_crit)}")
print(f"  Скрытый расход:        {len(df_hidden)}")
print(f"  К закупке позиций:     {len(df_buy)}")
print(f"  Сумма к закупке:       {total_sum:,.0f} тенге")
print(f"  Факт Q1 (шт):          {fact_q1_total:,.0f}")
print(f"  Прогноз год (шт):      {forecast_total:,.0f}")
print(f"  План год (шт):         {plan_total:,.0f}")
