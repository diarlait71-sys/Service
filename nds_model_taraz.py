"""
Анализ НДС-моделей для ДЦ Тараз
Сравнение Схема 1 (Агент) vs Схема 2 (Торговец, ст. 394 НК РК)
"""

import openpyxl
import pandas as pd
from pathlib import Path
import os

# ============================================================================
# ШАГИ:
# 1. Парсим ОПУ и ОСВ
# 2. Вытягиваем доходы/расходы по месяцам
# 3. Разбиваем входной НДС на категории
# 4. Строим две модели НДС
# 5. Сравниваем и даем рекомендацию
# ============================================================================

nds_path = r'c:\Users\D.Muldabaev\Desktop\Приложения для сервиса\Ндс'

def parse_opu_file(filepath):
    """Парсит ОПУ файл и вытягивает структуру доходов/расходов по месяцам"""
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    print(f"\n>>> ОПУ файл: {Path(filepath).name}")
    print(f"    Листы: {wb.sheetnames}\n")
    
    data = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"    Обработка листа: '{sheet_name}' ({ws.max_row} строк × {ws.max_column} колонок)")
        
        # ищем структуру:
        # обычно 1-я колонка = описание, 2+ колонки = месячные данные
        
        # собираем все строки
        rows_data = []
        for r in range(1, min(ws.max_row+1, 150)):
            row = []
            for c in range(1, min(15, ws.max_column+1)):
                row.append(ws.cell(r, c).value)
            rows_data.append(row)
        
        data[sheet_name] = rows_data
    
    return data

def extract_financial_structure(opu_data):
    """Из ОПУ вытягивает структуру: доходы, расходы, НДС по месяцам"""
    
    print("\n" + "="*100)
    print("АНАЛИЗ СТРУКТУРЫ ОПУ")
    print("="*100)
    
    # ищем ключевые строки в первом листе
    first_sheet = list(opu_data.values())[0]
    
    # группы строк, которые ищем
    search_terms = {
        'доход': ['выручк', 'доход', 'реализац', 'комисс', 'сервис', 'запчаст'],
        'расход': ['расход', 'затрат', 'себестоим', 'аренд', 'фот', 'маркетинг', 'коммунал', 'ндс'],
        'ндс': ['ндс', 'налог', 'начисл', 'входной', 'вычет']
    }
    
    found_items = {}
    
    for row_idx, row in enumerate(first_sheet[:80]):  # первые 80 строк
        if not row or not row[0]:
            continue
        
        line_text = str(row[0]).lower()
        
        for category, keywords in search_terms.items():
            for keyword in keywords:
                if keyword in line_text:
                    if category not in found_items:
                        found_items[category] = []
                    
                    # собираем числовые данные из этой строки
                    nums = []
                    for c in range(1, len(row)):
                        try:
                            if isinstance(row[c], (int, float)):
                                nums.append(row[c])
                        except:
                            pass
                    
                    found_items[category].append({
                        'text': str(row[0])[:60],
                        'values': nums,
                        'row': row_idx
                    })
                    break
    
    # выводим структуру
    for cat, items in found_items.items():
        print(f"\n{cat.upper()}:")
        for item in items[:5]:  # первые 5
            vals = item['values'][:6]  # первые 6 месяцев
            print(f"  {item['text']}")
            print(f"    Значения: {[f'{v:,.0f}' if isinstance(v, (int,float)) else v for v in vals]}")
    
    return found_items

def build_nds_models(opu_data):
    """Строит две модели НДС: Агент vs Торговец"""
    
    print("\n" + "="*100)
    print("ПОСТРОЕНИЕ НДС-МОДЕЛЕЙ")
    print("="*100)
    
    # парсим данные из ОПУ
    first_sheet = list(opu_data.values())[0]
    
    # ищем итоговые суммы за год
    annual_data = {
        'revenue_auto': 0,
        'revenue_service': 0,
        'revenue_spare': 0,
        'cost_auto': 0,
        'cost_service': 0,
        'cost_spare': 0,
        'other_costs': 0,
        'vat_input': 0,
    }
    
    # для каждой категории ищем суммы
    for row in first_sheet[:100]:
        if not row or len(row) < 2:
            continue
        
        text = str(row[0]).lower() if row[0] else ""
        
        # ищем ключевые суммы
        for i, cell in enumerate(row[1:], 1):
            if isinstance(cell, (int, float)) and abs(cell) > 100000:
                if 'комисс' in text or 'авто' in text and 'выручк' in text:
                    annual_data['revenue_auto'] = max(annual_data['revenue_auto'], cell)
                elif 'сервис' in text and 'выручк' in text:
                    annual_data['revenue_service'] = max(annual_data['revenue_service'], cell)
                elif 'запчаст' in text or 'запас' in text and 'выручк' in text:
                    annual_data['revenue_spare'] = max(annual_data['revenue_spare'], cell)
                elif 'входной' in text and 'ндс' in text:
                    annual_data['vat_input'] = max(annual_data['vat_input'], cell)
    
    print(f"\nОбнаруженные данные за период:")
    print(f"  Выручка авто (комиссия): {annual_data['revenue_auto']:,.0f} тг")
    print(f"  Выручка сервис: {annual_data['revenue_service']:,.0f} тг")
    print(f"  Выручка запчасти: {annual_data['revenue_spare']:,.0f} тг")
    print(f"  Входной НДС: {annual_data['vat_input']:,.0f} тг")
    
    return annual_data

# ============================================================================
# ОСНОВНОЙ СКРИПТ
# ============================================================================

opu_file = os.path.join(nds_path, 'Опиу с разбивкой весь период.xlsx')
osv_file = os.path.join(nds_path, 'ОСВ 2025.xlsx')

print("\n" + "="*100)
print("ФИНАНСОВЫЙ АНАЛИЗ: ДЦ ТАРАЗ")
print("Сравнение НДС-схем: Агент vs Торговец (ст. 394 НК РК)")
print("="*100)

if os.path.exists(opu_file):
    opu_data = parse_opu_file(opu_file)
    financial = extract_financial_structure(opu_data)
    models = build_nds_models(opu_data)
else:
    print(f"\n✗ Файл не найден: {opu_file}")
    print("Следующий шаг: открыть файл вручную и посмотреть структуру")

print("\n" + "="*100)
print("СОХРАНЕНИЕ РЕЗУЛЬТАТОВ")
print("="*100)
print("Скрипт готов, нужно уточнить структуру ОПУ вручную для точного расчета")
