"""
Быстрый парсер ОСВ и ОПУ - создает сводный файл
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
from pathlib import Path

def safe_parse_file(filepath):
    """Безопасно парсит файл и возвращает структуру"""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        result = {
            'file': Path(filepath).name,
            'status': 'OK',
            'sheets': wb.sheetnames,
            'rows': ws.max_row,
            'cols': ws.max_column,
            'data': []
        }
        
        # собираем первые значимые строки
        for r in range(1, min(100, ws.max_row+1)):
            row_vals = []
            for c in range(1, min(8, ws.max_column+1)):
                row_vals.append(ws.cell(r, c).value)
            
            result['data'].append(row_vals)
        
        return result
    
    except Exception as e:
        return {'file': Path(filepath).name, 'status': f'ERROR: {str(e)[:100]}'}

# парсим файлы
nds_path = r'c:\Users\D.Muldabaev\Desktop\Приложения для сервиса\Ндс'

files_info = {
    'ОСВ 2025.xlsx': safe_parse_file(os.path.join(nds_path, 'ОСВ 2025.xlsx')),
    'ОСВ 2026 1 квартал.xlsx': safe_parse_file(os.path.join(nds_path, 'ОСВ 2026 1 квартал.xlsx')),
    'Опиу с разбивкой весь период.xlsx': safe_parse_file(os.path.join(nds_path, 'Опиу с разбивкой весь период.xlsx')),
}

# создаем отчет
out_wb = Workbook()
ws = out_wb.active
ws.title = 'Структура файлов'

# заголовки
headers = ['Файл', 'Статус', 'Листы', 'Строк', 'Колонок', 'Примечание']
for col, h in enumerate(headers, 1):
    cell = ws.cell(1, col, h)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

# данные
row = 2
for fname, info in files_info.items():
    ws.cell(row, 1, fname)
    ws.cell(row, 2, info.get('status', 'Unknown'))
    ws.cell(row, 3, ', '.join(info.get('sheets', [])))
    ws.cell(row, 4, info.get('rows', ''))
    ws.cell(row, 5, info.get('cols', ''))
    ws.cell(row, 6, 'Проверить содержимое' if info.get('status') == 'OK' else 'Ошибка')
    row += 1

# добавляем лист с рекомендациями
ws2 = out_wb.create_sheet('Рекомендации')
recommendations = """
АНАЛИЗ ДОБАВЛЕННЫХ ФАЙЛОВ:

✓ ФАЙЛЫ УСПЕШНО ДОБАВЛЕНЫ:
  1. ОСВ 2025.xlsx - оборотно-сальдовая ведомость за весь год
  2. ОСВ 2026 1 квартал.xlsx - ОСВ за Q1 2026 (для тренда)
  3. Опиу с разбивкой весь период.xlsx - Отчет о прибылях и убытках

ПРОВЕРКА ДАННЫХ:
1. Откройте каждый файл и убедитесь:
   - ОСВ 2025: есть ли выручка по авто, сервис, запчасти, входной НДС
   - ОПУ: расчеты по доходам, расходам, НДС
   
2. Сверьте цифры:
   - Сумма выручки в ОПУ = сумма оборотов в Форме 300 (декларация НДС)
   - Входной НДС в ОСВ должен совпадать с вычетом в Форме 300

СЛЕДУЮЩИЙ ШАГ:
- Уточните, по какому ДЦ эти данные (или это по целой группе)
- Если есть, выгрузите ЭСФ-реестр (счета-фактуры) для верификации
- Готов считать модель как только убедимся в консистентности данных
""".split('\n')

for i, line in enumerate(recommendations, 1):
    ws2.cell(i, 1, line)

out_path = os.path.join(nds_path, '_Проверка_структуры_файлов.xlsx')
out_wb.save(out_path)

print(f"Отчет сохранен: {out_path}")
print(f"\nЛучше если я могу просмотреть данные, мне нужны уточнения:")
print("  1. По какому ДЦ эти данные (какое ТОО)?")
print("  2. Есть ли ЭСФ-реестр (счета-фактуры) для сверки?")
print("  3. Лист 'Опиу с разбивкой весь период' - это за какой период?")
