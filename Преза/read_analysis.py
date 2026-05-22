import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

path = r"c:\Users\D.Muldabaev\Desktop\Приложения для сервиса\Преза\Маржа апрель — копия.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

# Читаем только аналитические листы
analysis_sheets = ['Анализ по мес', 'Анализ Киа', 'Анализ Шеви', 'Анализ Джак', 'Анализ Джетур', 'Анализ Шкода']

for sh in analysis_sheets:
    if sh not in wb.sheetnames:
        continue
    ws = wb[sh]
    print(f"\n=== {sh} ({ws.max_row}x{ws.max_column}) ===")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), values_only=True):
        # Фильтруем пустые строки
        non_none = [v for v in row if v is not None]
        if non_none:
            print(non_none)
