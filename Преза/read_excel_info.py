import openpyxl, sys

path = r"c:\Users\D.Muldabaev\Desktop\Приложения для сервиса\Преза\Маржа апрель — копия.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
print("Листы:", wb.sheetnames)
for sh in wb.sheetnames:
    ws = wb[sh]
    print(f"\n=== {sh} ({ws.max_row}x{ws.max_column}) ===")
    for row in ws.iter_rows(min_row=1, max_row=min(15, ws.max_row), values_only=True):
        print(row)
